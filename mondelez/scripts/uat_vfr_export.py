"""
UAT VFR (Vehicle Fill Rate) export — sinh Excel test pack với số SQL thật cho khách Mondelez confirm.

Window: 2026-05-18 → 2026-05-22 (5 ngày) — cùng window với UAT OTIF pack.
Filter mặc định: ALL trên 5 chiều (Kho/Khu vực/NVC/Loại xe) · Loại ngày = 'ETA' → cột eta_vh.
Hai mode độc lập: GT (Chuyến gửi thầu, mv_vfr_gui_thau) + VH (Chuyến vận hành, mv_vfr_van_hanh).
Target VFR = 85% overall (PRD §3.6). RAG: 🟢 ≥85 / 🟡 75–<85 / 🔴 <75.

SQL provenance:
  • KPI (Avg VFR + 4 bucket), By Area, By Vehicle — load CANONICAL từ sql-registry.md (## vfr tender / ## vfr operation).
    Dynamic-load tại runtime; registry update → script tự pick up.
  • By Vendor, By Loading Type (window-level rollup), Trend daily, Detail — AD-HOC. By Vendor không có trong
    registry (chỉ ở vfr-spec.md §22). By Loading Type registry chỉ có day-level; ở đây rollup window-level dùng
    ĐÚNG công thức weighted của registry byArea, chỉ đổi GROUP BY. Note rõ "ad-hoc" ở sheet SQL Appendix.

Lưu ý weighted vs simple-avg (QUAN TRỌNG):
  • Avg VFR (hero) + 4 bucket = simple avg(vfr_max) / countIf — TÁI TẠO được bằng Excel formula trên sheet Detail.
  • VFR theo dimension (Area/Vehicle/Vendor/Loading Type) = WEIGHTED (Loose/FP × Khối/Tấn fill×mix) — KHÔNG
    tái tạo bằng Excel formula đơn giản. Cột "VFR weighted (SQL)" = số dashboard; cột "VFR simple-avg (crosscheck)"
    = AVERAGEIFS per-trip để sanity, KHÔNG bằng weighted. Diff giữa 2 cột chính là độ lệch do weighting (liên
    quan BUG-VFR-09 / công thức V3).

Output: projects/mondelez/01-sections/vfr/uat/vfr-uat-numbers-2026-05-18_to_2026-05-22.xlsx
"""
from __future__ import annotations

import os
import re
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_DIR = _SCRIPT_DIR.parent
_OUTPUT_DIR = _PROJECT_DIR / "01-sections" / "vfr" / "uat"

try:
    from dotenv import load_dotenv
    _env = _PROJECT_DIR / ".env"
    if _env.exists():
        load_dotenv(_env)
    else:
        load_dotenv()
except ImportError:
    pass

import clickhouse_connect
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config ───────────────────────────────────────────────
WINDOW_START_DATE = "2026-05-18"
WINDOW_END_DATE = "2026-05-22"
DATE_TYPE = "ETA"  # FE gửi 'ETA' | 'ATA' (spec §5.2), default ETA → cột eta_vh
TENANT = "Mondelez"
SOURCE_GT = "analytics_workspace.mv_vfr_gui_thau"
SOURCE_VH = "analytics_workspace.mv_vfr_van_hanh"

VFR_TARGET = 85.0       # PRD §3.6
RAG_GREEN = 85.0
RAG_YELLOW_LO = 75.0

OUT_FILE = _OUTPUT_DIR / f"vfr-uat-numbers-{WINDOW_START_DATE}_to_{WINDOW_END_DATE}.xlsx"

# Mode metadata: (label, MV table name, vehicle-type filter column)
MODES = {
    "GT": {"label": "Chuyến gửi thầu (Tender)", "mv": "mv_vfr_gui_thau",
           "veh_col": "ma_loai_xe_gui_thau", "scope": "## vfr tender", "suffix": "Tender"},
    "VH": {"label": "Chuyến vận hành (Operation)", "mv": "mv_vfr_van_hanh",
           "veh_col": "ma_loai_xe_van_hanh", "scope": "## vfr operation", "suffix": "Operation"},
}


# ── Number helpers ───────────────────────────────────────
def _pct2(v):
    if v is None:
        return None
    return Decimal(str(round(float(v), 2))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _num2(v):
    if v is None:
        return None
    try:
        return Decimal(str(round(float(v), 2))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (TypeError, ValueError):
        return None


def _f(v):
    """Safe float — dang_ky columns có thể là String trong MV."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def get_client():
    return clickhouse_connect.get_client(
        host=os.getenv("CLICKHOUSE_HOST", ""),
        port=int(os.getenv("CLICKHOUSE_PORT", "8443")),
        username=os.getenv("CLICKHOUSE_USER", ""),
        password=os.getenv("CLICKHOUSE_PASSWORD", ""),
        secure=os.getenv("CLICKHOUSE_SECURE", "true").lower() in ("1", "true", "yes"),
        connect_timeout=15, send_receive_timeout=60,
    )


# ── Registry SQL loader (scope vfr tender / vfr operation, ClickHouse block) ──
REGISTRY_PATH = _PROJECT_DIR / "02-data" / "data-sources" / "sql-registry.md"
_REGISTRY_TEXT_CACHE = None


def _registry_text():
    global _REGISTRY_TEXT_CACHE
    if _REGISTRY_TEXT_CACHE is None:
        _REGISTRY_TEXT_CACHE = REGISTRY_PATH.read_text(encoding="utf-8")
    return _REGISTRY_TEXT_CACHE


def load_registry_sql(section_title, scope_header):
    """Tìm `### <section_title>` trong scope `## vfr tender` / `## vfr operation`, lấy block dưới
    **ClickHouse SQL:** (KHÔNG lấy block Redshift đứng trước). Trả về (sql, line_ref)."""
    text = _registry_text()
    sm = re.search(rf"^{re.escape(scope_header)}\s*$", text, re.MULTILINE)
    if not sm:
        raise RuntimeError(f"Không tìm thấy scope '{scope_header}' trong registry")
    start = sm.end()
    nm = re.search(r"^## ", text[start:], re.MULTILINE)
    end = start + nm.start() if nm else len(text)
    region = text[start:end]
    pat = rf"^###\s+{re.escape(section_title)}.*?\*\*ClickHouse SQL:\*\*\s*```sql\s*\n(.*?)\n```"
    m = re.search(pat, region, re.MULTILINE | re.DOTALL)
    if not m:
        raise RuntimeError(f"Không tìm thấy section '{section_title}' (ClickHouse block) trong {scope_header}")
    sql = m.group(1).strip()
    line_no = text[:start + m.start(1)].count("\n") + 1
    return sql, line_no


# ── Filter membership (ALL bypass) + placeholder substitution ──
def _csv_from_array(client, expr):
    arr = client.query(expr).result_rows[0][0] or []
    return ",".join("'" + str(x).replace("'", "''") + "'" for x in arr)


def query_filter_values(client):
    """Lấy full membership list cho ALL bypass — match đúng subquery trong registry arraySort()."""
    return {
        "whseid": _csv_from_array(client, "SELECT groupUniqArray(ma_su_dung) FROM analytics_workspace.mv_masterdata_kho_stm"),
        "area": _csv_from_array(client, "SELECT groupArray(DISTINCT group_area_code) FROM analytics_workspace.mv_filter_region"),
        "transporter": _csv_from_array(client, "SELECT groupArray(DISTINCT vendor_code) FROM analytics_workspace.mv_filter_vendor"),
        "vehicle_type": _csv_from_array(client, "SELECT groupArray(DISTINCT code) FROM analytics_workspace.mv_filter_vehicle_type"),
    }


def substitute_placeholders(sql, filters, date_type=DATE_TYPE, from_date=WINDOW_START_DATE, to_date=WINDOW_END_DATE):
    sql = sql.replace("{{whseid}}", filters["whseid"])
    sql = sql.replace("{{area}}", filters["area"])
    sql = sql.replace("{{transporter}}", filters["transporter"])
    sql = sql.replace("{{vehicle_type_tender}}", filters["vehicle_type"])
    sql = sql.replace("{{vehicle_type_ops}}", filters["vehicle_type"])
    sql = sql.replace("{{date_type}}", f"'{date_type}'")
    sql = sql.replace("{{from_date}}", f"'{from_date}'")
    sql = sql.replace("{{to_date}}", f"'{to_date}'")
    sql = sql.replace("[[", "").replace("]]", "")
    return sql


# ── Reusable WHERE blocks (literal {{...}}, substitute_placeholders fills them) ──
MEMBERSHIP_WHERE = """
AND if(arraySort([{{whseid}}]) = (SELECT arraySort(groupUniqArray(ma_su_dung)) FROM analytics_workspace.mv_masterdata_kho_stm), 1=1, t.ma_diem_nhan IN ({{whseid}}))
AND if(arraySort([{{area}}]) = (SELECT arraySort(groupArray(DISTINCT group_area_code)) FROM analytics_workspace.mv_filter_region), 1=1, t.khu_vuc_doi_xe IN ({{area}}))
AND if(arraySort([{{transporter}}]) = (SELECT arraySort(groupArray(DISTINCT vendor_code)) FROM analytics_workspace.mv_filter_vendor), 1=1, t.nha_van_tai IN ({{transporter}}))
AND if(arraySort([{{vehicle_type_tender}}]) = (SELECT arraySort(groupArray(DISTINCT code)) FROM analytics_workspace.mv_filter_vehicle_type), 1=1, t.__VEH_COL__ IN ({{vehicle_type_tender}}))
"""

DATE_WHERE = """
AND ( toDate(CASE
    WHEN {{date_type}} = 'ETA' THEN t.eta_vh
    WHEN {{date_type}} = 'ATA' THEN t.ata_vh
    WHEN {{date_type}} = 'Ngày gửi thầu' THEN t.tender_date
  END) BETWEEN toDate(coalesce({{from_date}}, '1900-01-01')) AND toDate(coalesce({{to_date}}, '2999-12-31')) )
"""

# Weighted-dim template = registry byArea formula, GROUP BY tham số hoá (ad-hoc cho Vendor / Loading Type)
WEIGHTED_DIM = """
WITH base AS (
    SELECT
        t.__GROUP_COL__ AS dim_value,
        SUM(t.cbm_ke_hoach) AS total_cbm_ke_hoach,
        SUM(t.cbm_nhan)      AS total_cbm_nhan,
        SUM(IF(trim(t.loai_boc_xep) IN ('Loose','Losse'), t.cbm_nhan, 0)) AS loose_cbm_nhan,
        SUM(IF(trim(t.loai_boc_xep) IN ('Loose','Losse') AND t.phan_loai_vfr='Khối', t.cbm_nhan, 0)) AS loose_khoi_cbm_nhan,
        SUM(IF(trim(t.loai_boc_xep) IN ('Loose','Losse') AND t.phan_loai_vfr='Khối', toFloat64OrZero(toString(t.cbm_dang_ky)), 0)) AS loose_khoi_cbm_dk,
        SUM(IF(trim(t.loai_boc_xep) IN ('Loose','Losse') AND t.phan_loai_vfr='Tấn', t.tan_nhan, 0)) AS loose_tan_nhan,
        SUM(IF(trim(t.loai_boc_xep) IN ('Loose','Losse') AND t.phan_loai_vfr='Tấn', toFloat64OrZero(toString(t.tan_dang_ky)), 0)) AS loose_tan_dk,
        SUM(IF(trim(t.loai_boc_xep) IN ('Loose','Losse') AND t.phan_loai_vfr='Tấn', t.cbm_nhan, 0)) AS loose_tan_cbm_nhan,
        SUM(IF(t.loai_boc_xep='Full Pallet', t.cbm_nhan, 0)) AS fp_cbm_nhan,
        SUM(IF(t.loai_boc_xep='Full Pallet' AND t.phan_loai_vfr='Khối', t.cbm_nhan, 0)) AS fp_khoi_cbm_nhan,
        SUM(IF(t.loai_boc_xep='Full Pallet' AND t.phan_loai_vfr='Khối', toFloat64OrZero(toString(t.cbm_dang_ky)), 0)) AS fp_khoi_cbm_dk,
        SUM(IF(t.loai_boc_xep='Full Pallet' AND t.phan_loai_vfr='Tấn', t.tan_nhan, 0)) AS fp_tan_nhan,
        SUM(IF(t.loai_boc_xep='Full Pallet' AND t.phan_loai_vfr='Tấn', toFloat64OrZero(toString(t.tan_dang_ky)), 0)) AS fp_tan_dk,
        SUM(IF(t.loai_boc_xep='Full Pallet' AND t.phan_loai_vfr='Tấn', t.cbm_nhan, 0)) AS fp_tan_cbm_nhan
    FROM analytics_workspace.__MV__ t
    WHERE 1 = 1
__MEMBERSHIP__
__DATE__
    GROUP BY t.__GROUP_COL__
),
calc AS (
    SELECT
        dim_value, total_cbm_ke_hoach, total_cbm_nhan,
        if(total_cbm_nhan=0, 0, loose_cbm_nhan/total_cbm_nhan) AS loose_weight,
        if(total_cbm_nhan=0, 0, fp_cbm_nhan/total_cbm_nhan)    AS fp_weight,
        if(loose_khoi_cbm_dk=0, 0, loose_khoi_cbm_nhan/loose_khoi_cbm_dk) AS loose_khoi_fill_rate,
        if(loose_cbm_nhan=0, 0, loose_khoi_cbm_nhan/loose_cbm_nhan)       AS loose_khoi_mix_rate,
        if(loose_tan_dk=0, 0, loose_tan_nhan/loose_tan_dk)               AS loose_tan_fill_rate,
        if(loose_cbm_nhan=0, 0, loose_tan_cbm_nhan/loose_cbm_nhan)       AS loose_tan_mix_rate,
        if(fp_khoi_cbm_dk=0, 0, fp_khoi_cbm_nhan/fp_khoi_cbm_dk)         AS fp_khoi_fill_rate,
        if(fp_cbm_nhan=0, 0, fp_khoi_cbm_nhan/fp_cbm_nhan)               AS fp_khoi_mix_rate,
        if(fp_tan_dk=0, 0, fp_tan_nhan/fp_tan_dk)                       AS fp_tan_fill_rate,
        if(fp_cbm_nhan=0, 0, fp_tan_cbm_nhan/fp_cbm_nhan)               AS fp_tan_mix_rate
    FROM base
)
SELECT
    dim_value,
    total_cbm_ke_hoach AS planned,
    round(least(1.0, (
        (loose_khoi_fill_rate*loose_khoi_mix_rate + loose_tan_fill_rate*loose_tan_mix_rate) * loose_weight
        + (fp_khoi_fill_rate*fp_khoi_mix_rate + fp_tan_fill_rate*fp_tan_mix_rate) * fp_weight
    )) * 100, 2) AS vfr_ratio
FROM calc
__EXTRA_WHERE__
ORDER BY vfr_ratio ASC
"""

TREND_SQL = """
SELECT
    toDate(CASE
        WHEN {{date_type}} = 'ETA' THEN t.eta_vh
        WHEN {{date_type}} = 'ATA' THEN t.ata_vh
        WHEN {{date_type}} = 'Ngày gửi thầu' THEN t.tender_date
    END) AS d,
    round(avg(t.vfr_max), 2) AS avg_vfr,
    count() AS total_trips
FROM analytics_workspace.__MV__ t
WHERE 1 = 1
__MEMBERSHIP__
__DATE__
GROUP BY d
ORDER BY d
"""

DETAIL_SQL = """
SELECT
    t.id_chuyen_gui_thau AS trip_id,
    t.ma_chuyen_van_hanh AS ma_chuyen_vh,
    t.trang_thai_chuyen  AS trang_thai,
    t.ma_diem_nhan       AS wh_code,
    if(t.diem_nhan = '', '(rỗng)', t.diem_nhan)               AS wh_name,
    if(t.khu_vuc_doi_xe = '', '(rỗng)', t.khu_vuc_doi_xe)     AS khu_vuc,
    if(t.nha_van_tai = '', '(rỗng)', t.nha_van_tai)           AS nvc,
    if(t.loai_xe_gui_thau = '', '(rỗng)', t.loai_xe_gui_thau) AS loai_xe_gt,
    if(t.loai_xe_van_hanh = '', '(rỗng)', t.loai_xe_van_hanh) AS loai_xe_vh,
    if(t.loai_boc_xep = '', '(rỗng)', t.loai_boc_xep)         AS loai_boc_xep,
    coalesce(t.phan_loai_vfr, '')                            AS phan_loai_vfr,
    toDateTime(t.eta_vh, 'Asia/Ho_Chi_Minh')                AS eta_utc7,
    toDateTime(t.ata_vh, 'Asia/Ho_Chi_Minh')                AS ata_utc7,
    toString(t.tan_dang_ky)  AS tan_dang_ky,
    t.tan_nhan               AS tan_nhan,
    toString(t.cbm_dang_ky)  AS cbm_dang_ky,
    t.cbm_ke_hoach           AS cbm_ke_hoach,
    t.cbm_nhan               AS cbm_nhan,
    t.vfr_theo_tan           AS vfr_theo_tan,
    t.vfr_theo_khoi          AS vfr_theo_khoi,
    t.vfr_max                AS vfr_max
FROM analytics_workspace.__MV__ t
WHERE 1 = 1
__MEMBERSHIP__
__DATE__
ORDER BY eta_utc7 ASC, trip_id ASC
"""


def _assemble(template, mv, veh_col, group_col=None, extra_where=""):
    sql = template.replace("__MV__", mv)
    sql = sql.replace("__MEMBERSHIP__", MEMBERSHIP_WHERE.replace("__VEH_COL__", veh_col))
    sql = sql.replace("__DATE__", DATE_WHERE)
    if group_col is not None:
        sql = sql.replace("__GROUP_COL__", group_col)
    sql = sql.replace("__EXTRA_WHERE__", extra_where)
    return sql


def _bucket(vfr_max):
    v = _f(vfr_max)
    if v < 50:
        return "Low <50%"
    if v < 70:
        return "Medium 50-70%"
    if v < 95:
        return "High 70-95%"
    return "Excellent ≥95%"


# ── Runners ──────────────────────────────────────────────
def run_kpi(client, mode, filters):
    m = MODES[mode]
    title = f"Avg VFR (VFR by {m['suffix']} Trip)"
    tpl, line_ref = load_registry_sql(title, m["scope"])
    sql = substitute_placeholders(tpl, filters)
    r = client.query(sql)
    row = dict(zip(r.column_names, r.result_rows[0]))
    out = {
        "avg_vfr": _pct2(row.get("avg_vfr")),
        "low": row.get("cnt_vfr_50"),
        "medium": row.get("cnt_vfr_50_70"),
        "high": row.get("cnt_vfr_70_95"),
        "excellent": row.get("cnt_vfr_95"),
    }
    out["total"] = (out["low"] or 0) + (out["medium"] or 0) + (out["high"] or 0) + (out["excellent"] or 0)
    return out, sql, f"sql-registry.md:{line_ref}"


def run_registry_dim(client, mode, filters, dim):
    m = MODES[mode]
    titles = {
        "area": f"VFR {'gửi thầu' if mode == 'GT' else 'vận hành'} theo khu vực (VFR by {m['suffix']} Trip)",
        "vehicle": f"VFR {'gửi thầu' if mode == 'GT' else 'vận hành'} theo loại xe (VFR by {m['suffix']} Trip)",
    }
    tpl, line_ref = load_registry_sql(titles[dim], m["scope"])
    sql = substitute_placeholders(tpl, filters)
    r = client.query(sql)
    cols = list(r.column_names)
    rows = []
    for raw in r.result_rows:
        d = dict(zip(cols, raw))
        rows.append({"dim_value": d.get(cols[0]), "planned": _num2(d.get("total_cbm_ke_hoach")),
                     "vfr": _pct2(d.get("vfr_ratio"))})
    rows.sort(key=lambda x: (x["vfr"] if x["vfr"] is not None else Decimal(0)))
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_adhoc_dim(client, mode, filters, group_col):
    m = MODES[mode]
    sql = _assemble(WEIGHTED_DIM, m["mv"], m["veh_col"], group_col=group_col,
                    extra_where="WHERE dim_value != ''")
    sql = substitute_placeholders(sql, filters)
    r = client.query(sql)
    cols = list(r.column_names)
    rows = []
    for raw in r.result_rows:
        d = dict(zip(cols, raw))
        rows.append({"dim_value": d.get("dim_value"), "planned": _num2(d.get("planned")),
                     "vfr": _pct2(d.get("vfr_ratio"))})
    rows.sort(key=lambda x: (x["vfr"] if x["vfr"] is not None else Decimal(0)))
    return rows, sql.strip()


def run_trend(client, mode, filters):
    m = MODES[mode]
    sql = _assemble(TREND_SQL, m["mv"], m["veh_col"])
    sql = substitute_placeholders(sql, filters)
    r = client.query(sql)
    rows = [{"d": d, "avg_vfr": _pct2(a), "total": tt} for d, a, tt in r.result_rows]
    return rows, sql.strip()


def run_detail(client, mode, filters):
    m = MODES[mode]
    sql = _assemble(DETAIL_SQL, m["mv"], m["veh_col"])
    sql = substitute_placeholders(sql, filters)
    r = client.query(sql)
    cols = list(r.column_names)
    rows = []
    for raw in r.result_rows:
        d = dict(zip(cols, raw))
        for k in ("eta_utc7", "ata_utc7"):
            v = d.get(k)
            if v is not None and getattr(v, "tzinfo", None) is not None:
                d[k] = v.replace(tzinfo=None)
        d["mode"] = mode
        d["bucket"] = _bucket(d.get("vfr_max"))
        rows.append(d)
    return rows, sql.strip()


# ── Styles ───────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
SUB_FILL = PatternFill("solid", fgColor="14283F")
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="606060")
BODY_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="BFC9D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
GREEN = PatternFill("solid", fgColor="DCFCE7")
YELLOW = PatternFill("solid", fgColor="FEF3C7")
RED = PatternFill("solid", fgColor="FEE2E2")
GREY = PatternFill("solid", fgColor="F1F5F9")
EMPTY_CUSTOMER = PatternFill("solid", fgColor="FFFBEB")
PCT_FMT = '0.00"%"'
CNT_FMT = "#,##0"
NUM_FMT = "#,##0.00"


def rag_fill_vfr(value):
    if value is None:
        return GREY
    v = float(value)
    if v >= RAG_GREEN:
        return GREEN
    if v >= RAG_YELLOW_LO:
        return YELLOW
    return RED


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_row(ws, text, span):
    ws.append([text] + [""] * (span - 1))
    last = ws.max_row
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=span)
    c = ws.cell(row=last, column=1)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[last].height = 24


def header_row(ws, headers):
    ws.append(headers)
    last = ws.max_row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=last, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[last].height = 30


def sub_row(ws, text, span):
    ws.append([text] + [""] * (span - 1))
    last = ws.max_row
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=span)
    c = ws.cell(row=last, column=1)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = SUB_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[last].height = 22


def body_cell(c, value=None, fill=None, fmt=None, align=None):
    c.font = BODY_FONT
    c.border = BORDER
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if value is not None:
        c.value = value


# ── Detail sheet column map (for live formulas) ──────────
DETAIL_SHEET = "'04 — Detail Trips'"
# A=Mode B=Trip C=MãVH D=Trạng thái E=WH code F=WH name G=Khu vực H=NVC
# I=Loại xe GT J=Loại xe VH K=Loại bốc xếp L=Phân loại VFR M=ETA N=ATA
# O=Tấn đk P=Tấn nhận Q=CBM đk R=CBM kh S=CBM nhận T=VFR tấn U=VFR khối V=VFR max W=Bucket
COL = {"mode": "A", "khu_vuc": "G", "nvc": "H", "loai_xe_gt": "I", "loai_xe_vh": "J",
       "loai_boc_xep": "K", "eta": "M", "cbm_kh": "R", "vfr_max": "V"}


def f_total(mode):
    return f'=COUNTIF({DETAIL_SHEET}!A:A,"{mode}")'


def f_avg_vfr(mode):
    return f'=IFERROR(ROUND(AVERAGEIFS({DETAIL_SHEET}!V:V,{DETAIL_SHEET}!A:A,"{mode}"),2),0)'


def f_bucket(mode, lo, hi):
    crit = [f'{DETAIL_SHEET}!A:A,"{mode}"']
    if lo is not None:
        crit.append(f'{DETAIL_SHEET}!V:V,">={lo}"')
    if hi is not None:
        crit.append(f'{DETAIL_SHEET}!V:V,"<{hi}"')
    return f'=COUNTIFS({",".join(crit)})'


def f_dim_avg(mode, dim_col, value_cell):
    return (f'=IFERROR(ROUND(AVERAGEIFS({DETAIL_SHEET}!V:V,{DETAIL_SHEET}!A:A,"{mode}",'
            f'{DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell}),2),0)')


def f_dim_planned(mode, dim_col, value_cell):
    return (f'=SUMIFS({DETAIL_SHEET}!R:R,{DETAIL_SHEET}!A:A,"{mode}",'
            f'{DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell})')


def _date_args(date_str):
    y, m, d = date_str.split("-")
    return (f'{DETAIL_SHEET}!M:M,">="&DATE({y},{m},{d}),'
            f'{DETAIL_SHEET}!M:M,"<"&DATE({y},{m},{d})+1')


def f_trend_total(mode, date_str):
    return f'=COUNTIFS({DETAIL_SHEET}!A:A,"{mode}",{_date_args(date_str)})'


def f_trend_avg(mode, date_str):
    return (f'=IFERROR(ROUND(AVERAGEIFS({DETAIL_SHEET}!V:V,{DETAIL_SHEET}!A:A,"{mode}",'
            f'{_date_args(date_str)}),2),0)')


# ── Excel build ──────────────────────────────────────────
def build_readme(ws, ctx):
    ws.title = "00 — README"
    set_widths(ws, [24, 88])
    title_row(ws, "UAT VFR — Số SQL thật cho khách Mondelez confirm", 2)
    meta = [
        ("Section", "VFR (Vehicle Fill Rate) — Smartlog Control Tower"),
        ("Tenant", TENANT),
        ("Window test", f"{WINDOW_START_DATE} → {WINDOW_END_DATE} (5 ngày) — cùng window UAT OTIF"),
        ("Loại ngày", f"'{DATE_TYPE}'  → cột MV: eta_vh (ETA chuyến vận hành)"),
        ("Filter còn lại", "ALL (Kho / Khu vực / NVC / Loại xe không chọn)"),
        ("Hai mode độc lập", "GT = Chuyến gửi thầu (mv_vfr_gui_thau) · VH = Chuyến vận hành (mv_vfr_van_hanh). "
                              "Widget toggle 1 mode/lần; hero luôn hiện cả 2 + Δ GT−VH."),
        ("Nguồn dữ liệu", f"{SOURCE_GT} · {SOURCE_VH}"),
        ("Target VFR", f"{VFR_TARGET:.0f}% overall (PRD §3.6). RAG: 🟢 ≥{RAG_GREEN:.0f} / 🟡 {RAG_YELLOW_LO:.0f}–<{RAG_GREEN:.0f} / 🔴 <{RAG_YELLOW_LO:.0f}."),
        ("Công thức VFR", "Per-trip vfr_max = MAX(VFR tấn, VFR khối) (pre-computed trong MV). "
                          "Avg VFR (hero) = simple avg(vfr_max). VFR theo dimension = WEIGHTED (Loose/FP × Khối/Tấn fill×mix)."),
        ("SQL provenance", "KPI (Avg + 4 bucket), By Area, By Vehicle = CANONICAL load từ sql-registry.md "
                           "(## vfr tender / ## vfr operation). By Vendor + By Loading Type (rollup window-level) + Trend + "
                           "Detail = AD-HOC (xem sheet 07). Khi registry update → re-run script tự pick up."),
        ("⚠ Weighted vs simple-avg", "Sheet 02 Dimension: cột 'VFR weighted (SQL)' = số dashboard (canonical). Cột "
                                      "'VFR simple-avg (crosscheck)' = AVERAGEIFS per-trip — KHÔNG bằng weighted, chỉ để sanity. "
                                      "Diff giữa 2 cột = độ lệch do weighting (liên quan BUG-VFR-09 / công thức V3 Bước 4-5)."),
        ("⚠ Bucket ≥95 vs >95", "Registry hiện dùng countIf(vfr_max >= 95) (BUG-VFR-08 đã fix trong registry). Nếu config "
                                 "runtime của widget chưa re-paste (rollout pending) → dashboard có thể dùng >95, lệch đúng "
                                 "số trip có vfr_max = 95.0. Cần đối chiếu khi reconcile bucket Excellent."),
        ("⚠ Không clamp >100%", "vfr_max > 100 render raw + cảnh báo icon, vẫn đếm vào Excellent. Trip over-registered (đk < thực "
                                 "nhận) đẩy Avg VFR lên. Xem sheet 04 Detail filter V>100 + insight pack."),
        ("Run timestamp", ctx["run_ts"]),
        ("CH server now (UTC+7)", str(ctx["server_now"]) if ctx["server_now"] else "n/a"),
        ("Tổng chuyến (window)", f"GT = {ctx['kpi']['GT']['total']:,} · VH = {ctx['kpi']['VH']['total']:,}"),
        ("", ""),
        ("Mục đích file", "PM/BA mang số SQL thật xuống khách Mondelez confirm: (1) Định nghĩa VFR (MAX tấn/khối, weighted "
                          "Loose/FP) có khớp công thức MDLZ; (2) Avg VFR + 4 nhóm tải có đúng kỳ vọng; (3) Khu vực / loại xe / "
                          "NVC / loại bốc xếp nào non tải; (4) GT vs VH chênh lệch có hợp lý không."),
        ("Hướng dẫn cho khách", "Khách nhìn cột 'SQL value', so số nội bộ MDLZ. Khớp trong tolerance → Status = OK. Lệch → ghi rõ "
                                "vào cột 'Khách ghi chú' để Smartlog điều tra."),
        ("Tolerance default", "% VFR ≤ 0.5pp · Số đếm (bucket / chuyến) ≤ 1% · Top N dim ≥ 4/5 tên match (thứ tự có thể lệch)"),
        ("", ""),
        ("Sheet trong file", "01 — KPI Hero & Buckets (Avg VFR + Δ GT−VH + 4 bucket, GT & VH, SQL + formula live + RAG) · "
                             "02 — Dimension Matrix (By Area / Vehicle / Vendor / Loading Type × 2 mode, weighted SQL + simple-avg "
                             "crosscheck, worst-first) · 03 — Trend daily (Avg VFR/ngày, GT & VH) · 04 — Detail Trips (raw, cột Mode, "
                             "autofilter + freeze, là nguồn mọi formula) · 05 — Formula Guide · 06 — UX & Filter checklist (khách fill "
                             "trong session) · 07 — SQL Appendix (provenance registry vs ad-hoc)"),
        ("Insight file riêng", f"Operational insight (master data scorecard, red flags non-tải, KPI extension, open questions) ở "
                               f"file md riêng: vfr-uat-ops-insight-{WINDOW_START_DATE}_to_{WINDOW_END_DATE}.md (same folder) — "
                               f"KHÔNG nhồi vào Excel để giữ pack tập trung reconciliation."),
        ("Cách verify dual", "Sheet 01 + 03: cột 'SQL value' (frozen, Python ghi) phải = cột 'Công thức Excel' (live từ sheet 04). "
                             "Lệch → có gap filter. Sheet 02 dimension: SQL weighted ≠ simple-avg formula là BÌNH THƯỜNG (xem note weighted)."),
        ("Format chuẩn", "% VFR: '0.00\"%\"' (vd 94.49%, 82.30%). Count: '#,##0' (vd 1,234). CBM/số có thập phân: '#,##0.00'. "
                         "Datetime ETA/ATA: 'yyyy-mm-dd hh:mm:ss' (datetime cell, sort/filter range OK, đã strip tzinfo)."),
        ("Re-run", "python scripts/uat_vfr_export.py — đổi WINDOW_START_DATE / WINDOW_END_DATE / DATE_TYPE ở đầu file rồi chạy lại."),
    ]
    for k, v in meta:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP
    for r_ in range(2, ws.max_row + 1):
        ln = len(str(ws.cell(row=r_, column=2).value or ""))
        ws.row_dimensions[r_].height = max(18, min(96, 18 + ln // 64 * 16))


def _add_kpi_row(ws, label, sql_val, formula, is_pct=False, rag=False, tol=""):
    ws.append([label, sql_val, formula, "", "", "", tol, ""])
    r_ = ws.max_row
    for col in range(1, 9):
        body_cell(ws.cell(row=r_, column=col))
    fmt = PCT_FMT if is_pct else CNT_FMT
    ws.cell(row=r_, column=2).number_format = fmt
    ws.cell(row=r_, column=3).number_format = fmt
    if rag and sql_val is not None:
        ws.cell(row=r_, column=2).fill = rag_fill_vfr(float(sql_val))
    for col in (4, 5):
        ws.cell(row=r_, column=col).fill = EMPTY_CUSTOMER
    df = ws.cell(row=r_, column=6)
    df.value = f'=IFERROR(D{r_}-B{r_},"")'
    df.number_format = fmt


def build_kpi_sheet(ws, ctx):
    ws.title = "01 — KPI Hero & Buckets"
    set_widths(ws, [40, 16, 56, 16, 16, 16, 16, 26])
    title_row(ws, f"L1 Hero KPI + L2 Buckets — {WINDOW_START_DATE} → {WINDOW_END_DATE} (Loại ngày {DATE_TYPE}) — target {VFR_TARGET:.0f}%", 8)

    for mode in ("GT", "VH"):
        k = ctx["kpi"][mode]
        sub_row(ws, f"Mode {mode} — {MODES[mode]['label']}  (nguồn: analytics_workspace.{MODES[mode]['mv']})", 8)
        header_row(ws, ["Metric", "SQL value\n(frozen)", "Công thức Excel\n(live từ sheet 04)",
                        "Dashboard\n(khách verify)", "Số MDLZ\n(golden)", "Diff\n(Dash − SQL)",
                        "Tolerance", "Khách ghi chú / Status"])
        _add_kpi_row(ws, f"Avg VFR ({mode}) — simple avg(vfr_max)", k["avg_vfr"], f_avg_vfr(mode),
                     is_pct=True, rag=True, tol="≤ 0.5 pp")
        _add_kpi_row(ws, f"Tổng chuyến ({mode})", k["total"], f_total(mode), tol="≤ 1%")
        _add_kpi_row(ws, f"Low <50% (đỏ) ({mode})", k["low"], f_bucket(mode, None, 50), tol="≤ 1%")
        _add_kpi_row(ws, f"Medium 50-70% (vàng) ({mode})", k["medium"], f_bucket(mode, 50, 70), tol="≤ 1%")
        _add_kpi_row(ws, f"High 70-95% (xanh) ({mode})", k["high"], f_bucket(mode, 70, 95), tol="≤ 1%")
        _add_kpi_row(ws, f"Excellent ≥95% (xanh đậm) ({mode})", k["excellent"], f_bucket(mode, 95, None), tol="≤ 1%")
        ws.append([])

    # GT vs VH delta (hero highlight)
    gt, vh = ctx["kpi"]["GT"]["avg_vfr"], ctx["kpi"]["VH"]["avg_vfr"]
    delta = _pct2(float(gt) - float(vh)) if gt is not None and vh is not None else None
    sub_row(ws, "Hero — GT vs VH (luôn hiển thị bất kể mode toggle)", 8)
    ws.append(["Δ GT − VH (chênh lệch Avg VFR)", delta,
               f'=IFERROR(ROUND({f_avg_vfr("GT")[1:]}-({f_avg_vfr("VH")[1:]}),2),0)'.replace("==", "="),
               "", "", "", "≤ 0.5 pp", ""])
    r_ = ws.max_row
    for col in range(1, 9):
        body_cell(ws.cell(row=r_, column=col))
    ws.cell(row=r_, column=2).number_format = PCT_FMT
    ws.cell(row=r_, column=2).value = delta
    # Recompute the live delta formula cleanly
    ws.cell(row=r_, column=3).value = (
        f'=IFERROR(ROUND(AVERAGEIFS({DETAIL_SHEET}!V:V,{DETAIL_SHEET}!A:A,"GT")'
        f'-AVERAGEIFS({DETAIL_SHEET}!V:V,{DETAIL_SHEET}!A:A,"VH"),2),0)')
    ws.cell(row=r_, column=3).number_format = PCT_FMT
    for col in (4, 5):
        ws.cell(row=r_, column=col).fill = EMPTY_CUSTOMER

    ws.append([])
    ws.append(["Note", "Avg VFR + bucket = simple avg / countIf → SQL value khớp công thức Excel live. "
                       "Bucket boundary: <50 / [50,70) / [70,95) / ≥95 (registry). "
                       "Δ GT−VH cho biết chất lượng kế hoạch (GT) vs thực thi (VH)."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)


def build_dim_sheet(ws, ctx):
    ws.title = "02 — Dimension Matrix"
    set_widths(ws, [30, 34, 18, 18, 22, 16, 16, 26])
    title_row(ws, "Dimension VFR (worst-first) — WEIGHTED SQL = dashboard · simple-avg = crosscheck (≠ weighted)", 8)

    dim_blocks = [
        ("By Area — VFR theo Khu vực", "khu_vuc", "area"),
        ("By Vehicle — VFR theo Loại xe", None, "vehicle"),       # dim_col resolved per mode below
        ("By Vendor — VFR theo Nhà vận tải", "nvc", "vendor"),
        ("By Loading Type — VFR theo Loại bốc xếp", "loai_boc_xep", "loading_type"),
    ]
    for label, dim_col_key, dim_key in dim_blocks:
        for mode in ("GT", "VH"):
            rows = ctx["dim"][dim_key][mode]
            sub_row(ws, f"{label}  —  Mode {mode} ({MODES[mode]['label']})", 8)
            header_row(ws, ["Giá trị", "(loại)", "VFR weighted\n(SQL = dashboard)",
                            "VFR simple-avg\n(crosscheck, ≠ weighted)", "CBM kế hoạch\n(planned)",
                            "Dashboard\n(khách)", "Diff\n(Dash − SQL)", "Khách ghi chú"])
            if dim_col_key is not None:
                dim_col = COL[dim_col_key]
            else:  # By Vehicle: GT→loai_xe_gt (I), VH→loai_xe_vh (J)
                dim_col = COL["loai_xe_gt"] if mode == "GT" else COL["loai_xe_vh"]
            if not rows:
                ws.append(["(không có data)", "", "", "", "", "", "", ""])
                for col in range(1, 9):
                    body_cell(ws.cell(row=ws.max_row, column=col))
                continue
            for row in rows:
                val = row["dim_value"] if row["dim_value"] not in (None, "") else "(rỗng)"
                ws.append([val, "", row["vfr"], "", row["planned"], "", "", ""])
                r_ = ws.max_row
                v_ref = f"$A{r_}"
                ws.cell(row=r_, column=4).value = f_dim_avg(mode, dim_col, v_ref)
                ws.cell(row=r_, column=6).value = ""   # Dashboard (khách)
                ws.cell(row=r_, column=7).value = f'=IFERROR(F{r_}-C{r_},"")'
                for col in range(1, 9):
                    body_cell(ws.cell(row=r_, column=col))
                ws.cell(row=r_, column=3).number_format = PCT_FMT
                ws.cell(row=r_, column=4).number_format = PCT_FMT
                ws.cell(row=r_, column=5).number_format = NUM_FMT
                ws.cell(row=r_, column=7).number_format = PCT_FMT
                if row["vfr"] is not None:
                    ws.cell(row=r_, column=3).fill = rag_fill_vfr(float(row["vfr"]))
                ws.cell(row=r_, column=6).fill = EMPTY_CUSTOMER
                ws.cell(row=r_, column=8).fill = EMPTY_CUSTOMER
            ws.append([])


def build_trend_sheet(ws, ctx):
    ws.title = "03 — Trend daily"
    set_widths(ws, [16, 16, 16, 56, 16, 16, 24])
    title_row(ws, "Trend Avg VFR theo ngày (sparkline source) — GT & VH — số live từ sheet 04", 7)
    for mode in ("GT", "VH"):
        sub_row(ws, f"Mode {mode} — {MODES[mode]['label']}", 7)
        header_row(ws, ["Ngày", "Avg VFR\n(SQL frozen)", "Tổng chuyến\n(SQL)",
                        "Công thức Avg VFR\n(live từ sheet 04)", "Dashboard\n(khách)",
                        "Diff\n(Dash − SQL)", "Khách ghi chú"])
        for row in ctx["trend"][mode]:
            ds = row["d"].strftime("%Y-%m-%d")
            ws.append([row["d"], row["avg_vfr"], row["total"], f_trend_avg(mode, ds), "", "", ""])
            r_ = ws.max_row
            ws.cell(row=r_, column=4).value = f_trend_avg(mode, ds)
            ws.cell(row=r_, column=6).value = f'=IFERROR(E{r_}-B{r_},"")'
            for col in range(1, 8):
                body_cell(ws.cell(row=r_, column=col))
            ws.cell(row=r_, column=1).number_format = "yyyy-mm-dd"
            ws.cell(row=r_, column=2).number_format = PCT_FMT
            ws.cell(row=r_, column=3).number_format = CNT_FMT
            ws.cell(row=r_, column=4).number_format = PCT_FMT
            ws.cell(row=r_, column=6).number_format = PCT_FMT
            if row["avg_vfr"] is not None:
                ws.cell(row=r_, column=2).fill = rag_fill_vfr(float(row["avg_vfr"]))
            ws.cell(row=r_, column=5).fill = EMPTY_CUSTOMER
        ws.append([])
    ws.append(["Note", "Sheet này là Avg VFR/ngày trong window 5 ngày. Hero sparkline trên widget hiển thị 14 ngày gần "
                       "nhất (VFR_SPARKLINE_DAYS=14) tính tới hôm nay — khác window UAT. Khi reconcile sparkline, dùng window "
                       "14d riêng (re-run script đổi date range)."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=7)


def build_detail_sheet(ws, ctx):
    ws.title = "04 — Detail Trips"
    headers = [
        "Mode", "Trip ID", "Mã chuyến VH", "Trạng thái", "WH code", "WH name", "Khu vực", "NVC",
        "Loại xe GT", "Loại xe VH", "Loại bốc xếp", "Phân loại VFR", "ETA (UTC+7)", "ATA (UTC+7)",
        "Tấn đăng ký", "Tấn nhận", "CBM đăng ký", "CBM kế hoạch", "CBM nhận",
        "VFR tấn", "VFR khối", "VFR max", "Bucket", "Khớp golden? (Y/N)", "Khách ghi chú",
    ]
    widths = [7, 18, 18, 16, 12, 24, 18, 20, 16, 16, 16, 14, 20, 20,
              13, 13, 14, 14, 13, 12, 12, 12, 16, 16, 30]
    set_widths(ws, widths)
    total = len(ctx["detail"])
    title_row(ws, f"Detail Trips — {total:,} chuyến (GT + VH) cho khách đối chiếu line-by-line — nguồn mọi formula", len(headers))
    header_row(ws, headers)

    status_fill = {"Low <50%": RED, "Medium 50-70%": YELLOW, "High 70-95%": GREEN, "Excellent ≥95%": GREEN}
    for row in ctx["detail"]:
        ws.append([
            row["mode"], row.get("trip_id"), row.get("ma_chuyen_vh"), row.get("trang_thai"),
            row.get("wh_code"), row.get("wh_name"), row.get("khu_vuc"), row.get("nvc"),
            row.get("loai_xe_gt"), row.get("loai_xe_vh"), row.get("loai_boc_xep"), row.get("phan_loai_vfr"),
            row.get("eta_utc7"), row.get("ata_utc7"),
            _num2(_f(row.get("tan_dang_ky"))), _num2(_f(row.get("tan_nhan"))),
            _num2(_f(row.get("cbm_dang_ky"))), _num2(_f(row.get("cbm_ke_hoach"))), _num2(_f(row.get("cbm_nhan"))),
            _pct2(row.get("vfr_theo_tan")), _pct2(row.get("vfr_theo_khoi")), _pct2(row.get("vfr_max")),
            row.get("bucket"), "", "",
        ])
        r_ = ws.max_row
        for col in range(1, len(headers) + 1):
            body_cell(ws.cell(row=r_, column=col))
        for col in (13, 14):
            ws.cell(row=r_, column=col).number_format = "yyyy-mm-dd hh:mm:ss"
        for col in (15, 16, 17, 18, 19):
            ws.cell(row=r_, column=col).number_format = NUM_FMT
        for col in (20, 21, 22):
            ws.cell(row=r_, column=col).number_format = PCT_FMT
        # VFR max RAG + bucket color
        if ws.cell(row=r_, column=22).value is not None:
            ws.cell(row=r_, column=22).fill = rag_fill_vfr(_f(row.get("vfr_max")))
        ws.cell(row=r_, column=23).fill = status_fill.get(row.get("bucket"), GREY)
        for col in (24, 25):
            ws.cell(row=r_, column=col).fill = EMPTY_CUSTOMER
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"


def build_formula_guide_sheet(ws, ctx):
    ws.title = "05 — Formula Guide"
    set_widths(ws, [30, 92])
    title_row(ws, "Excel Formula Guide — copy paste khi PM/khách verify thủ công", 2)
    intro = [
        ("Mục đích", "Số ở sheet 01 (KPI) + 03 (Trend) đều có công thức live tính lại từ sheet 04 (Detail Trips). "
                     "Copy công thức từ đây để audit."),
        ("⚠ Weighted vs simple-avg", "VFR theo dimension (sheet 02) là WEIGHTED (Loose/FP × Khối/Tấn fill×mix) — KHÔNG tái tạo "
                                      "bằng AVERAGEIFS. Cột simple-avg ở sheet 02 chỉ là sanity bound. Avg VFR hero (sheet 01) "
                                      "LÀ simple avg(vfr_max) → AVERAGEIFS khớp."),
        ("ETA/ATA = datetime", "Cột M + N sheet 04 là datetime cell (đã strip tzinfo, UTC+7). Sort/filter range OK. Trend dùng "
                               "COUNTIFS/AVERAGEIFS với DATE() bound thay vì prefix-text."),
        ("Sheet name reference", "Tên sheet detail trong công thức: '04 — Detail Trips' (giữ nguyên hoặc Find&Replace đồng bộ)."),
    ]
    for k, v in intro:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP

    ws.append([])
    ws.append(["Bảng 1 — Mapping cột Detail Trips", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = SUB_FILL
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    col_map = [
        ("A", "Mode (GT / VH)"), ("B", "Trip ID"), ("C", "Mã chuyến VH"), ("D", "Trạng thái chuyến"),
        ("E", "WH code (ma_diem_nhan)"), ("F", "WH name"), ("G", "Khu vực (khu_vuc_doi_xe)"),
        ("H", "NVC (nha_van_tai)"), ("I", "Loại xe GT (loai_xe_gui_thau)"), ("J", "Loại xe VH (loai_xe_van_hanh)"),
        ("K", "Loại bốc xếp (Loose / Full Pallet / ...)"), ("L", "Phân loại VFR (Tấn / Khối)"),
        ("M", "ETA UTC+7 (datetime)"), ("N", "ATA UTC+7 (datetime)"), ("O", "Tấn đăng ký"), ("P", "Tấn nhận"),
        ("Q", "CBM đăng ký"), ("R", "CBM kế hoạch"), ("S", "CBM nhận"), ("T", "VFR theo tấn"),
        ("U", "VFR theo khối"), ("V", "VFR max (= MAX tấn/khối) — KEY cho KPI/bucket"), ("W", "Bucket (Low/Medium/High/Excellent)"),
        ("X", "Khớp golden? (Y/N) — khách điền"), ("Y", "Khách ghi chú — khách điền"),
    ]
    for col, desc in col_map:
        ws.append([col, desc])
        for c in range(1, 3):
            body_cell(ws.cell(row=ws.max_row, column=c))

    ws.append([])
    ws.append(["Bảng 2 — Pattern công thức chuẩn", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = SUB_FILL
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    patterns = [
        ("Avg VFR mode GT", f_avg_vfr("GT")),
        ("Tổng chuyến GT", f_total("GT")),
        ("Low <50% (GT)", f_bucket("GT", None, 50)),
        ("Medium 50-70% (GT)", f_bucket("GT", 50, 70)),
        ("High 70-95% (GT)", f_bucket("GT", 70, 95)),
        ("Excellent ≥95% (GT)", f_bucket("GT", 95, None)),
        ("Δ GT − VH Avg VFR", '=ROUND(AVERAGEIFS(\'04 — Detail Trips\'!V:V,\'04 — Detail Trips\'!A:A,"GT")'
                              '-AVERAGEIFS(\'04 — Detail Trips\'!V:V,\'04 — Detail Trips\'!A:A,"VH"),2)'),
        ("VFR simple-avg theo 1 khu vực GT ($A5)", f_dim_avg("GT", "G", "$A5")),
        ("VFR simple-avg theo 1 NVC VH ($A5)", f_dim_avg("VH", "H", "$A5")),
        ("Avg VFR theo 1 ngày ETA (GT)", f_trend_avg("GT", WINDOW_START_DATE)),
        ("Tổng chuyến >100% (over-registered, GT)", f'=COUNTIFS({DETAIL_SHEET}!A:A,"GT",{DETAIL_SHEET}!V:V,">100")'),
        ("Tổng chuyến Loose (GT)", f'=COUNTIFS({DETAIL_SHEET}!A:A,"GT",{DETAIL_SHEET}!K:K,"Loose")'),
    ]
    for label, formula in patterns:
        ws.append([label, formula])
        r_ = ws.max_row
        body_cell(ws.cell(row=r_, column=1))
        c = ws.cell(row=r_, column=2)
        c.font = Font(name="Consolas", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        ws.row_dimensions[r_].height = max(20, min(70, 16 + len(formula) // 80 * 14))

    ws.append([])
    ws.append(["Cảnh báo", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="DC2626")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    warnings = [
        ("Weighted không tái tạo", "Dimension VFR sheet 02 KHÔNG verify được bằng AVERAGEIFS. Để audit weighted phải dùng SQL ở "
                                   "sheet 07 (CTE base→calc). Simple-avg chỉ là chặn trên/dưới."),
        ("Mode column bắt buộc", "Mọi formula KPI/trend phải kèm tiêu chí A:A=\"GT\" hoặc \"VH\" — sheet 04 trộn cả 2 mode. Bỏ "
                                 "tiêu chí Mode → số gộp GT+VH sai."),
        ("Sheet rename", "Đổi tên sheet 04 → công thức 01/02/03 broken. Find&Replace '04 — Detail Trips' nếu phải đổi."),
        ("Recalculation", "Mở file: Update external links = KHÔNG. Công thức không refresh → F9 / Ctrl+Alt+F9."),
    ]
    for k, v in warnings:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP
        for c in range(1, 3):
            body_cell(ws.cell(row=ws.max_row, column=c))


UX_CHECKLIST_ITEMS = [
    # (category, what_to_check, expected, severity, ref)
    # === UX Visual ===
    ("UX Visual", "Hero card hiển thị Avg VFR (mode đang xem) + RAG dot + gap-vs-target 85% + sparkline",
     "Số Avg VFR to readable; RAG dot (🟢≥85/🟡75–<85/🔴<75); gap ±Xpt vs 85; sparkline visible.", "Major", "AC-S2"),
    ("UX Visual", "Hero LUÔN hiển thị GT% + VH% + Δ GT−VH bất kể đang toggle mode nào",
     "Cả 3 con số (GT, VH, Δ) visible đồng thời; toggle mode KHÔNG ẩn mất 1 trong 2.", "Major", "AC-S2"),
    ("UX Visual", "Sparkline trong hero — xác nhận cửa sổ thời gian (code = 14 ngày, KHÔNG phải 13 tuần)",
     "Sparkline hiển thị 14 ngày gần nhất (VFR_SPARKLINE_DAYS=14). ⚠ PRD §18 ghi '13w' — drift; lấy 14d làm chuẩn.", "Minor", "AC-S2 (drift)"),
    ("UX Visual", "4 bucket chip (Low/Medium/High/Excellent) dùng RAG color, KHÔNG phải 5 màu rời; Avg gộp vào hero",
     "4 chip: Low🔴 / Medium🟡 / High🟢 / Excellent🟢đậm. Không còn card Avg đứng riêng.", "Minor", "AC-S7"),
    ("UX Visual", "3 bar chart (By Area/Vehicle/Vendor) có ReferenceLine dashed tại target 85 + bar tô RAG",
     "Mỗi bar chart có đường target 85 (dashed) + bar màu theo band. By Loading Type×Time KHÔNG có target line (theo code).", "Major", "AC-S4"),
    ("UX Visual", "Action title chart nói insight (worst dim / gap) thay vì label tĩnh 'VFR theo Khu vực'",
     "Title chứa tên dim xấu nhất + magnitude + gap to 85. Empty data → fallback label tĩnh.", "Major", "AC-S1"),
    ("UX Visual", "Time × Area table tô RAG theo target 85 + có hàng/cột trung bình",
     "Cell màu theo band; average row + column giữ màu sky.", "Minor", "AC-S4"),
    # === Filter ===
    ("Filter", "Mode toggle GT ↔ VH — đổi KPI/charts/table/detail; default = GT (tender)",
     "Click toggle đổi toàn bộ section sang mode kia; mở widget default = GT.", "Major", "Spec §3.1"),
    ("Filter", "Filter Kho / Khu vực / Vendor / Loại xe — dropdown đủ option + ALL",
     "Mỗi dropdown đủ giá trị thật từ MV; chọn ALL = không lọc.", "Major", "PRD §4.5"),
    ("Filter", "Loại xe Tender disabled khi mode=Operation và ngược lại (conditional disable)",
     "vehicle_type_tender mờ/disabled ở mode VH; vehicle_type_ops mờ ở mode GT.", "Minor", "PRD §4.5"),
    ("Filter", "Loại ngày — 2 option ETA / ATA; default ETA",
     "Đổi ETA→ATA re-fetch toàn widget với cột ngày tương ứng.", "Minor", "Spec §5.2"),
    ("Filter", "Date range cap 12 tháng — chọn sớm hơn today−365d → clamp + toast warning",
     "from_date bị kéo về today−365d, toast cảnh báo (getVfrHistoryFloor / VFR_HISTORY_CAP_DAYS=365).", "Major", "PRD §5.3 / BUG-VFR-05"),
    ("Filter", "Date range default = đầu tháng → hôm nay",
     "Mở widget lần đầu: from = startOfMonth(today), to = today.", "Minor", "PRD §4.5"),
    ("Filter", "Multi-select 2+ value 1 dim (vd Khu vực = A + B) → IN logic đúng",
     "Số dashboard = SQL ...IN ('A','B'); verify với tổng dimension.", "Major", "PRD §4.5"),
    ("Filter", "Filter combo 5-dim response time < 3s; page load < 5s",
     "Trace Network tab; >3s flag perf defect (12 query song song).", "Minor", "Spec §15"),
    # === Interaction ===
    ("Interaction", "Exception panel <50% — count chuyến + window 7 ngày + top 3 chip area/vendor/vehicle",
     "Panel hiện N chuyến <50% trong 7d + 3 dim chip mỗi loại (sort count DESC).", "Major", "AC-S3"),
    ("Interaction", "Exception CTA 'Xem N chuyến →' — 1 click sang Detail tab + pre-filter vfr_max<50 + range 7d",
     "Click → chuyển Detail, áp filter vfr_max<50 và date 7 ngày, không cần thao tác thêm.", "Major", "AC-S3"),
    ("Interaction", "Exception empty state — badge xanh khi 0 chuyến <50% trong 7 ngày",
     "Hiện 'Hệ thống ổn — không có chuyến <50%' thay vì panel trống.", "Minor", "AC-S3"),
    ("Interaction", "By Vendor chart — nếu admin CHƯA cấu hình SQL → empty state, không crash",
     "Chart 'By Vendor' render khi có data; nếu vendorBarRows rỗng → placeholder/empty state (gated).", "Major", "AC-S5"),
    ("Interaction", "By Loading Type×Time — toggle group day/week/month đổi grouping, KHÔNG gọi lại API",
     "Toggle dùng data đã fetch; không có request Network mới.", "Minor", "Spec §8.3"),
    ("Interaction", "Detail grid 30 cột — sort theo column; pagination 10/20/50; export all rows",
     "Click header sort; pager đổi page size; Export all chạy được. ⚠ Code = 30 cột (PRD/spec ghi 29 — drift).", "Minor", "Spec §9 (drift)"),
    ("Interaction", "4 persona preset (BOD / Ops / Carrier / Planning) render đúng section theo ma trận §18.3",
     "Chọn preset → chỉ hiện section khai báo (visibleSections). Default = Planning (full).", "Major", "AC-S6"),
    # === Storytelling ===
    ("Storytelling", "Q1 'VFR tổng có thấp không?' trả lời ≤ 2s từ hero Avg VFR + RAG",
     "Avg VFR + màu RAG visible ngay, không scroll.", "Major", "AC-S2"),
    ("Storytelling", "Q2 'Kế hoạch hay thực thi kém?' trả lời ≤ 10s từ GT vs VH + Δ",
     "Khách đọc được GT cao/thấp vs VH; Δ chỉ rõ lệch.", "Major", "AC-S2"),
    ("Storytelling", "Q3 'Chiều nào kéo VFR xuống?' trả lời ≤ 20s từ dimension charts worst-first",
     "Bar sorted ascending; khách chỉ đúng dim/giá trị non-tải nhất.", "Major", "AC-S1"),
    ("Storytelling", "Q4 'Có chuyến nguy hiểm <50% không? Bao nhiêu?' từ Exception panel ≤ 20s",
     "Count + top dim chips trả lời ngay; CTA drill xem chi tiết.", "Major", "AC-S3"),
    ("Storytelling", "Q5 'Xu hướng VFR đang lên hay xuống?' từ sparkline / trend ≤ 30s",
     "Sparkline 14d cho thấy hướng; (lưu ý 14d không phải 13w).", "Minor", "AC-S2"),
    # === Master Data ===
    ("Master Data", "Tên NVC (nha_van_tai) hiển thị đầy đủ, tiếng Việt có dấu, không phải code",
     "Vd 'NGUYEN PHAT' / 'NINJAVAN' hiển thị tên, không phải mã raw.", "Minor", "Master data"),
    ("Master Data", "Loại bốc xếp — Loose vs Full Pallet phân biệt rõ; 'Losse' (typo) gom vào Loose",
     "Chart/table phân biệt Loose / Full Pallet; record 'Losse' không tách thành nhóm riêng.", "Minor", "Spec §22 / registry"),
    ("Master Data", "Chuyến thiếu loại bốc xếp → gom nhóm '(rỗng)' / 'Không xác định', không drop âm thầm",
     "loai_boc_xep rỗng vẫn xuất hiện trong nhóm rõ ràng.", "Minor", "BUG-VFR-04"),
    ("Master Data", "VFR > 100% — hiển thị raw + cảnh báo icon (không clamp), vẫn đếm Excellent",
     "Trip vfr_max 142% hiện 142% + AlertTriangle; nằm trong bucket ≥95.", "Minor", "BUG-VFR-01"),
    ("Master Data", "% / số format — '94.49%' (2 thập phân) · count '1,234' · CBM '1,326.50'",
     "% có 2 decimal + ký hiệu %; count có dấu phẩy ngàn.", "Cosmetic", "i18n"),
]


def build_ux_checklist_sheet(ws):
    ws.title = "06 — UX & Filter checklist"
    set_widths(ws, [6, 18, 50, 50, 14, 12, 18, 28])
    title_row(ws, "UX / Filter / Interaction / Storytelling / Master Data — khách fill trong session (KHÔNG compute từ data)", 8)
    header_row(ws, ["#", "Category", "Mục cần kiểm tra (Quan sát thế nào?)", "Expected behavior",
                    "Customer\nverdict", "Severity\nnếu Fail", "Reference\n(AC / PRD)", "Customer ghi chú"])
    category_fill = {
        "UX Visual": PatternFill("solid", fgColor="DBEAFE"),
        "Filter": PatternFill("solid", fgColor="FEF3C7"),
        "Interaction": PatternFill("solid", fgColor="E0E7FF"),
        "Storytelling": PatternFill("solid", fgColor="FCE7F3"),
        "Master Data": PatternFill("solid", fgColor="DCFCE7"),
    }
    severity_fill = {"Critical": RED, "Major": YELLOW, "Minor": GREEN, "Cosmetic": GREY}
    for i, (cat, check, expected, severity, ref) in enumerate(UX_CHECKLIST_ITEMS, 1):
        ws.append([i, cat, check, expected, "", severity, ref, ""])
        r_ = ws.max_row
        for col in range(1, 9):
            body_cell(ws.cell(row=r_, column=col))
            ws.cell(row=r_, column=col).alignment = WRAP
        ws.cell(row=r_, column=2).fill = category_fill.get(cat, GREY)
        ws.cell(row=r_, column=2).font = Font(bold=True)
        ws.cell(row=r_, column=5).fill = EMPTY_CUSTOMER
        ws.cell(row=r_, column=5).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r_, column=6).fill = severity_fill.get(severity, GREY)
        ws.cell(row=r_, column=6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_, column=8).fill = EMPTY_CUSTOMER
        max_text = max(len(check), len(expected))
        ws.row_dimensions[r_].height = max(34, min(96, 20 + max_text // 50 * 16))
    ws.freeze_panes = "A3"
    ws.append([])
    ws.append(["Hướng dẫn", "Khách điền cột Verdict (Pass / Fail / N/A / cần làm rõ) + ghi chú. Sheet này tập trung lớp C "
                            "(UX & Storytelling) + D (Filter/Perf), bổ sung cho sheet 01-04 (lớp A Data + B Logic)."])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)


def build_sql_sheet(ws, sql_map):
    ws.title = "07 — SQL Appendix"
    set_widths(ws, [44, 34, 110])
    title_row(ws, "SQL queries — provenance rõ ràng (canonical từ sql-registry.md OR ad-hoc)", 3)
    header_row(ws, ["Query ID — mục đích", "Source / provenance", "SQL (đã substitute placeholder)"])
    for qid, (q, src) in sql_map.items():
        ws.append([qid, src, q])
        r_ = ws.max_row
        ws.cell(row=r_, column=1).font = Font(bold=True)
        ws.cell(row=r_, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        src_cell = ws.cell(row=r_, column=2)
        src_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if src.startswith("sql-registry.md"):
            src_cell.fill = GREEN
            src_cell.font = Font(bold=True)
        else:
            src_cell.fill = YELLOW
        for col in (1, 2, 3):
            ws.cell(row=r_, column=col).border = BORDER
        sql_cell = ws.cell(row=r_, column=3)
        sql_cell.font = Font(name="Consolas", size=9)
        sql_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_].height = max(40, min(460, 18 + q.count("\n") * 12))


# ── Main ─────────────────────────────────────────────────
def main():
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("[*] Connect CH…")
    client = get_client()
    ctx = {
        "run_ts": datetime.now(timezone(timedelta(hours=7))).strftime("%Y-%m-%d %H:%M UTC+7"),
        "server_now": client.query("SELECT now('Asia/Ho_Chi_Minh')").result_rows[0][0],
    }
    print("[*] Query membership lists for ALL bypass…")
    filters = query_filter_values(client)

    sql_map = {}
    ctx["kpi"] = {}
    ctx["trend"] = {}
    ctx["detail"] = []
    ctx["dim"] = {"area": {}, "vehicle": {}, "vendor": {}, "loading_type": {}}

    for mode in ("GT", "VH"):
        print(f"[*] Mode {mode}…")
        kpi, sql_k, src_k = run_kpi(client, mode, filters)
        ctx["kpi"][mode] = kpi
        sql_map[f"Q-KPI-{mode} — Avg VFR + 4 bucket ({mode})"] = (sql_k, src_k)
        print(f"    KPI {mode}: avg={kpi['avg_vfr']} buckets L/M/H/E={kpi['low']}/{kpi['medium']}/{kpi['high']}/{kpi['excellent']} total={kpi['total']}")

        area, sql_a, src_a = run_registry_dim(client, mode, filters, "area")
        ctx["dim"]["area"][mode] = area
        sql_map[f"Q-AREA-{mode} — VFR theo khu vực ({mode})"] = (sql_a, src_a)

        veh, sql_v, src_v = run_registry_dim(client, mode, filters, "vehicle")
        ctx["dim"]["vehicle"][mode] = veh
        sql_map[f"Q-VEHICLE-{mode} — VFR theo loại xe ({mode})"] = (sql_v, src_v)

        ven, sql_vn = run_adhoc_dim(client, mode, filters, "nha_van_tai")
        ctx["dim"]["vendor"][mode] = ven
        sql_map[f"Q-VENDOR-{mode} — VFR theo NVC ({mode})"] = (
            sql_vn, "ad-hoc (KHÔNG có trong registry; weighted formula = registry byArea, GROUP BY nha_van_tai; "
                    "tham chiếu vfr-spec.md §22)")

        lt, sql_lt = run_adhoc_dim(client, mode, filters, "loai_boc_xep")
        ctx["dim"]["loading_type"][mode] = lt
        sql_map[f"Q-LOADTYPE-{mode} — VFR theo loại bốc xếp ({mode})"] = (
            sql_lt, "ad-hoc (registry byLoadingType là day-level; đây rollup window-level dùng weighted formula registry byArea, "
                    "GROUP BY loai_boc_xep)")

        trend, sql_t = run_trend(client, mode, filters)
        ctx["trend"][mode] = trend
        sql_map[f"Q-TREND-{mode} — Avg VFR daily ({mode})"] = (
            sql_t, "ad-hoc (simple avg(vfr_max) per ngày; registry không có section trend avg-VFR thuần)")

        detail, sql_d = run_detail(client, mode, filters)
        ctx["detail"].extend(detail)
        sql_map[f"Q-DETAIL-{mode} — Detail trips ({mode})"] = (
            sql_d, "ad-hoc (bám pattern report registry + membership filter; thêm cột vfr_max + bucket)")
        print(f"    Detail {mode}: {len(detail):,} rows")

    print("[*] Build workbook…")
    wb = Workbook()
    build_readme(wb.active, ctx)
    build_kpi_sheet(wb.create_sheet(), ctx)
    build_dim_sheet(wb.create_sheet(), ctx)
    build_trend_sheet(wb.create_sheet(), ctx)
    build_detail_sheet(wb.create_sheet(), ctx)
    build_formula_guide_sheet(wb.create_sheet(), ctx)
    build_ux_checklist_sheet(wb.create_sheet())
    build_sql_sheet(wb.create_sheet(), sql_map)

    wb.save(OUT_FILE)
    print(f"[OK] Wrote {OUT_FILE}  ({len(ctx['detail']):,} detail trips)")


if __name__ == "__main__":
    main()
