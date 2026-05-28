"""
load_tms_report_to_ch.py
────────────────────────
Nạp dữ liệu report TMS #25 (Báo cáo theo đơn hàng và chuyến) vào ClickHouse
`analytics_workspace.tms_report_25_trip_order`.

Chiến lược:
  - Chia khoảng ngày thành cửa sổ <=5 ngày, mỗi cửa sổ gọi nguồn 1 lần
    (nghỉ giữa các lần để nhẹ cho hệ thống nguồn STM).
  - Phase 1: nạp raw 144 cột String vào bảng `_stg` (full refresh).
  - Phase 2: dựng bảng cuối = raw + 4 cột CSE (kh_cse/pick_cse/giao_cse/chenh_cse)
    join từ mv_otif_swm_stm_data theo OrderCode = `Mã đơn hàng`. CSE đã quy đổi UOM
    (CSE/PCE/PALLET) qua mv_masterdata_sku — KHÔNG sum QuantityBBGN raw (lẫn đơn vị).
  - 144 cột String (trung thực theo file export) + COMMENT tiếng Việt từng cột;
    2 cột meta: _src_window, _loaded_at.

Schema canonical: knowledge-base/tms/reports/25-trip-and-order/25-tms-trip-and-order.columns.json

Cách chạy:
    python projects/mondelez/scripts/load_tms_report_to_ch.py
    python projects/mondelez/scripts/load_tms_report_to_ch.py --from 2026-05-01 --to 2026-05-26
    python projects/mondelez/scripts/load_tms_report_to_ch.py --recreate     # DROP + tạo lại (khi đổi schema/comment)

Env (projects/mondelez/.env): TMS_* + CLICKHOUSE_*
"""

import os
import sys
import json
import time
import argparse
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
import clickhouse_connect

_TENANT_DIR = Path(__file__).resolve().parent.parent       # projects/mondelez/
_PROJECTS_DIR = _TENANT_DIR.parent                          # projects/
_ENV = _TENANT_DIR / ".env"
if _ENV.exists():
    load_dotenv(_ENV)
else:
    load_dotenv()

_REPORT_DIR = _PROJECTS_DIR / "knowledge-base" / "tms" / "reports" / "25-trip-and-order"
COLUMNS_JSON = _REPORT_DIR / "25-tms-trip-and-order.columns.json"
REQUEST_JSON = _REPORT_DIR / "samples" / "request.json"

CH_DATABASE = "analytics_workspace"
CH_TABLE = "mdlz_tms_report_25_trip_order"
TABLE_COMMENT = (
    "MDLZ TMS report #25 (Bao cao theo don hang va chuyen) - staging dump String trung thuc. "
    "Refresh: TRUNCATE+reload theo cua so <=5 ngay. "
    "Nguon: REPDIOPSPlan_SettingDownload functionid=78 TypeExport=25. "
    "Doc: mondelez/02-data/data-sources/mdlz_tms_report_25_trip_order.md"
)
FINAL_TABLE_COMMENT = (
    "MDLZ TMS report #25 + CSE enrich. 144 cot String raw tu export + "
    "kh_cse/pick_cse/giao_cse/chenh_cse (muc DON, join mv_otif_swm_stm_data theo OrderCode='Ma don hang'; "
    "CSE da quy doi UOM qua mv_masterdata_sku). CSE la grain DON: nhieu dong cung OrderCode = cung gia tri "
    "-> dedupe theo OrderCode khi sum. ~77% don khop (con lai ngoai scope OTIF: tra ve/dich vu khac -> NULL). "
    "Refresh: TRUNCATE+reload theo cua so <=5 ngay."
)
ENDPOINT = "REPDIOPSPlan_SettingDownload"
FUNCTIONID = "78"
CHUNK_DAYS = 5
SLEEP_BETWEEN = 3.0   # giây nghỉ giữa các lần gọi nguồn
DEFAULT_FROM = "2026-05-01"
DEFAULT_TO = "2026-05-26"

CH_TABLE_STG = CH_TABLE + "_stg"
CSE_SOURCE = "mv_otif_swm_stm_data"   # canonical line-level CSE, đã quy đổi UOM qua mv_masterdata_sku
CSE_COLUMNS = [
    ("kh_cse", "Kế hoạch đặt (CSE) — Σ ORIGINAL CSE theo Mã đơn hàng"),
    ("pick_cse", "Kho đã pick (CSE) — Σ SHIPPED CSE"),
    ("giao_cse", "Giao nhận thực BBGN (CSE) — Σ Sản lượng giao CSE"),
    ("chenh_cse", "Chênh lệch giao nhận vs kế hoạch (giao_cse − kh_cse); âm = giao thiếu"),
]

AUTH_URL = os.getenv("TMS_AUTH_URL")
REPORT_HOST = (os.getenv("TMS_REPORT_HOST") or "").rstrip("/")
TENANT_HOST = os.getenv("TMS_TENANT_HOST")
CLIENT_ID = os.getenv("TMS_CLIENT_ID")
SCOPE = os.getenv("TMS_SCOPE", "openid profile email address phone role Auth offline_access")
USERNAME = os.getenv("TMS_USERNAME")
PASSWORD = os.getenv("TMS_PASSWORD")


def validate_env() -> None:
    required = {
        "TMS_AUTH_URL": AUTH_URL, "TMS_REPORT_HOST": REPORT_HOST,
        "TMS_TENANT_HOST": TENANT_HOST, "TMS_CLIENT_ID": CLIENT_ID,
        "TMS_USERNAME": USERNAME, "TMS_PASSWORD": PASSWORD,
        "CLICKHOUSE_HOST": os.getenv("CLICKHOUSE_HOST"),
        "CLICKHOUSE_USER": os.getenv("CLICKHOUSE_USER"),
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        print(f"[ERROR] Thiếu env trong {_ENV}: {', '.join(missing)}")
        sys.exit(1)


def load_columns() -> list[dict]:
    return json.loads(COLUMNS_JSON.read_text(encoding="utf-8"))


def ch_client():
    return clickhouse_connect.get_client(
        host=os.getenv("CLICKHOUSE_HOST", ""),
        port=int(os.getenv("CLICKHOUSE_PORT", "8443")),
        username=os.getenv("CLICKHOUSE_USER", ""),
        password=os.getenv("CLICKHOUSE_PASSWORD", ""),
        secure=os.getenv("CLICKHOUSE_SECURE", "true").lower() not in ("false", "0", "no"),
        connect_timeout=30,
    )


def _esc(s: str) -> str:
    return s.replace(chr(39), chr(39) * 2)


def build_ddl(columns: list[dict], table: str, with_cse: bool = False, comment: str = TABLE_COMMENT) -> str:
    defs = [f"  `{c['code']}` String COMMENT '{_esc(c['label'])}'" for c in columns]
    defs.append("  `_src_window` String COMMENT 'Cửa sổ ngày nguồn (yyyy-mm-dd..yyyy-mm-dd)'")
    defs.append("  `_loaded_at` DateTime DEFAULT now() COMMENT 'Thời điểm nạp dữ liệu'")
    if with_cse:
        defs += [f"  `{code}` Nullable(Float64) COMMENT '{_esc(cmt)}'" for code, cmt in CSE_COLUMNS]
    return (
        f"CREATE TABLE IF NOT EXISTS `{CH_DATABASE}`.`{table}` (\n"
        + ",\n".join(defs)
        + "\n)\nENGINE = MergeTree\nORDER BY (`MasterCode`, `OrderCode`)\n"
        + f"COMMENT '{_esc(comment)}'"
    )


def build_enrich_sql(columns: list[dict]) -> str:
    raw = [f"`{c['code']}`" for c in columns] + ["`_src_window`", "`_loaded_at`"]
    insert_cols = ", ".join(raw + [code for code, _ in CSE_COLUMNS])
    select_raw = ", ".join(f"s.{x}" for x in raw)
    return f"""
INSERT INTO `{CH_DATABASE}`.`{CH_TABLE}` ({insert_cols})
SELECT {select_raw},
       agg.kh_cse, agg.pick_cse, agg.giao_cse, (agg.giao_cse - agg.kh_cse) AS chenh_cse
FROM `{CH_DATABASE}`.`{CH_TABLE_STG}` AS s
LEFT JOIN (
    SELECT `Mã đơn hàng` AS oc,
           round(sum(ifNull(`ORIGINAL CSE`, 0)), 2) AS kh_cse,
           round(sum(ifNull(`SHIPPED CSE`, 0)), 2) AS pick_cse,
           round(sum(ifNull(`Sản lượng giao CSE`, 0)), 2) AS giao_cse
    FROM `{CH_DATABASE}`.`{CSE_SOURCE}`
    WHERE `Mã đơn hàng` IS NOT NULL AND `Mã đơn hàng` != ''
    GROUP BY `Mã đơn hàng`
) AS agg ON s.OrderCode = agg.oc
SETTINGS join_use_nulls = 1
""".strip()


def get_token() -> str:
    print("[INFO] Lấy access token (password grant)...")
    resp = requests.post(
        AUTH_URL,
        data={"grant_type": "password", "username": USERNAME, "password": PASSWORD,
              "client_id": CLIENT_ID, "scope": SCOPE},
        headers={"content-type": "application/x-www-form-urlencoded", "d": TENANT_HOST},
        timeout=60,
    )
    if resp.status_code != 200:
        print(f"[ERROR] Token thất bại {resp.status_code}: {resp.text[:300]}")
        sys.exit(1)
    token = resp.json().get("access_token")
    if not token:
        print(f"[ERROR] Response không có access_token: {resp.text[:300]}")
        sys.exit(1)
    print("[OK] Đã có token.")
    return token


def fetch_report_url(body: dict, token: str) -> str:
    url = f"{REPORT_HOST}/api/REP/{ENDPOINT}"
    resp = requests.post(
        url, json=body,
        headers={
            "authorization": f"Bearer {token}",
            "content-type": "application/json; charset=UTF-8",
            "d": TENANT_HOST, "functionid": FUNCTIONID,
            "origin": f"https://{TENANT_HOST}", "referer": f"https://{TENANT_HOST}/",
        },
        timeout=300,
    )
    if resp.status_code != 200:
        print(f"[ERROR] Export thất bại {resp.status_code}: {resp.text[:400]}")
        sys.exit(1)
    try:
        file_url = resp.json()
    except ValueError:
        file_url = resp.text.strip().strip('"')
    if not isinstance(file_url, str) or not file_url.startswith("http"):
        print(f"[ERROR] Response không phải URL: {str(file_url)[:300]}")
        sys.exit(1)
    return file_url


# Cửa sổ local [S, E] bao gồm 2 đầu: 17:00Z = 00:00 +07
def to_dtfrom(d: datetime) -> str:
    return (d - timedelta(days=1)).strftime("%Y-%m-%dT17:00:00.000Z")


def to_dtto(d: datetime) -> str:
    return d.strftime("%Y-%m-%dT17:00:00.000Z")


def date_windows(start: datetime, end: datetime, days: int):
    cur = start
    while cur <= end:
        w_end = min(cur + timedelta(days=days - 1), end)
        yield cur, w_end
        cur = w_end + timedelta(days=1)


def fetch_window_rows(body_template: dict, token: str, s: datetime, e: datetime, ncols: int) -> list[list]:
    body = json.loads(json.dumps(body_template))  # deep copy
    body["dtfrom"] = to_dtfrom(s)
    body["dtto"] = to_dtto(e)
    file_url = fetch_report_url(body, token)
    content = requests.get(file_url, timeout=300).content
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    next(it, None)  # bỏ header
    rows = []
    for r in it:
        vals = [("" if v is None else str(v)) for v in r]
        vals = (vals + [""] * ncols)[:ncols]
        rows.append(vals)
    wb.close()
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Nạp report TMS #25 vào ClickHouse.")
    parser.add_argument("--from", dest="dfrom", default=DEFAULT_FROM, help="Ngày local bắt đầu YYYY-MM-DD")
    parser.add_argument("--to", dest="dto", default=DEFAULT_TO, help="Ngày local kết thúc YYYY-MM-DD")
    parser.add_argument("--recreate", action="store_true", help="DROP + tạo lại bảng (khi đổi schema/comment)")
    parser.add_argument("--sleep", type=float, default=SLEEP_BETWEEN, help="Giây nghỉ giữa các lần gọi nguồn")
    args = parser.parse_args()

    print("=" * 60)
    print("load_tms_report_to_ch.py — MDLZ Control Tower")
    print("=" * 60)
    validate_env()

    columns = load_columns()
    codes = [c["code"] for c in columns]
    ncols = len(codes)
    start = datetime.strptime(args.dfrom, "%Y-%m-%d")
    end = datetime.strptime(args.dto, "%Y-%m-%d")
    windows = list(date_windows(start, end, CHUNK_DAYS))
    print(f"[INFO] Khoảng: {args.dfrom} → {args.dto} | {ncols} cột | {len(windows)} cửa sổ ≤{CHUNK_DAYS} ngày")

    client = ch_client()
    print(f"[INFO] ClickHouse: {os.getenv('CLICKHOUSE_HOST')} / {CH_DATABASE}.{CH_TABLE}")

    # ── Phase 1: nạp raw 144 cột vào bảng staging ──
    if args.recreate:
        print("[INFO] DROP staging (recreate)...")
        client.command(f"DROP TABLE IF EXISTS `{CH_DATABASE}`.`{CH_TABLE_STG}`")
    client.command(build_ddl(columns, CH_TABLE_STG, comment="Staging raw report#25 (String) cho enrich CSE."))
    print("[INFO] TRUNCATE staging (full refresh)...")
    client.command(f"TRUNCATE TABLE IF EXISTS `{CH_DATABASE}`.`{CH_TABLE_STG}`")

    token = get_token()
    body_template = json.loads(REQUEST_JSON.read_text(encoding="utf-8"))
    insert_cols = codes + ["_src_window"]
    total = 0
    for i, (s, e) in enumerate(windows):
        if i > 0 and args.sleep > 0:
            time.sleep(args.sleep)
        win = f"{s:%Y-%m-%d}..{e:%Y-%m-%d}"
        print(f"[INFO] [{i + 1}/{len(windows)}] cửa sổ {win} ...")
        rows = fetch_window_rows(body_template, token, s, e, ncols)
        data = [r + [win] for r in rows]
        if data:
            client.insert(CH_TABLE_STG, data, column_names=insert_cols, database=CH_DATABASE)
        total += len(rows)
        print(f"        ↳ {len(rows):,} dòng")

    # ── Phase 2: bảng cuối = raw + CSE (join mv_otif_swm_stm_data theo OrderCode = `Mã đơn hàng`) ──
    print(f"[INFO] Enrich CSE từ {CSE_SOURCE} (join OrderCode = `Mã đơn hàng`)...")
    client.command(f"DROP TABLE IF EXISTS `{CH_DATABASE}`.`{CH_TABLE}`")
    client.command(build_ddl(columns, CH_TABLE, with_cse=True, comment=FINAL_TABLE_COMMENT))
    client.command(build_enrich_sql(columns))
    client.command(f"DROP TABLE IF EXISTS `{CH_DATABASE}`.`{CH_TABLE_STG}`")

    final = client.query(f"SELECT count() FROM `{CH_DATABASE}`.`{CH_TABLE}`").first_row[0]
    matched = client.query(f"SELECT countIf(kh_cse IS NOT NULL) FROM `{CH_DATABASE}`.`{CH_TABLE}`").first_row[0]
    pct = (100 * matched / final) if final else 0
    print(f"\n[OK] Nạp xong: {total:,} dòng raw → bảng cuối {final:,} dòng; "
          f"CSE khớp {matched:,} ({pct:.1f}%). Phần còn lại ngoài scope OTIF (trả về / dịch vụ khác) → CSE NULL.")


if __name__ == "__main__":
    main()
