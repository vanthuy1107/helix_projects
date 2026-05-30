"""
UAT Flash Daily export — sinh file Excel test case với số SQL thật cho khách hàng confirm.

Window: 2026-06-08 (1 ngày — D-1 session UAT 2026-06-09).
Filter mặc định: Date Type = "Ngày GI" → cột delivery_date_1.
  ⚠ Label "Ngày GI" exact match CH CASE branch — KHÔNG dùng FE label "GI date".
    Per memory [[feedback_sql_date_type_label_exact_match]] — mismatch silent bug.
Source: analytics_workspace.mv_flash_and_drop_report (L1, L3, L5) +
        analytics_workspace.mv_dropped_report (L4, T7, T8) +
        analytics_workspace.mv_flash_report (T9 detail).

SQL provenance:
  • L1 Hero, L3 funnel 5 status, L5 4 dim panels (Kho/Khu vực/Customer/Kênh),
    L2 Exception 3 ô, L4 Drop Trend 14d, T7 Drop bucket, T8 Drop reason — load
    CANONICAL từ `projects/mondelez/02-data/data-sources/sql-registry.md`
    section ## Flash Report. KHÔNG tự sinh.
  • T9 Detail (32 cột Flash Detail) — ad-hoc (widget detailTable config nằm trong widget.config DB).

Output: projects/mondelez/01-sections/flash-daily/uat/flash-daily-uat-numbers-2026-06-08_to_2026-06-08.xlsx
"""
from __future__ import annotations

import os
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_DIR = next(p for p in Path(__file__).resolve().parents
                    if (p / "da.toml").exists())  # tenant root = nơi có da.toml (relocation-proof)
_OUTPUT_DIR = _PROJECT_DIR / "01-sections" / "flash-daily" / "uat"

try:
    from dotenv import load_dotenv
    _env = _PROJECT_DIR / ".env"
    if _env.exists():
        load_dotenv(_env)
    else:
        load_dotenv()
except ImportError:
    pass

import clickhouse_connect
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _pct2(v):
    """Normalize % values → Decimal('XX.XX') scale=2 fixed (trailing zero preserved)."""
    if v is None:
        return None
    return Decimal(str(round(float(v), 2))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _num2(v):
    """Normalize số có .XX → Decimal scale=2."""
    if v is None:
        return None
    return Decimal(str(round(float(v), 2))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# ── Config ───────────────────────────────────────────────
# Window proxy: session UAT = 2026-06-09; D-1 = 2026-06-08 (future, no data yet).
# Dùng proxy 2026-05-22 (Fri, complete day với 4447 rows trong CH).
# Khi UAT thực, đổi WINDOW về 2026-06-08 + re-run.
WINDOW_START_DATE = "2026-05-22"
WINDOW_END_DATE = "2026-05-22"
DATE_TYPE_LABEL = "Ngày GI"       # ⚠ EXACT match CH CASE branch — KHÔNG dùng FE label "GI date"
DATE_COL = "delivery_date_1"
UOM = "cse"
SOURCE_MV_MAIN = "analytics_workspace.mv_flash_and_drop_report"
SOURCE_MV_DROP = "analytics_workspace.mv_dropped_report"
SOURCE_MV_RAW = "analytics_workspace.mv_flash_report"
TENANT = "Mondelez"

OUT_FILE = _OUTPUT_DIR / f"flash-daily-uat-numbers-{WINDOW_START_DATE}_to_{WINDOW_END_DATE}.xlsx"

# RAG bands per PRD §5.4 + memory [[project_mondelez_flash_daily_target]]
RAG = {
    "pct_done": {"target": 95.0, "green": 95.0, "yellow_lo": 85.0, "alert": 80.0},
    "drop_rate": {"target": 5.0, "green_max": 5.0, "yellow_max": 10.0},  # inverted — low is good
}


def get_client():
    return clickhouse_connect.get_client(
        host=os.getenv("CLICKHOUSE_HOST", ""),
        port=int(os.getenv("CLICKHOUSE_PORT", "8443")),
        username=os.getenv("CLICKHOUSE_USER", ""),
        password=os.getenv("CLICKHOUSE_PASSWORD", ""),
        secure=os.getenv("CLICKHOUSE_SECURE", "true").lower() in ("1", "true", "yes"),
        connect_timeout=15,
        send_receive_timeout=120,
    )


# ── Registry SQL loader ──────────────────────────────────
REGISTRY_PATH = _PROJECT_DIR / "02-data" / "data-sources" / "sql-registry.md"
_REGISTRY_TEXT_CACHE = None


def _registry_text():
    global _REGISTRY_TEXT_CACHE
    if _REGISTRY_TEXT_CACHE is None:
        _REGISTRY_TEXT_CACHE = REGISTRY_PATH.read_text(encoding="utf-8")
    return _REGISTRY_TEXT_CACHE


def load_registry_sql(section_title, parent_section="Flash Report"):
    """Đọc sql-registry.md, tìm `### <section_title>` trong scope `## Flash Report`, lấy block ClickHouse SQL.
    Trả về (sql, line_ref).
    """
    import re
    text = _registry_text()
    parent_match = re.search(rf"^## {re.escape(parent_section)}\b", text, re.MULTILINE)
    if not parent_match:
        raise RuntimeError(f"sql-registry.md không có `## {parent_section}` section")
    parent_start = parent_match.start()
    # End of parent section = next `## ` heading
    end_match = re.search(r"^## (?!#)", text[parent_start + 5:], re.MULTILINE)
    parent_end = (parent_start + 5 + end_match.start()) if end_match else len(text)
    scope_text = text[parent_start:parent_end]
    # Pattern: ### <title> ... **ClickHouse SQL:** ```sql ... ```
    pat = rf"^###\s+{re.escape(section_title)}\b.*?\*\*ClickHouse SQL:\*\*\s*```sql\s*\n(.*?)\n```"
    m = re.search(pat, scope_text, re.MULTILINE | re.DOTALL)
    if not m:
        raise RuntimeError(f"Không tìm được section '{section_title}' trong '{parent_section}' (ClickHouse block)")
    sql = m.group(1).strip()
    abs_offset = parent_start + m.start(1)
    line_no = text[:abs_offset].count("\n") + 1
    return sql, line_no


def query_filter_values(client):
    """Query mv_filter_* để substitute placeholder khi ALL filter (bypass-detect pattern registry).
    Column mapping per actual schema verified 2026-05-26:
      mv_filter_cargo_brand → group_of_cargo_code, group_of_cargo_name, brand_code, brand_name
      mv_filter_channel    → channel_code, channel_name
      mv_filter_warehouse  → whseid, whseid_name, group_whseid_name
      mv_filter_region     → group_area_code, group_area_name
    """
    out = {}
    src = [
        ("whseid", "whseid", "mv_filter_warehouse"),
        ("group_of_cargo", "group_of_cargo_code", "mv_filter_cargo_brand"),
        ("brand", "brand_code", "mv_filter_cargo_brand"),
        ("region", "group_area_code", "mv_filter_region"),
        ("group_name", "channel_code", "mv_filter_channel"),
    ]
    for key, col, tbl in src:
        try:
            rows = client.query(
                f"SELECT DISTINCT {col} FROM analytics_workspace.{tbl} WHERE {col} IS NOT NULL ORDER BY 1"
            ).result_rows
            vals = ",".join("'" + str(r[0]).replace("'", "''") + "'" for r in rows)
            out[key] = vals if vals else "''"
        except Exception as e:
            print(f"    [warn] mv_filter_{key} unavailable ({e}); fallback empty string")
            out[key] = "''"
    return out


def substitute_placeholders(sql, filters, date_type, from_date, to_date, uom=UOM):
    sql = sql.replace("{{whseid}}", filters["whseid"])
    sql = sql.replace("{{group_of_cargo}}", filters["group_of_cargo"])
    sql = sql.replace("{{brand}}", filters["brand"])
    sql = sql.replace("{{region}}", filters["region"])
    sql = sql.replace("{{group_name}}", filters["group_name"])
    sql = sql.replace("{{date_type}}", f"'{date_type}'")
    sql = sql.replace("{{from_date}}", f"'{from_date}'")
    sql = sql.replace("{{to_date}}", f"'{to_date}'")
    sql = sql.replace("{{uom}}", f"'{uom}'")
    sql = sql.replace("[[", "").replace("]]", "")
    return sql


# ── Excel formula helpers ────────────────────────────────
DETAIL_SHEET = "'05 — Detail Orders'"
# Detail Orders column letters (fixed by build_detail_sheet order, updated after schema verify 2026-05-26):
# A=SO B=Kho C=Cargo group D=Brand E=Channel F=Region G=Customer code H=Customer name
# I=ETA(UTC+7) J=ATA(UTC+7) K=Actual Ship Date L=E2E label M=Status (Cancel/Close/...)
# N=Order type O=Type description P=Original CSE Q=Shipped CSE R=Sản lượng giao CSE
# S=Remark 2 (drop reason for status='Cancel'); T=Khớp Y/N U=Khách ghi chú
DETAIL_DIM_COL = {
    "Kho":     "B",
    "Khu vực": "F",
    "Customer": "H",
    "Kênh":    "E",
}
# E2E label is col L; status enum is col M; CSE volume cols = P(orig), Q(shipped), R(delivered)
E2E_COL = "L"
STATUS_COL = "M"
VOL_ORIG = "P"
VOL_SHIPPED = "Q"
VOL_DELIVERED = "R"
REMARK_COL = "S"


def f_count_e2e(label):
    return f'=COUNTIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"{label}")'


def f_volume_e2e(label, vol_col):
    """SUMIF volume cho 1 e2e_label (label in col L, volume in col P/Q/R)."""
    return f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"{label}",{DETAIL_SHEET}!{vol_col}:{vol_col})'


def f_dim_total_volume(dim_col, value_cell, vol_col=None):
    vol = vol_col or VOL_ORIG
    return f'=SUMIF({DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell},{DETAIL_SHEET}!{vol}:{vol})'


def f_dim_done_volume(dim_col, value_cell):
    """SUMIFS volume where e2e_label = Đã vận chuyển AND dim = value."""
    return (f'=SUMIFS({DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED},'
            f'{DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",'
            f'{DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell})')


def f_dim_pct_done(dim_col, value_cell):
    return (f'=IFERROR(ROUND(SUMIFS({DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED},'
            f'{DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",'
            f'{DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell})'
            f'/SUMIF({DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell},{DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})*100,2),0)')


def _date_range_args(date_str):
    y, m, d = date_str.split("-")
    next_d = f'DATE({y},{m},{d})+1'
    return f'">="&DATE({y},{m},{d}),{DETAIL_SHEET}!I:I,"<"&{next_d}'


# ── Styles ───────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="606060")
BODY_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="BFC9D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
GREEN = PatternFill("solid", fgColor="DCFCE7")
YELLOW = PatternFill("solid", fgColor="FEF3C7")
RED = PatternFill("solid", fgColor="FEE2E2")
GREY = PatternFill("solid", fgColor="F1F5F9")
ALERT_RED = PatternFill("solid", fgColor="FCA5A5")
EMPTY_CUSTOMER = PatternFill("solid", fgColor="FFFBEB")


def rag_fill_pct_done(value):
    if value is None:
        return GREY
    v = float(value)
    if v >= RAG["pct_done"]["green"]:
        return GREEN
    if v >= RAG["pct_done"]["yellow_lo"]:
        return YELLOW
    if v >= RAG["pct_done"]["alert"]:
        return RED
    return ALERT_RED


def rag_fill_drop_rate(value):
    """Inverted — low drop_rate = good."""
    if value is None:
        return GREY
    v = float(value)
    if v <= RAG["drop_rate"]["green_max"]:
        return GREEN
    if v <= RAG["drop_rate"]["yellow_max"]:
        return YELLOW
    return RED


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_row(ws, text, span):
    ws.append([text] + [""] * (span - 1))
    last = ws.max_row
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=span)
    c = ws.cell(row=last, column=1)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[last].height = 24


def header_row(ws, headers):
    ws.append(headers)
    last = ws.max_row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=last, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[last].height = 28


def body_cell(c, value=None, fill=None, fmt=None, align=None):
    c.font = BODY_FONT
    c.border = BORDER
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if value is not None:
        c.value = value


# ── Queries — load canonical từ sql-registry.md ──────────
def _normalize_pct(row, keys=("pct_done", "drop_rate")):
    for k in keys:
        if k in row:
            row[k] = _pct2(row[k])
    return row


def _normalize_num(row, keys):
    for k in keys:
        if k in row:
            row[k] = _num2(row[k])
    return row


def run_l1_l3_funnel(client, filters):
    """L1 Hero + L3 funnel — registry "Tổng Volume" trả 6 rows breakdown trong 1 query.
    Mỗi row có (trang_thai_don_do, value_uom). Parse ra dict theo trang_thai_don_do.

    Lưu ý: registry section "Tổng Volume" / "Chưa xuất kho" / "Đã vận chuyển" / ... đều là
    CÙNG 1 SQL trả 6 rows. Trước đây script load 6 lần và sum all → over-count 2×.
    Fix: chỉ chạy 1 lần, parse breakdown từ "trang_thai_don_do" column.
    """
    tpl, line_ref = load_registry_sql("Tổng Volume")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    cols = r.column_names
    # Build dict: trang_thai_don_do → value_uom
    raw_map = {}
    for row in r.result_rows:
        d = dict(zip(cols, row))
        raw_map[str(d.get("trang_thai_don_do") or "")] = float(d.get("value_uom") or 0)

    # Canonical 5 status + "Kế hoạch xuất" → frontend uses "Tổng Volume" label
    out = {
        "Tổng Volume":     _num2(raw_map.get("Kế hoạch xuất", 0)),  # Plan tổng
        "Chưa xuất kho":   _num2(raw_map.get("Chưa xuất kho", 0)),
        "Đang xuất kho":   _num2(raw_map.get("Đang xuất kho", 0)),
        "Đã xuất kho":     _num2(raw_map.get("Đã xuất kho", 0)),
        "Đang vận chuyển": _num2(raw_map.get("Đang vận chuyển", 0)),
        "Đã vận chuyển":   _num2(raw_map.get("Đã vận chuyển", 0)),
    }
    # SQL blocks for appendix — single SQL with breakdown note
    sql_blocks = [
        ("L1+L3 (Tổng Volume 1 SQL trả 6 rows breakdown)", sql, f"sql-registry.md:{line_ref}"),
    ]
    return out, sql_blocks


def run_l5_panel(client, filters, dim_key):
    """L5 dim panel — load registry "Báo cáo tổng hợp theo X"."""
    section_map = {
        "Kho":     "Báo cáo tổng hợp theo kho hệ thống",
        "Customer": "Báo cáo tổng hợp theo NPP",
        "Khu vực": "Báo cáo tổng hợp theo khu vực",
        "Kênh":    "Báo cáo tổng hợp theo kênh bán hàng",
    }
    section_name = section_map[dim_key]
    tpl, line_ref = load_registry_sql(section_name)
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [_normalize_pct(_normalize_num(dict(zip(r.column_names, row)),
                                          ("total_volume", "done_volume", "pending_volume")))
            for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_l4_drop_trend(client, filters):
    """L4 Drop Trend 14 ngày."""
    tpl, line_ref = load_registry_sql("L4 Trend tỷ lệ rớt 14 ngày")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = []
    for row in r.result_rows:
        d = dict(zip(r.column_names, row))
        d["drop_rate"] = _pct2(d.get("drop_rate"))
        d["drop_rate_30d_avg"] = _pct2(d.get("drop_rate_30d_avg"))
        d["total_plan"] = _num2(d.get("total_plan"))
        d["total_failed"] = _num2(d.get("total_failed"))
        # Strip tzinfo
        date_v = d.get("date")
        if date_v is not None and hasattr(date_v, "tzinfo") and date_v.tzinfo is not None:
            d["date"] = date_v.replace(tzinfo=None)
        rows.append(d)
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_l2_hotspot_kho(client, filters):
    """L2 Exception Spotlight ô (a) — Top N kho off-target."""
    tpl, line_ref = load_registry_sql("L2 Điểm nóng — Kho")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [_normalize_pct(dict(zip(r.column_names, row))) for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_l2_hotspot_drop(client, filters):
    """L2 ô (b) — Đơn rớt + lý do."""
    tpl, line_ref = load_registry_sql("L2 Điểm nóng — Drop + Lý do")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [dict(zip(r.column_names, row)) for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_l2_hotspot_region(client, filters):
    """L2 ô (c) — Khu vực dưới target."""
    tpl, line_ref = load_registry_sql("L2 Điểm nóng — Khu vực")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [_normalize_pct(dict(zip(r.column_names, row))) for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_t7_drop_report(client, filters):
    """T7 Drop Report bucket."""
    tpl, line_ref = load_registry_sql("Bổ sung report hàng rớt")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [_normalize_num(dict(zip(r.column_names, row)),
                           ("dry_fresh_cse", "posm_pc")) for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_t8_drop_reason(client, filters):
    """T8 Drop Reason."""
    tpl, line_ref = load_registry_sql("Bổ sung report lý do rớt đơn")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [_normalize_num(dict(zip(r.column_names, row)),
                           ("dry_fresh_cse", "posm_pc")) for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_detail(client, filters):
    """Detail Orders — ad-hoc (T9 widget config trong widget.config DB).
    Pattern filter bám registry "Tổng Volume" — same date_type + UOM + multi-filter.
    """
    f_w = filters["whseid"]
    f_c = filters["group_of_cargo"]
    f_b = filters["brand"]
    f_r = filters["region"]
    f_g = filters["group_name"]
    # Schema verified 2026-05-26: mv_flash_and_drop_report KHÔNG có trang_thai_don_do
    # (chỉ có e2e_label làm bucket E2E pre-computed). Detail query bám đúng schema thực.
    # Status column = raw enum (Active/Cancel/Close/...), thay role của trang_thai_don_do trong dim categorization.
    q = f"""
    SELECT
      so                                                       AS so,
      coalesce(whseid, '(rỗng)')                               AS kho,
      coalesce(group_of_cago, '(rỗng)')                        AS cargo_group,
      coalesce(brand, '(rỗng)')                                AS brand,
      coalesce(group_name, '(rỗng)')                           AS kenh,
      coalesce(khu_vuc_doi_xe, '(rỗng)')                       AS khu_vuc,
      coalesce(customer_code, '')                              AS customer_code,
      coalesce(customer_name, '')                              AS customer_name,
      toDateTime(eta_giao_hang_cho_npp, 'Asia/Ho_Chi_Minh')    AS eta_utc7,
      toDateTime(ata_den, 'Asia/Ho_Chi_Minh')                  AS ata_utc7,
      toDateTime(actual_ship_date, 'Asia/Ho_Chi_Minh')         AS asd_utc7,
      coalesce(e2e_label, '')                                  AS e2e_label,
      coalesce(status, '')                                     AS status,
      coalesce(type, '')                                       AS order_type,
      coalesce(type_description, '')                           AS type_description,
      toFloat64(coalesce(original_cse, 0))                     AS original_cse,
      toFloat64(coalesce(shipped_cse, 0))                      AS shipped_cse,
      toFloat64(coalesce(san_luong_giao_cse, 0))               AS san_luong_giao_cse,
      coalesce(remark_2, '')                                   AS remark_2
    FROM {SOURCE_MV_MAIN}
    WHERE 1 = 1
      AND ( toDate(
        CASE
          WHEN '{DATE_TYPE_LABEL}' = 'Ngày GI'              THEN delivery_date_1
          WHEN '{DATE_TYPE_LABEL}' = 'Actual Ship Date'     THEN actual_ship_date
          WHEN '{DATE_TYPE_LABEL}' = 'ETD gửi thầu (đơn)'  THEN etd_chuyen_gui_thau
          WHEN '{DATE_TYPE_LABEL}' = 'ETA gửi thầu (đơn)'  THEN eta_giao_hang_cho_npp
          WHEN '{DATE_TYPE_LABEL}' = 'ATA đơn'              THEN ata_den
          ELSE delivery_date_1
        END)
        BETWEEN toDate('{WINDOW_START_DATE}') AND toDate('{WINDOW_END_DATE}')
      )
    ORDER BY delivery_date_1 ASC, so ASC
    LIMIT 50000
    """
    r = client.query(q)
    rows = []
    for row in r.result_rows:
        d = dict(zip(r.column_names, row))
        for k in ("original_cse", "shipped_cse", "san_luong_giao_cse"):
            d[k] = _num2(d.get(k))
        for k in ("eta_utc7", "ata_utc7", "asd_utc7"):
            v = d.get(k)
            if v is not None and hasattr(v, "tzinfo") and v.tzinfo is not None:
                d[k] = v.replace(tzinfo=None)
        rows.append(d)
    return rows, q.strip(), "ad-hoc (T9 detailTable widget config — pattern filter bám registry Flash Report)"


def run_mv_now(client):
    try:
        r = client.query("SELECT now('Asia/Ho_Chi_Minh') AS server_now")
        return r.result_rows[0][0]
    except Exception:
        return None


# ── Excel build ──────────────────────────────────────────
def build_readme(ws, ctx):
    ws.title = "00 — README"
    set_widths(ws, [24, 88])
    title_row(ws, "UAT Flash Daily — Số SQL thật cho khách Mondelez confirm", 2)

    funnel = ctx["funnel"]
    plan_total = float(funnel["Tổng Volume"] or 0)
    done = float(funnel["Đã vận chuyển"] or 0)
    pct_done = round(100.0 * done / plan_total, 2) if plan_total > 0 else 0

    meta = [
        ("Section", "Flash Daily — Smartlog Control Tower"),
        ("Tenant", TENANT),
        ("Window test", f"{WINDOW_START_DATE} → {WINDOW_END_DATE} (1 ngày — D-1 session UAT 2026-06-09)"),
        ("Date Type", f"{DATE_TYPE_LABEL}  (cột MV: {DATE_COL}) — ⚠ exact CH CASE branch label"),
        ("UOM", f"{UOM} (mặc định)"),
        ("Filter còn lại", "ALL (Kho/Sales Channel/Cargo Group/Brand/Region không chọn)"),
        ("Nguồn dữ liệu", f"{SOURCE_MV_MAIN} (L1/L3/L5) + {SOURCE_MV_DROP} (L4/T7/T8) + {SOURCE_MV_RAW} (T9)"),
        ("SQL provenance",
         "L1 Hero + L3 funnel 5 status + L5 4 dim panels + L2 Exception 3 ô + L4 Drop Trend 14d + "
         "T7 Drop bucket + T8 Drop reason — load CANONICAL từ sql-registry.md (## Flash Report, "
         "xem sheet 08). T9 Detail orders — ad-hoc (widget config-specific). Khi registry update, "
         "re-run script để pick up."),
        ("KPI vs Detail row count",
         f"L1 Hero Plan = {plan_total:,.2f} {UOM} (registry 'Tổng Volume'). Detail Orders sheet có "
         f"{len(ctx['detail']):,} row (1 row = 1 SO+whseid key). Diff acceptable nếu 1 SO split nhiều whseid."),
        ("Target % Hoàn thành", "95% overall (per memory [[project_mondelez_flash_daily_target]])"),
        ("RAG bands per metric",
         "% Hoàn thành: Green ≥ 95 / Yellow 85–<95 / Red 80–<85 / Alert (red đậm) <80. "
         "drop_rate (L4): Green ≤ 5 / Yellow 5–10 / Red > 10."),
        ("Format chuẩn (consistent)",
         "% metric: 2 decimal places fixed, format '0.00\"%\"' (display 93.30%). "
         "Count: format '#,##0'. CSE / số có thập phân: format '#,##0.00'. "
         "Datetime ETA/ATA/ASD: format 'yyyy-mm-dd hh:mm:ss' (datetime cell, sort + filter OK)."),
        ("Run timestamp", ctx["run_ts"]),
        ("CH server now (UTC+7)", str(ctx["server_now"]) if ctx["server_now"] else "n/a"),
        ("L1 Hero % Hoàn thành (frozen)", f"{pct_done:.2f}%"),
        ("L1 Hero Plan (frozen)", f"{plan_total:,.2f} {UOM}"),
        ("L1 Hero Đã giao (frozen)", f"{done:,.2f} {UOM}"),
        ("", ""),
        ("Mục đích file",
         "PM/BA mang số SQL thật xuống cho khách Mondelez confirm 4 việc: (1) Định nghĩa 5 trạng thái "
         "E2E + STM mutually exclusive có khớp mental model Ops không; (2) % Hoàn thành + Plan/Đã giao "
         "khớp golden file 3 nguồn (SAP plan + WMS shipped + STM giao); (3) Top N kho off-target + "
         "khu vực dưới target có đúng kỳ vọng vận hành không; (4) Drop bucket + reason có map đúng "
         "với business process MDLZ không."),
        ("", ""),
        ("Hướng dẫn cho khách",
         "Khách chỉ cần nhìn cột 'SQL value', so với số nội bộ MDLZ (Excel/pipeline báo cáo). Nếu khớp → "
         "tick Status = OK. Nếu lệch ngoài tolerance → ghi rõ vào cột 'Khách ghi chú' để Smartlog "
         "điều tra. Sheet 07 (UX & Filter checklist) khách verify visual/storytelling/master data, "
         "KHÔNG compute từ data."),
        ("", ""),
        ("Tolerance default",
         "Số đếm tuyệt đối ≤ 1% · % metric ≤ 0.5pp · Top N ≥ 4/5 tên match · "
         "L5 panel SUM(total_volume) parity L1 Plan ≤ 1% (regression bug 2026-05-18 check)"),
        ("", ""),
        ("Regression gate quan trọng",
         "Bug 2026-05-18: L5 4 panel SUM(total_volume) phải PARITY L1 Plan (≤ 1%). Nếu lệch → CTE "
         "thiếu `trang_thai_don_do IN (5 canonical status)` filter (memory "
         "[[feedback_l5_sql_canonical_status_filter]]). Verify trong sheet 02 cell footer 'Σ check'."),
        ("", ""),
        ("Sheet trong file",
         "00 — README · 01 — L1+L3 KPI (% Hoàn thành + 5 status funnel, có cột công thức live) · "
         "02 — L5 Health Matrix 4 dim (cell tính live từ sheet 05 + Σ check parity) · "
         "03 — L2 Exception Spotlight 3 ô (kho/drop/khu vực off-target) · "
         "04 — L4 Drop Trend 14 ngày (line chart data + rolling 30d) · "
         "05 — Detail Orders (raw — 1 row = 1 DO, autofilter, là nguồn cho mọi formula) · "
         "06 — Formula Guide (mapping cột + pattern formula) · "
         "07 — UX & Filter checklist (~35 mục cho khách verify visual/filter/interaction/storytelling) · "
         "08 — SQL Appendix (provenance: registry vs ad-hoc) + "
         "09 — T7/T8 Drop Report & Reason (bucket pattern check)"),
        ("Insight file riêng",
         "Operational insight (master data scorecard, 6 red flags, 9 KPI extension, 7 open questions) "
         "ở file md riêng: flash-daily-uat-ops-insight-2026-06-08_to_2026-06-08.md (same folder) — "
         "KHÔNG nhồi vào Excel để giữ UAT pack tập trung vào reconciliation."),
        ("Cách verify",
         "Số trong sheet 01-04 + 09 đều có 2 dạng — SQL value (frozen, do Python ghi) và Excel formula "
         "(live, tính lại từ sheet 05). Hai phải khớp. Nếu lệch → có gap filter ở 1 trong 2. Xem "
         "sheet 06 Formula Guide để copy công thức custom."),
    ]
    for k, v in meta:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP
    for r_ in range(2, ws.max_row + 1):
        ws.row_dimensions[r_].height = max(18, min(120, 18 + len(str(ws.cell(row=r_, column=2).value or "")) // 70 * 16))


def build_l1_l3_sheet(ws, ctx):
    """Sheet 01 — L1 Hero (4 KPI: %Done, Plan, Done, Remaining) + L3 funnel (5 status)."""
    ws.title = "01 — L1+L3 KPI"
    set_widths(ws, [38, 16, 60, 16, 16, 16, 18, 24])
    title_row(ws, f"L1 Hero + L3 Funnel — {WINDOW_START_DATE} ({DATE_TYPE_LABEL}, UOM={UOM})", 8)
    header_row(ws, [
        "Metric", "SQL value\n(frozen)", "Công thức Excel\n(live từ sheet 05)",
        "Dashboard\n(khách verify)", "Số MDLZ\n(golden file)",
        "Diff\n(Dashboard − SQL)", "Tolerance", "Khách ghi chú / Status",
    ])

    f = ctx["funnel"]
    plan = float(f["Tổng Volume"] or 0)
    done = float(f["Đã vận chuyển"] or 0)
    remaining = plan - done
    pct_done = _pct2(100.0 * done / plan) if plan > 0 else _pct2(0)

    # L1 Hero block — header
    ws.append(["▼ L1 HERO (% Hoàn thành + 3 sub-numbers)"])
    last = ws.max_row
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=8)
    ws.cell(row=last, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=last, column=1).fill = PatternFill("solid", fgColor="14283F")
    ws.row_dimensions[last].height = 22

    def _add_row(label, sql_val, formula, fmt, tolerance, rag_fn=None):
        ws.append([label, sql_val, formula, "", "", "", tolerance, ""])
        r_ = ws.max_row
        for col in range(1, 9):
            body_cell(ws.cell(row=r_, column=col))
        sv = ws.cell(row=r_, column=2)
        fv = ws.cell(row=r_, column=3)
        sv.number_format = fmt
        fv.number_format = fmt
        if rag_fn and isinstance(sql_val, (int, float, Decimal)):
            sv.fill = rag_fn(float(sql_val))
        for col in (4, 5):
            ws.cell(row=r_, column=col).fill = EMPTY_CUSTOMER
        df = ws.cell(row=r_, column=6)
        df.value = f"=IFERROR(D{r_}-B{r_}, \"\")"
        df.number_format = fmt

    # L1 cards — formulas dùng E2E_COL=L (e2e_label), VOL_ORIG=P, VOL_DELIVERED=R
    pct_formula = (f'=IFERROR(ROUND(SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED})'
                   f'/SUM({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})*100,2),0)')
    plan_formula = f'=SUM({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})'
    done_formula = f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED})'
    remaining_formula = f'=SUM({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})-SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED})'

    _add_row("% Hoàn thành (target 95%)", pct_done, pct_formula, "0.00\"%\"", "≤ 0.5 pp", rag_fill_pct_done)
    _add_row(f"Plan ({UOM}) — tổng kế hoạch", _num2(plan), plan_formula, "#,##0.00", "≤ 1%")
    _add_row(f"Đã giao ({UOM}) — đã vận chuyển", _num2(done), done_formula, "#,##0.00", "≤ 1%")
    _add_row(f"Còn lại ({UOM}) — Plan − Đã giao", _num2(remaining), remaining_formula, "#,##0.00", "≤ 1%")

    ws.append([])

    # L3 Funnel block — header
    ws.append(["▼ L3 FUNNEL — 5 trạng thái E2E (volume + % share)"])
    last = ws.max_row
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=8)
    ws.cell(row=last, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=last, column=1).fill = PatternFill("solid", fgColor="14283F")
    ws.row_dimensions[last].height = 22

    # Volume column per status (per PRD §3.1):
    #   Chưa xuất / Đang xuất: original_cse (col P = VOL_ORIG)
    #   Đã xuất / Đang vận:    shipped_cse  (col Q = VOL_SHIPPED)
    #   Đã vận:                san_luong_giao_cse (col R = VOL_DELIVERED)
    status_cfg = [
        ("Chưa xuất kho",    f["Chưa xuất kho"],    VOL_ORIG,      "(QTY OPEN)"),
        ("Đang xuất kho",    f["Đang xuất kho"],    VOL_ORIG,      "(QTY PICKDETAILED)"),
        ("Đã xuất kho",      f["Đã xuất kho"],      VOL_SHIPPED,   "(QTY SHIPPED, chưa ATD STM)"),
        ("Đang vận chuyển",  f["Đang vận chuyển"],  VOL_SHIPPED,   "(QTY SHIPPED, có ATD chưa ATA)"),
        ("Đã vận chuyển",    f["Đã vận chuyển"],    VOL_DELIVERED, "(Sản lượng giao, có ATA STM)"),
    ]

    for status, val, vol_col, note in status_cfg:
        formula = f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"{status}",{DETAIL_SHEET}!{vol_col}:{vol_col})'
        _add_row(f"{status} {note}", _num2(val), formula, "#,##0.00", "≤ 1%")
        # Share % row referencing Plan via SUM(VOL_ORIG)
        share = _pct2(100.0 * float(val or 0) / plan) if plan > 0 else _pct2(0)
        ws.append([f"   └ % share của {status}", share,
                   (f'=IFERROR(ROUND(SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"{status}",{DETAIL_SHEET}!{vol_col}:{vol_col})'
                    f'/SUM({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})*100,2),0)'),
                   "", "", "", "≤ 0.5pp", ""])
        share_r = ws.max_row
        for col in range(1, 9):
            body_cell(ws.cell(row=share_r, column=col))
        ws.cell(row=share_r, column=1).font = Font(italic=True, color="606060")
        ws.cell(row=share_r, column=2).number_format = "0.00\"%\""
        ws.cell(row=share_r, column=3).number_format = "0.00\"%\""

    ws.append([])
    ws.append(["Note STM lag",
               "2 status 'Đã xuất kho' + 'Đang vận chuyển' phụ thuộc STM signal (ATD/ATA). Nếu signal lag >12h, "
               "đơn vật lý đã rời kho nhưng vẫn ở bucket 'Đã xuất kho'. CH MV pre-computes `e2e_label` mutually "
               "exclusive qua `thoi_gian_di IS NULL` → KHÔNG inflate count tổng (PRD §3.1 v1.1.0 + Audit A3)."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)


def build_l5_sheet(ws, ctx):
    """Sheet 02 — L5 Health Matrix 4 dim với DUAL SQL frozen + Excel formula (per skill spec)."""
    ws.title = "02 — L5 Health Matrix"
    set_widths(ws, [22, 26, 13, 13, 13, 13, 13, 13, 16, 24])
    title_row(ws,
              "L5 4 dim panels — DUAL: SQL value (frozen) | Excel formula (live từ sheet 05). "
              "Σ check vs L1 Plan ở cell cuối mỗi panel.",
              10)
    header_row(ws, [
        "Dimension", "Giá trị",
        "Total\n(SQL)", "Total\n(formula)",
        "Done\n(SQL)", "Done\n(formula)",
        "% Done\n(SQL)", "% Done\n(formula)",
        "Σ check\nvs L1", "Khách ghi chú",
    ])

    sections = [
        ("Kho hệ thống", ctx["dim_kho"], "Kho", DETAIL_DIM_COL["Kho"]),
        ("Khu vực", ctx["dim_khu_vuc"], "Khu vực", DETAIL_DIM_COL["Khu vực"]),
        ("Customer (NPP = Customer per Mondelez)", ctx["dim_customer"], "Customer", DETAIL_DIM_COL["Customer"]),
        ("Kênh bán hàng", ctx["dim_kenh"], "Kênh", DETAIL_DIM_COL["Kênh"]),
    ]
    plan_total = float(ctx["funnel"]["Tổng Volume"] or 0)

    for label, rows, dim_key, dim_col in sections:
        ws.append([label])
        last = ws.max_row
        ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=10)
        c = ws.cell(row=last, column=1)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="14283F")
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[last].height = 22

        if not rows:
            ws.append(["", "(không có data)", "", "", "", "", "", "", "", ""])
            continue

        rows_sorted = sorted(rows, key=lambda r: float(r.get("pct_done") or 0))
        dim_value_key = list(rows_sorted[0].keys())[0]

        sum_total = 0.0
        for row in rows_sorted[:30]:
            dim_val = row.get(dim_value_key, "")
            total_sql = float(row.get("total_volume") or 0)
            done_sql = float(row.get("done_volume") or 0)
            pct_sql = float(row.get("pct_done") or 0)

            ws.append(["", dim_val, total_sql, "", done_sql, "", pct_sql, "", "", ""])
            r_ = ws.max_row
            v_ref = f"$B{r_}"

            # Formula cells (live recompute từ Detail)
            ws.cell(row=r_, column=4).value = f_dim_total_volume(dim_col, v_ref)
            ws.cell(row=r_, column=6).value = f_dim_done_volume(dim_col, v_ref)
            ws.cell(row=r_, column=8).value = f_dim_pct_done(dim_col, v_ref)
            # Σ check: diff % giữa SQL total và formula total — phát hiện gap filter
            ws.cell(row=r_, column=9).value = f'=IFERROR(ROUND((D{r_}-C{r_})/C{r_}*100,2),0)'

            for col in range(1, 11):
                body_cell(ws.cell(row=r_, column=col))
            # Number formats
            for col in (3, 4, 5, 6):
                ws.cell(row=r_, column=col).number_format = "#,##0.00"
            for col in (7, 8):
                ws.cell(row=r_, column=col).number_format = "0.00\"%\""
            ws.cell(row=r_, column=9).number_format = "+0.00\"%\";-0.00\"%\";0\"%\""
            # RAG color on % Done SQL cell
            ws.cell(row=r_, column=7).fill = rag_fill_pct_done(pct_sql)
            ws.cell(row=r_, column=8).fill = rag_fill_pct_done(pct_sql)
            sum_total += total_sql

        # Σ panel row — parity check vs L1 Plan
        diff_pct = round(100.0 * (sum_total - plan_total) / plan_total, 2) if plan_total > 0 else 0
        ws.append([
            "",
            f"Σ {label} (top {min(len(rows_sorted), 30)})",
            sum_total,
            f'=SUMPRODUCT(C{ws.max_row - len(rows_sorted[:30]) + 0}:C{ws.max_row - 1}*1)',
            "", "", "", "",
            f"L1 Plan {plan_total:,.2f} · diff {diff_pct:+.2f}%",
            "",
        ])
        sum_r = ws.max_row
        for col in range(1, 11):
            body_cell(ws.cell(row=sum_r, column=col))
        ws.cell(row=sum_r, column=2).font = Font(bold=True)
        for col in (3, 4):
            ws.cell(row=sum_r, column=col).number_format = "#,##0.00"
            ws.cell(row=sum_r, column=col).font = Font(bold=True)
        ws.cell(row=sum_r, column=9).fill = GREEN if abs(diff_pct) <= 1.0 else (
            YELLOW if abs(diff_pct) <= 5.0 else RED)
        ws.cell(row=sum_r, column=9).font = Font(bold=True)

        ws.append([])

    ws.append([
        "Note dual",
        "Mỗi metric có 2 cell side-by-side: SQL value (frozen, từ Python) | Excel formula (live, "
        "compute từ Detail sheet 05). Hai phải khớp ±0.01 cho mọi dim row. Cột 'Σ check' = "
        "(formula − SQL) / SQL × 100 — nếu > ±0.1% → có gap filter giữa registry SQL và Detail "
        "query. Cột 'Σ check' panel cuối = parity vs L1 Plan (per memory "
        "[[feedback_l5_sql_canonical_status_filter]]) — bug 2026-05-18 regression gate.",
    ])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=10)
    ws.row_dimensions[ws.max_row].height = 60


def build_l2_sheet(ws, ctx):
    """Sheet 03 — L2 Exception Spotlight 3 ô với DUAL SQL frozen + Excel formula (per skill spec)."""
    ws.title = "03 — L2 Exception"
    set_widths(ws, [22, 28, 14, 14, 14, 14, 22])
    title_row(ws,
              "L2 Exception Spotlight — DUAL SQL value | Excel formula. "
              "3 ô (Top kho off-target / Drop + lý do / Khu vực dưới target)",
              7)

    def _add_dim_section(title, rows, dim_col_letter, dim_value_key):
        """Section cho dim-based (kho/region): % Done SQL | % Done formula."""
        ws.append([title])
        last = ws.max_row
        ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=7)
        c = ws.cell(row=last, column=1)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="14283F")
        ws.row_dimensions[last].height = 22
        header_row(ws, ["", "Giá trị",
                        "% Done\n(SQL)", "% Done\n(formula)",
                        "Total\n(formula)", "Done\n(formula)",
                        "Khách ghi chú"])
        if not rows:
            ws.append(["", "(không có data)", "", "", "", "", ""])
            return
        for row in rows[:10]:
            v = row.get(dim_value_key, "")
            pct_sql = float(row.get("pct_done") or 0) if row.get("pct_done") is not None else None
            ws.append(["", v, pct_sql, "", "", "", ""])
            r_ = ws.max_row
            v_ref = f"$B{r_}"
            ws.cell(row=r_, column=4).value = f_dim_pct_done(dim_col_letter, v_ref)
            ws.cell(row=r_, column=5).value = f_dim_total_volume(dim_col_letter, v_ref)
            ws.cell(row=r_, column=6).value = f_dim_done_volume(dim_col_letter, v_ref)
            for col in range(1, 8):
                body_cell(ws.cell(row=r_, column=col))
            for col in (3, 4):
                ws.cell(row=r_, column=col).number_format = "0.00\"%\""
                if pct_sql is not None:
                    ws.cell(row=r_, column=col).fill = rag_fill_pct_done(pct_sql)
            for col in (5, 6):
                ws.cell(row=r_, column=col).number_format = "#,##0.00"
        ws.append([])

    def _add_drop_section(title, rows, dim_value_key):
        """Section cho drop: SQL count | Excel formula = COUNTIF status='Cancel' với reason match."""
        ws.append([title])
        last = ws.max_row
        ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=7)
        c = ws.cell(row=last, column=1)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="14283F")
        ws.row_dimensions[last].height = 22
        header_row(ws, ["", "Lý do (raw remark_2)",
                        "Count\n(SQL)", "Count\n(formula)",
                        "Volume CSE\n(SQL)", "Volume CSE\n(formula)",
                        "Khách ghi chú"])
        if not rows:
            ws.append(["", "(không có data)", "", "", "", "", ""])
            return
        keys = list(rows[0].keys())
        count_key = next((k for k in keys if "count" in k.lower() or "so" in k.lower() or k == keys[1]), keys[1] if len(keys) > 1 else None)
        vol_key = next((k for k in keys if "cse" in k.lower() or "volume" in k.lower()), keys[2] if len(keys) > 2 else None)
        for row in rows[:10]:
            reason = row.get(dim_value_key, "") or "(rỗng)"
            cnt_sql = row.get(count_key) if count_key else None
            vol_sql = row.get(vol_key) if vol_key and vol_key != count_key else None
            # Excel formulas: count đơn có status='Cancel' AND remark khớp (col U sau khi add)
            cnt_formula = (f'=COUNTIFS({DETAIL_SHEET}!{STATUS_COL}:{STATUS_COL},"Cancel",'
                           f'{DETAIL_SHEET}!{REMARK_COL}:{REMARK_COL},$B{ws.max_row + 1})')
            vol_formula = (f'=SUMIFS({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG},'
                           f'{DETAIL_SHEET}!{STATUS_COL}:{STATUS_COL},"Cancel",'
                           f'{DETAIL_SHEET}!{REMARK_COL}:{REMARK_COL},$B{ws.max_row + 1})')
            ws.append(["", reason, cnt_sql, cnt_formula,
                       vol_sql if isinstance(vol_sql, (int, float, Decimal)) else "",
                       vol_formula, ""])
            r_ = ws.max_row
            for col in range(1, 8):
                body_cell(ws.cell(row=r_, column=col))
            ws.cell(row=r_, column=3).number_format = "#,##0"
            ws.cell(row=r_, column=4).number_format = "#,##0"
            ws.cell(row=r_, column=5).number_format = "#,##0.00"
            ws.cell(row=r_, column=6).number_format = "#,##0.00"
        ws.append([])

    # Ô (a) Top kho off-target
    kho_rows = ctx["l2_kho"]
    first_key_kho = list(kho_rows[0].keys())[0] if kho_rows else "name"
    _add_dim_section("Ô (a) — Top kho off-target (< 85% pct_done)",
                     kho_rows, DETAIL_DIM_COL["Kho"], first_key_kho)

    # Ô (b) Drop + lý do — formula compute từ Detail col U (remark_2) + M (status)
    drop_rows = ctx["l2_drop"]
    first_key_drop = list(drop_rows[0].keys())[0] if drop_rows else "remark"
    _add_drop_section("Ô (b) — Đơn rớt + lý do (top N theo count)", drop_rows, first_key_drop)

    # Ô (c) Khu vực dưới target
    region_rows = ctx["l2_region"]
    first_key_region = list(region_rows[0].keys())[0] if region_rows else "name"
    _add_dim_section("Ô (c) — Khu vực dưới target (< 95% pct_done)",
                     region_rows, DETAIL_DIM_COL["Khu vực"], first_key_region)

    ws.append([
        "Note dual",
        "Mỗi metric có 2 cell SQL | formula side-by-side. Formula tính live từ Detail sheet 05: "
        "% Done dùng SUMIFS pattern; Drop count dùng COUNTIFS status='Cancel' + remark_2 (col U). "
        "Hai cell phải khớp ±0.01.",
    ])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=7)
    ws.row_dimensions[ws.max_row].height = 48


def build_l4_sheet(ws, ctx):
    """Sheet 04 — L4 Drop Trend 14 ngày với DUAL SQL frozen + Excel formula (chỉ ngày trong window có data Detail)."""
    ws.title = "04 — L4 Drop Trend"
    set_widths(ws, [13, 14, 14, 14, 14, 12, 12, 12, 22])
    title_row(ws,
              "L4 Drop Trend 14 ngày — DUAL: SQL frozen | Excel formula (ngày trong Detail window). "
              "FAIL = status='Cancel' only (H1), target ≤5%.",
              9)
    header_row(ws, [
        "Date",
        "Total plan\n(SQL)", "Total plan\n(formula)",
        "Total failed\n(SQL)", "Total failed\n(formula)",
        "drop_rate %\n(SQL)", "drop_rate %\n(formula)",
        "30d avg %\n(SQL)",
        "Khách ghi chú",
    ])

    from datetime import date as date_cls
    try:
        window_date = date_cls.fromisoformat(WINDOW_START_DATE)
    except Exception:
        window_date = None

    for row in ctx["l4_trend"]:
        d = row["date"]
        d_only = d.date() if isinstance(d, datetime) else d
        in_window = (window_date is not None and d_only == window_date)

        if in_window:
            y, m, dd = window_date.year, window_date.month, window_date.day
            plan_formula = (f'=SUMIFS({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG},'
                            f'{DETAIL_SHEET}!I:I,">="&DATE({y},{m},{dd}),'
                            f'{DETAIL_SHEET}!I:I,"<"&DATE({y},{m},{dd})+1)')
            failed_formula = (f'=SUMIFS({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG},'
                              f'{DETAIL_SHEET}!I:I,">="&DATE({y},{m},{dd}),'
                              f'{DETAIL_SHEET}!I:I,"<"&DATE({y},{m},{dd})+1,'
                              f'{DETAIL_SHEET}!{STATUS_COL}:{STATUS_COL},"Cancel")')
            drop_formula = (f'=IFERROR(ROUND(SUMIFS({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG},'
                            f'{DETAIL_SHEET}!I:I,">="&DATE({y},{m},{dd}),'
                            f'{DETAIL_SHEET}!I:I,"<"&DATE({y},{m},{dd})+1,'
                            f'{DETAIL_SHEET}!{STATUS_COL}:{STATUS_COL},"Cancel")'
                            f'/SUMIFS({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG},'
                            f'{DETAIL_SHEET}!I:I,">="&DATE({y},{m},{dd}),'
                            f'{DETAIL_SHEET}!I:I,"<"&DATE({y},{m},{dd})+1)*100,2),0)')
        else:
            plan_formula = "n/a (out of Detail window)"
            failed_formula = "n/a"
            drop_formula = "n/a"

        ws.append([d, row["total_plan"], plan_formula,
                   row["total_failed"], failed_formula,
                   row["drop_rate"], drop_formula,
                   row["drop_rate_30d_avg"], ""])
        r_ = ws.max_row
        for col in range(1, 10):
            body_cell(ws.cell(row=r_, column=col))
        ws.cell(row=r_, column=1).number_format = "yyyy-mm-dd"
        for col in (2, 3, 4, 5):
            ws.cell(row=r_, column=col).number_format = "#,##0.00"
        for col in (6, 7, 8):
            ws.cell(row=r_, column=col).number_format = "0.00\"%\""
        dr = row.get("drop_rate")
        if dr is not None:
            ws.cell(row=r_, column=6).fill = rag_fill_drop_rate(dr)
            if in_window:
                ws.cell(row=r_, column=7).fill = rag_fill_drop_rate(dr)
        if not in_window:
            for col in (3, 5, 7):
                ws.cell(row=r_, column=col).font = Font(italic=True, color="808080")
                ws.cell(row=r_, column=col).number_format = "@"

    ws.append([])
    ws.append([
        "Note L4 dual",
        f"Formula chỉ feasible cho ngày trong Detail window ({WINDOW_START_DATE}) — Detail sheet 05 "
        f"chỉ chứa 1 ngày data. Các ngày khác trong 14-day trend chỉ có SQL frozen (từ registry "
        f"\"L4 Trend tỷ lệ rớt 14 ngày\"). Khi UAT thực, đổi WINDOW = 14 ngày, re-run script để mọi "
        f"cell L4 đều có formula. Reference lines dashboard: y=5% solid red (target ≤5%) + rolling "
        f"30d avg dashed grey. FAIL chỉ tính status='Cancel' (H1) — KHÔNG include Close.",
    ])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=9)
    ws.row_dimensions[ws.max_row].height = 72


def build_detail_sheet(ws, ctx):
    ws.title = "05 — Detail Orders"
    headers = [
        "SO", "Kho", "Cargo group", "Brand", "Kênh", "Khu vực",
        "Customer code", "Customer name",
        "ETA (UTC+7)", "ATA (UTC+7)", "Actual Ship Date (UTC+7)",
        "E2E label", "Status",
        "Order type", "Type description",
        "Original CSE", "Shipped CSE", "Sản lượng giao CSE",
        "Remark 2 (drop reason)",
        "Khớp golden? (Y/N)", "Khách ghi chú",
    ]
    widths = [
        16, 12, 14, 14, 14, 18,
        14, 28,
        20, 20, 20,
        18, 14,
        14, 20,
        14, 14, 16,
        28,
        18, 30,
    ]
    set_widths(ws, widths)
    title_row(ws,
              f"L6 Detail Orders — {len(ctx['detail']):,} đơn (1 row = 1 DO) cho khách đối chiếu line-by-line",
              len(headers))
    header_row(ws, headers)

    e2e_fill = {
        "Chưa xuất kho":   GREY,
        "Đang xuất kho":   YELLOW,
        "Đã xuất kho":     PatternFill("solid", fgColor="EDE7F6"),  # light violet
        "Đang vận chuyển": PatternFill("solid", fgColor="DBEAFE"),  # light blue
        "Đã vận chuyển":   GREEN,
    }

    for row in ctx["detail"]:
        ws.append([
            row["so"], row["kho"], row["cargo_group"], row["brand"], row["kenh"], row["khu_vuc"],
            row["customer_code"], row["customer_name"],
            row["eta_utc7"], row["ata_utc7"], row["asd_utc7"],
            row["e2e_label"], row["status"],
            row["order_type"], row["type_description"],
            row["original_cse"], row["shipped_cse"], row["san_luong_giao_cse"],
            row["remark_2"],
            "", "",
        ])
        r_ = ws.max_row
        for col in range(1, len(headers) + 1):
            body_cell(ws.cell(row=r_, column=col))
        # Datetime cols I/J/K (9/10/11)
        for col in (9, 10, 11):
            ws.cell(row=r_, column=col).number_format = "yyyy-mm-dd hh:mm:ss"
        # CSE cols P/Q/R (16/17/18)
        for col in (16, 17, 18):
            ws.cell(row=r_, column=col).number_format = "#,##0.00"
        # E2E label color (col 12 = L)
        fill = e2e_fill.get(ws.cell(row=r_, column=12).value)
        if fill:
            ws.cell(row=r_, column=12).fill = fill
        # Empty customer cols T/U → now V/W (after adding remark_2 at U)
        for col in (20, 21):
            ws.cell(row=r_, column=col).fill = EMPTY_CUSTOMER

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"


def build_formula_guide_sheet(ws, ctx):
    ws.title = "06 — Formula Guide"
    set_widths(ws, [28, 90])
    title_row(ws, "Excel Formula Guide — copy paste khi PM/Khách hàng cần verify thủ công", 2)

    intro = [
        ("Mục đích",
         "Tất cả số trong sheet 01-04 + 09 đều có công thức live tính lại từ sheet 05 (Detail Orders). "
         "PM/khách có thể copy công thức từ đây sang Excel nội bộ hoặc dùng để audit số dashboard."),
        ("ETA/ATA/ASD dtype = datetime",
         "Cột I + J + K ở sheet 05 là datetime cell (yyyy-mm-dd hh:mm:ss). Sort + filter range OK. "
         "Trend daily formulas dùng COUNTIFS với DATE() bounds thay vì prefix-text match."),
        ("Sheet name reference",
         "Tên sheet detail: '05 — Detail Orders' (có space + em-dash, BẮT BUỘC giữ nguyên)."),
        ("STM lag note",
         "2 status 'Đã xuất kho' + 'Đang vận chuyển' phụ thuộc STM signal ATD/ATA. CH MV "
         "pre-computes `e2e_label` mutually exclusive — KHÔNG inflate count tổng (Audit A3)."),
        ("Multi-unit T7/T8",
         "Drop Report T7 + Drop Reason T8 có 2 unit độc lập DRY&FRESH (cse) + POSM (pc). "
         "Render riêng 'N cse · M pc', KHÔNG sum (per memory [[feedback_mondelez_dropreport_cse_pc_split]])."),
    ]
    for k, v in intro:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP

    ws.append([])
    ws.append(["Bảng 1 — Mapping cột Detail Orders", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="14283F")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    col_map = [
        ("A", "SO (mã đơn)"),
        ("B", "Kho (whseid)"),
        ("C", "Cargo group (group_of_cago)"),
        ("D", "Brand"),
        ("E", "Kênh bán (group_name)"),
        ("F", "Khu vực (khu_vuc_doi_xe)"),
        ("G", "Customer code"),
        ("H", "Customer name"),
        ("I", "ETA UTC+7 (datetime)"),
        ("J", "ATA UTC+7 (datetime)"),
        ("K", "Actual Ship Date UTC+7 (datetime)"),
        ("L", "E2E label (Chưa xuất / Đang xuất / Đã xuất / Đang vận / Đã vận — bucket pre-computed)"),
        ("M", "Status (raw: Active / Cancel / Close / ...)"),
        ("N", "Order type (raw type code)"),
        ("O", "Type description"),
        ("P", "Original CSE (Plan volume)"),
        ("Q", "Shipped CSE"),
        ("R", "Sản lượng giao CSE (Done volume — chỉ có khi e2e_label='Đã vận chuyển')"),
        ("S", "Remark 2 (drop reason — non-empty khi status='Cancel')"),
        ("T", "Khớp golden? (Y/N) — khách điền"),
        ("U", "Khách ghi chú — khách điền"),
    ]
    for col, desc in col_map:
        ws.append([col, desc])
        for c in range(1, 3):
            body_cell(ws.cell(row=ws.max_row, column=c))

    ws.append([])
    ws.append(["Bảng 2 — Pattern công thức chuẩn", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="14283F")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)

    patterns = [
        ("Plan tổng (UOM=cse)",
         f'=SUM({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})'),
        ("Đã giao (UOM=cse)",
         f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED})'),
        ("% Hoàn thành",
         f'=ROUND(SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED})/SUM({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})*100,2)'),
        ("Volume status 'Chưa xuất kho'",
         f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Chưa xuất kho",{DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})'),
        ("Volume status 'Đang xuất kho'",
         f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đang xuất kho",{DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})'),
        ("Volume status 'Đã xuất kho'",
         f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã xuất kho",{DETAIL_SHEET}!{VOL_SHIPPED}:{VOL_SHIPPED})'),
        ("Volume status 'Đang vận chuyển'",
         f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đang vận chuyển",{DETAIL_SHEET}!{VOL_SHIPPED}:{VOL_SHIPPED})'),
        ("Volume status 'Đã vận chuyển'",
         f'=SUMIF({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED})'),
        ("% Done theo 1 Kho cụ thể ($B5)",
         f'=IFERROR(ROUND(SUMIFS({DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED},{DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!B:B,$B5)'
         f'/SUMIF({DETAIL_SHEET}!B:B,$B5,{DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})*100,2),0)'),
        ("% Done theo 1 Khu vực ($B5)",
         f'=IFERROR(ROUND(SUMIFS({DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED},{DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!F:F,$B5)'
         f'/SUMIF({DETAIL_SHEET}!F:F,$B5,{DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})*100,2),0)'),
        ("% Done theo 1 Customer ($B5)",
         f'=IFERROR(ROUND(SUMIFS({DETAIL_SHEET}!{VOL_DELIVERED}:{VOL_DELIVERED},{DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã vận chuyển",{DETAIL_SHEET}!H:H,$B5)'
         f'/SUMIF({DETAIL_SHEET}!H:H,$B5,{DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG})*100,2),0)'),
        ("Đơn theo ngày GI (datetime range)",
         f'=COUNTIFS({DETAIL_SHEET}!I:I,">="&DATE(2026,6,8),{DETAIL_SHEET}!I:I,"<"&DATE(2026,6,9))'),
        ("Drop count (status='Cancel') — col M = status raw",
         f'=COUNTIF({DETAIL_SHEET}!{STATUS_COL}:{STATUS_COL},"Cancel")'),
        ("STM lag check — đơn 'Đã xuất kho' chưa có ATA (col J trống)",
         f'=COUNTIFS({DETAIL_SHEET}!{E2E_COL}:{E2E_COL},"Đã xuất kho",{DETAIL_SHEET}!J:J,"")'),
    ]
    for label, formula in patterns:
        ws.append([label, formula])
        r_ = ws.max_row
        body_cell(ws.cell(row=r_, column=1))
        c = ws.cell(row=r_, column=2)
        c.font = Font(name="Consolas", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        ws.row_dimensions[r_].height = max(20, min(80, 16 + len(formula) // 80 * 14))

    ws.append([])
    ws.append(["Cảnh báo", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="DC2626")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    warnings = [
        ("Sheet rename",
         "Đổi tên sheet 05 → công thức ở 01-04 + 09 sẽ broken. Find&Replace formula bar: "
         "'05 — Detail Orders' → tên mới."),
        ("UOM scope",
         f"Sheet 05 chỉ chứa CSE (UOM={UOM}). Nếu cần test UOM khác (ton/cbm/pallet/do) → re-run script Python "
         f"với UOM = 'ton' (hoặc đổi var UOM trong file)."),
        ("STM exclusion",
         "Sheet 05 KHÔNG loại đơn STM-pending (khác OTIF). 2 status 'Đã xuất kho' + 'Đang vận chuyển' "
         "có thể bao gồm đơn lag STM. Verify với customer định nghĩa."),
        ("L5 Σ parity",
         "Sheet 02 có cột 'Σ check vs L1' — nếu RED (diff > 5%) → bug 2026-05-18 CTE L5 thiếu canonical status filter."),
    ]
    for k, v in warnings:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP
        for c in range(1, 3):
            body_cell(ws.cell(row=ws.max_row, column=c))


UX_CHECKLIST_ITEMS = [
    # (category, what_to_check, expected, severity_if_fail, ref)
    # === UX Visual ===
    ("UX Visual",
     "L1 Hero — % Hoàn thành full-width, RAG color visible HiDPI, value đủ to readable trên 1366×768",
     "Headline ≥ display weight (lớn hơn body-large). RAG color border/background rõ ràng trên retina, không dùng border 0.5px.",
     "Major", "PRD §5.4 L1"),
    ("UX Visual",
     "L1 Hero KHÔNG có delta vs hôm qua / KHÔNG có as-of timestamp",
     "Chỉ value + target + RAG + 3 sub-numbers (Plan/Đã giao/Còn lại). KHÔNG '+5% vs yesterday'. KHÔNG 'Cập nhật lúc HH:mm'.",
     "Major", "PRD §5.4 F2 + G7"),
    ("UX Visual",
     "L1 Alert banner full-width xuất hiện khi % Hoàn thành < 80%",
     "Banner đỏ alert ngay dưới Hero khi metric drop critical.",
     "Major", "PRD §5.4 storytelling"),
    ("UX Visual",
     "L2 Exception 3 ô có headline + list 3-max + dim tên + value",
     "3 ô riêng biệt; ô (a) kho off-target sort worst-first; ô (b) drop có lý do; ô (c) khu vực <95%.",
     "Major", "PRD §5.4 L2"),
    ("UX Visual",
     "L3 Funnel strip 5 status đúng thứ tự STATUS_ORDER (Chưa xuất → Đang xuất → Đã xuất → Đang vận → Đã vận)",
     "5 entry liên tiếp 1 dòng; KHÔNG có 'Kế hoạch xuất' hay 'Thực xuất' trong funnel (đã thay 6 KPI baseline).",
     "Critical", "PRD §5.4 L3 + Spec §20 Drift #11"),
    ("UX Visual",
     "L4 Drop Trend chart — 2 reference lines visible (target ≤5% solid red + rolling 30d dashed grey)",
     "2 lines render đúng style (solid vs dashed); label 'Target ≤5%' visible. KHÔNG nhầm vai 2 lines.",
     "Major", "Spec §6.7"),
    ("UX Visual",
     "L5 4 dim panels render grid 2×2 đồng thời (KHÔNG dùng Tabs)",
     "4 panels visible 1 frame trên ≥1920×1080. KHÔNG có dropdown chuyển panel.",
     "Major", "Memory [[feedback_l5_dimension_panels_over_tabs]]"),
    ("UX Visual",
     "L5 panel bar label 2 dòng — top bold '%pct_done', dưới muted 'done / total uom'",
     "Bar label hiển thị cả % và volume context, KHÔNG chỉ % hoặc chỉ count.",
     "Minor", "Spec §6.8"),
    ("UX Visual",
     "Action title trên mỗi section — insight statement (vd 'Hôm nay 73% — DƯỚI target 95%') thay vì static label",
     "Title đổi theo state Green/Red; KHÔNG dùng cố định 'Tiến độ E2E'.",
     "Minor", "PRD §5.4 storytelling"),
    ("UX Visual",
     "Tier 1 (L1 + L2 + L3) fit 1 fold viewport 1366×768 — không cần scroll để trả lời Q1+Q2",
     "L1 + L2 visible trên fold đầu, L3 strip có thể overflow xuống nhưng vẫn close enough.",
     "Major", "PRD §5.4 + UAT-FLASH-020"),
    # === Filter behavior ===
    ("Filter",
     "Filter Kho dropdown — full options (BKD1, BKD2, BKD3, NKD, VN821, VN831) + ALL",
     "6 kho + ALL. Verify dropdown hiển thị đúng tên, không pseudo.",
     "Major", "PRD §4"),
    ("Filter",
     "Filter Cargo Group dropdown — đúng 7 option (FRESH, DRY, MOONCAKE, POSM/OFFBOM, TEST, PM, EQUIPMENT) + ALL",
     "Đủ 7 type + ALL. Match alpha hoặc by priority.",
     "Minor", "PRD §4"),
    ("Filter",
     "Filter Brand depend Cargo Group — chọn Cargo=DRY → Brand chỉ hiện brands DRY",
     "Brand dropdown filter theo `parentKey: 'group_of_cargo'`. AC-11.",
     "Major", "PRD §4 + AC-11"),
    ("Filter",
     "Filter Region — 10 option đúng (South East, SE-Lam Dong, Ha Noi, Central highland, Mekong 1, HCM, NE-NW, NCC, SCC, Mekong 2) + ALL",
     "Đủ 10 region + ALL. Dropdown sort alpha.",
     "Minor", "PRD §4"),
    ("Filter",
     "Filter UOM — 5 option (cse, ton, cbm, pallet, do) + default = cse",
     "Đủ 5 UOM. KHÔNG có option khác.",
     "Minor", "PRD §3.3 + AC-02"),
    ("Filter",
     "Filter Date Type — 5 option (GI date, Actual Ship date, ETD gửi thầu, ATA đơn, ETA gửi thầu) + default = GI date",
     "5 option. Default GI date.",
     "Minor", "PRD §4 + AC-07"),
    ("Filter",
     "Filter Date Range default = tháng hiện tại tới hôm nay (`thisMonthToTodayRange`)",
     "Default state ngày 1 tháng hiện tại → today.",
     "Minor", "PRD §4 + AC-07"),
    ("Filter",
     "Filter persist sau F5 reload (localStorage `dashboard-widget-filter:{dashboardId}:{widgetId}`)",
     "Quay lại trang giữ filter state. Reset chỉ khi click Reset.",
     "Minor", "PRD §12 + AC-08"),
    ("Filter",
     "Filter combo 5-dim cross response time < 3s",
     "Trace Network tab; nếu > 3s flag perf defect.",
     "Major", "Plan §8 + UAT-FLASH-016"),
    ("Filter",
     "L4 Drop Trend chart date type guard — disable ETD/ETA gửi thầu options khi user xem L4",
     "Chọn ETD/ETA filter overall → L4 ẩn 2 option hoặc fallback GI date (H2 chốt).",
     "Major", "PRD §6.6 H2 + UAT-FLASH-008"),
    # === Interaction ===
    ("Interaction",
     "Click row L5 panel Kho → smooth scroll lên L2 + highlight kho đó",
     "Click kho A → outline + bold row A trong L2 ô (a); row khác giảm opacity 50%.",
     "Minor", "PRD §5.4 + UAT-FLASH-004"),
    ("Interaction",
     "Click 1 DO trong Detail Table → mở Order Monitor với context DO + load < 2s",
     "URL/state có DO code; load < 2s.",
     "Minor", "Spec §5.3"),
    ("Interaction",
     "Nút Export trên mỗi chart — PNG (DOM-to-image) hoặc CSV, filename `flash-daily-{slug}`",
     "Export PNG file size > 0, mở được. Export CSV có header + rows.",
     "Minor", "AC-10"),
    ("Interaction",
     "Hover bar chart L5 → tooltip hiển thị %pct_done + done_volume + total_volume + uom",
     "Tooltip có 4 thông tin context, không bị clipped ngoài viewport.",
     "Cosmetic", "UX standard"),
    ("Interaction",
     "Detail Table T9 — sort theo bất kỳ column nào (SO, ETA, ATA, Status, ...)",
     "Click header → asc; click 2 → desc; sort indicator visible.",
     "Minor", "UX standard"),
    ("Interaction",
     "Đổi UOM filter (cse → ton → cbm → pallet → do) → tất cả L1/L3/L5/T1-T9 refetch + format đổi",
     "5 UOM switch consistent. UOM=do → cột ORIGINAL/SHIPPED/Sản lượng giao = '-'.",
     "Major", "AC-02 + UAT-FLASH-019"),
    ("Interaction",
     "Subtitle 'Số kế hoạch (CBM)' trên chart vẫn dùng CBM bất kể UOM người dùng chọn",
     "Subtitle CBM cố định, không đổi theo UOM bar.",
     "Minor", "Spec §6.4"),
    # === Storytelling (mental model) ===
    ("Storytelling",
     "Q1 — 'Hôm nay đang đi tới đâu?' trả lời được ≤ 5 giây nhìn L1 Hero",
     "% Hoàn thành + RAG color visible ngay, không scroll/decode.",
     "Major", "PRD §5.4 + UAT-FLASH-020"),
    ("Storytelling",
     "Q2 — 'Có rủi ro gì không?' trả lời được ≤ 10 giây từ L2 Exception",
     "Customer chỉ đúng 3 ô (kho/drop/khu vực) off-target.",
     "Major", "PRD §5.4 + UAT-FLASH-020"),
    ("Storytelling",
     "Q3 — 'Đơn đang kẹt ở đâu trong luồng E2E?' trả lời được ≤ 15 giây từ L3 Funnel",
     "Customer spot status có volume lớn nhất + STM lag context tooltip.",
     "Minor", "PRD §5.4"),
    ("Storytelling",
     "Q4 — 'Xu hướng rớt 14 ngày?' trả lời được ≤ 20 giây từ L4 Drop Trend",
     "Customer thấy pattern lên/xuống + so với target ≤5%.",
     "Minor", "PRD §5.4"),
    ("Storytelling",
     "Q5 — 'Chiều nào kéo % xuống nhất?' trả lời được ≤ 20 giây từ L5 4 panels",
     "Customer spot panel có % thấp nhất; KHÔNG cần chuyển Tabs.",
     "Major", "UAT-FLASH-021"),
    # === Master Data ===
    ("Master Data",
     "Customer name (NPP = Customer per Mondelez) hiển thị có dấu tiếng Việt đầy đủ",
     "Verify 5 customer mẫu: 'Cty TNHH ...' có dấu, không bị strip diacritics.",
     "Minor", "Memory [[project_mondelez_npp_eq_customer]] + RF-04"),
    ("Master Data",
     "Khu vực hiển thị đúng 10 enum (không xuất hiện region khác hoặc 'Unclassified' > 5%)",
     "10 region match FE expect; Unclassified < 5%.",
     "Minor", "Q-INS-08 trong ops insight"),
    ("Master Data",
     "Cargo group fill rate ≥ 95% — đơn rỗng cargo_group fallback 'Unclassified'",
     "Verify Unclassified < 5% trong T9 Detail.",
     "Minor", "Q-INS-06"),
    ("Master Data",
     "L5 panel Customer KHÔNG có dropdown NPP/Customer toggle",
     "Per OQ-07 dropped — Mondelez NPP=Customer, KHÔNG cần dropdown phân biệt.",
     "Major", "Memory [[project_mondelez_npp_eq_customer]] + UAT-FLASH-012"),
    ("Master Data",
     "Number format — số đếm có dấu phẩy ngăn ngàn (vd 2,984), % format '93.30%' 2 decimal",
     "Convention VN/EN — chốt với khách.",
     "Cosmetic", "i18n"),
]


def build_ux_checklist_sheet(ws, ctx):
    """Sheet 07 — UX & Filter checklist (customer fill trong session)."""
    ws.title = "07 — UX & Filter checklist"
    set_widths(ws, [6, 20, 48, 48, 14, 12, 16, 30])
    title_row(ws,
              "UX / Filter / Interaction / Storytelling / Master Data checklist — KHÔNG compute từ data, customer fill trong session UAT",
              8)
    header_row(ws, [
        "#", "Category",
        "Mục cần kiểm tra (Quan sát thế nào?)",
        "Expected behavior",
        "Customer\nverdict",
        "Severity\nnếu Fail",
        "Reference\n(PRD/AC/source)",
        "Customer ghi chú",
    ])

    category_fill = {
        "UX Visual":    PatternFill("solid", fgColor="DBEAFE"),
        "Filter":       PatternFill("solid", fgColor="FEF3C7"),
        "Interaction":  PatternFill("solid", fgColor="E0E7FF"),
        "Storytelling": PatternFill("solid", fgColor="FCE7F3"),
        "Master Data":  PatternFill("solid", fgColor="DCFCE7"),
    }
    severity_fill = {"Critical": RED, "Major": YELLOW, "Minor": GREEN, "Cosmetic": GREY}

    for i, (cat, check, expected, severity, ref) in enumerate(UX_CHECKLIST_ITEMS, 1):
        ws.append([i, cat, check, expected, "", severity, ref, ""])
        r_ = ws.max_row
        for col in range(1, 9):
            body_cell(ws.cell(row=r_, column=col))
            ws.cell(row=r_, column=col).alignment = WRAP
        ws.cell(row=r_, column=2).fill = category_fill.get(cat, GREY)
        ws.cell(row=r_, column=2).font = Font(bold=True)
        ws.cell(row=r_, column=5).fill = EMPTY_CUSTOMER
        ws.cell(row=r_, column=5).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r_, column=6).fill = severity_fill.get(severity, GREY)
        ws.cell(row=r_, column=6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_, column=8).fill = EMPTY_CUSTOMER
        max_text = max(len(check), len(expected))
        ws.row_dimensions[r_].height = max(36, min(96, 20 + max_text // 50 * 16))

    ws.freeze_panes = "A3"

    ws.append([])
    ws.append(["Hướng dẫn customer",
               "Mỗi row = 1 mục verify trên dashboard. Customer điền 'Verdict' (Pass / Fail / N/A / cần làm rõ) "
               "+ 'Customer ghi chú' nếu có. Severity cho biết mức độ nếu Fail. Reference trỏ về acceptance "
               "criteria gốc trong PRD."])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)


def build_sql_sheet(ws, sql_map):
    ws.title = "08 — SQL Appendix"
    set_widths(ws, [40, 32, 110])
    title_row(ws, "SQL queries — provenance rõ ràng (canonical từ sql-registry.md OR ad-hoc)", 3)
    header_row(ws, ["Query ID — mục đích", "Source / provenance", "SQL (đã substitute placeholder)"])
    for qid, (q, src) in sql_map.items():
        ws.append([qid, src, q])
        r_ = ws.max_row
        ws.cell(row=r_, column=1).font = Font(bold=True)
        ws.cell(row=r_, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        src_cell = ws.cell(row=r_, column=2)
        src_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if src.startswith("sql-registry.md"):
            src_cell.fill = GREEN
            src_cell.font = Font(bold=True)
        else:
            src_cell.fill = YELLOW
        for col in (1, 2, 3):
            ws.cell(row=r_, column=col).border = BORDER
        sql_cell = ws.cell(row=r_, column=3)
        sql_cell.font = Font(name="Consolas", size=9)
        sql_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_].height = max(40, min(480, 18 + q.count("\n") * 13))


def build_t7_t8_sheet(ws, ctx):
    """Sheet 09 — T7 Drop Report bucket (frozen, delivery_to_customer not in Detail) +
    T8 Drop Reason (DUAL với formula compute từ Detail col S remark_2)."""
    ws.title = "09 — T7 Drop & T8 Reason"
    set_widths(ws, [40, 14, 14, 14, 14, 14, 14, 24])
    title_row(ws,
              "T7 Drop Report (frozen — delivery_to_customer field không có trong Detail) + "
              "T8 Drop Reason (DUAL via remark_2 col S). 2 unit độc lập DRY&FRESH (cse) + POSM (pc).",
              8)

    # T7 — frozen only with rationale
    ws.append(["T7 — Drop Report bucket (FROZEN ONLY)"])
    last = ws.max_row
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=8)
    ws.cell(row=last, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=last, column=1).fill = PatternFill("solid", fgColor="14283F")
    header_row(ws, [
        "Bucket", "DRY&FRESH\n(cse SQL)", "POSM\n(pc SQL)",
        "% DRY&FRESH\n(SQL)", "% POSM\n(SQL)",
        "", "", "Khách ghi chú",
    ])
    if not ctx["t7_drop"]:
        ws.append(["(không có data)"] + [""] * 7)
    else:
        for row in ctx["t7_drop"]:
            first_key = list(row.keys())[0]
            ws.append([
                row.get(first_key, ""),
                row.get("dry_fresh_cse") or 0,
                row.get("posm_pc") or 0,
                row.get("pct_dry_fresh_cse"),
                row.get("pct_posm_pc"),
                "", "", "",
            ])
            r_ = ws.max_row
            for col in range(1, 9):
                body_cell(ws.cell(row=r_, column=col))
            ws.cell(row=r_, column=2).number_format = "#,##0.00"
            ws.cell(row=r_, column=3).number_format = "#,##0.00"
            ws.cell(row=r_, column=4).number_format = "0\"%\""
            ws.cell(row=r_, column=5).number_format = "0\"%\""
    ws.append(["", "Lý do: field `delivery_to_customer` không có trong mv_flash_and_drop_report → "
               "Detail sheet 05 không thể chứa cột này → formula compute lại không feasible. Customer "
               "verify trực tiếp bằng SQL trong sheet 08 SQL Appendix.", "", "", "", "", "", ""])
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)
    ws.row_dimensions[ws.max_row].height = 36

    ws.append([])

    # T8 — DUAL with formula via Detail col S (remark_2)
    ws.append(["T8 — Drop Reason (DUAL via Detail!S remark_2)"])
    last = ws.max_row
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=8)
    ws.cell(row=last, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=last, column=1).fill = PatternFill("solid", fgColor="14283F")
    header_row(ws, [
        "Lý do (remark_2)",
        "DRY&FRESH\n(cse SQL)", "DRY&FRESH\n(cse formula)",
        "POSM\n(pc SQL)", "POSM\n(pc formula)",
        "Count đơn\n(SQL)", "Count đơn\n(formula)",
        "Khách ghi chú",
    ])
    if not ctx["t8_reason"]:
        ws.append(["(không có data)"] + [""] * 7)
    else:
        for row in ctx["t8_reason"]:
            first_key = list(row.keys())[0]
            reason = row.get(first_key, "") or "(rỗng)"
            dry_sql = row.get("dry_fresh_cse") or 0
            posm_sql = row.get("posm_pc") or 0
            # Formula: SUMIFS Detail!P:P where Status=Cancel + Remark=reason
            # POSM formula proxy uses Q (shipped_cse) — for cse-vs-pc differentiation, we'd need
            # a uom flag in Detail. Keep formula limited to cse total via VOL_ORIG.
            r_target = ws.max_row + 1
            dry_formula = (f'=SUMIFS({DETAIL_SHEET}!{VOL_ORIG}:{VOL_ORIG},'
                           f'{DETAIL_SHEET}!{STATUS_COL}:{STATUS_COL},"Cancel",'
                           f'{DETAIL_SHEET}!{REMARK_COL}:{REMARK_COL},$A{r_target})')
            # POSM pc — no direct mapping in Detail (cargo_group used as proxy)
            posm_formula = "n/a (POSM pc mapping cần cargo_group filter)"
            count_formula = (f'=COUNTIFS({DETAIL_SHEET}!{STATUS_COL}:{STATUS_COL},"Cancel",'
                             f'{DETAIL_SHEET}!{REMARK_COL}:{REMARK_COL},$A{r_target})')
            ws.append([reason, dry_sql, dry_formula, posm_sql, posm_formula, "", count_formula, ""])
            r_ = ws.max_row
            for col in range(1, 9):
                body_cell(ws.cell(row=r_, column=col))
            for col in (2, 3, 4):
                ws.cell(row=r_, column=col).number_format = "#,##0.00"
            ws.cell(row=r_, column=5).font = Font(italic=True, color="808080")
            ws.cell(row=r_, column=5).number_format = "@"
            for col in (6, 7):
                ws.cell(row=r_, column=col).number_format = "#,##0"

    ws.append([
        "Note T8 dual",
        "T8 reason formula tính `SUMIFS Detail!P (Original CSE) WHERE Status=Cancel AND Remark_2=match`. "
        "POSM (pc unit) formula KHÔNG feasible từ Detail vì Detail dùng cse-only volume cols — UAT verify "
        "POSM qua SQL appendix sheet 08. DRY&FRESH SQL frozen vs formula phải khớp ±0.01.",
    ])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)
    ws.row_dimensions[ws.max_row].height = 48

    ws.append([])
    ws.append([
        "Note 2 unit",
        "DRY&FRESH (cse) + POSM (pc) là 2 unit độc lập, KHÔNG sum. Render 'N cse · M pc' riêng "
        "(per memory [[feedback_mondelez_dropreport_cse_pc_split]]). Bucket pattern match substring "
        "fragile — xem ops insight RF-02.",
    ])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)
    ws.row_dimensions[ws.max_row].height = 48


# ── Main ─────────────────────────────────────────────────
def main():
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"[*] Connect CH…")
    client = get_client()
    print(f"[*] Run queries window {WINDOW_START_DATE}→{WINDOW_END_DATE} ({DATE_TYPE_LABEL}, UOM={UOM})…")

    ctx = {
        "run_ts": datetime.now(timezone(timedelta(hours=7))).strftime("%Y-%m-%d %H:%M UTC+7"),
        "server_now": run_mv_now(client),
    }

    print(f"[*] Query mv_filter_* for ALL substitution…")
    filters = query_filter_values(client)
    ctx["filters"] = filters

    funnel, funnel_sqls = run_l1_l3_funnel(client, filters)
    ctx["funnel"] = funnel
    plan = float(funnel["Tổng Volume"] or 0)
    done = float(funnel["Đã vận chuyển"] or 0)
    print(f"    L1 Hero: Plan={plan:,.2f} Done={done:,.2f} pct_done={(100*done/plan if plan>0 else 0):.2f}%")

    kho, sql_kho, src_kho = run_l5_panel(client, filters, "Kho"); ctx["dim_kho"] = kho
    kv, sql_kv, src_kv = run_l5_panel(client, filters, "Khu vực"); ctx["dim_khu_vuc"] = kv
    cust, sql_cust, src_cust = run_l5_panel(client, filters, "Customer"); ctx["dim_customer"] = cust
    kenh, sql_kenh, src_kenh = run_l5_panel(client, filters, "Kênh"); ctx["dim_kenh"] = kenh
    print(f"    L5: kho={len(kho)} khu_vuc={len(kv)} customer={len(cust)} kenh={len(kenh)}")

    l2_kho, sql_l2k, src_l2k = run_l2_hotspot_kho(client, filters); ctx["l2_kho"] = l2_kho
    l2_drop, sql_l2d, src_l2d = run_l2_hotspot_drop(client, filters); ctx["l2_drop"] = l2_drop
    l2_region, sql_l2r, src_l2r = run_l2_hotspot_region(client, filters); ctx["l2_region"] = l2_region

    l4, sql_l4, src_l4 = run_l4_drop_trend(client, filters); ctx["l4_trend"] = l4
    print(f"    L4 Drop Trend: {len(l4)} ngày")

    try:
        t7, sql_t7, src_t7 = run_t7_drop_report(client, filters)
    except Exception as e:
        print(f"    [warn] T7 Drop Report skipped: {e}")
        t7, sql_t7, src_t7 = [], "-- skipped due to error --", f"error: {str(e)[:200]}"
    ctx["t7_drop"] = t7

    try:
        t8, sql_t8, src_t8 = run_t8_drop_reason(client, filters)
    except Exception as e:
        print(f"    [warn] T8 Drop Reason skipped: {e}")
        t8, sql_t8, src_t8 = [], "-- skipped due to error --", f"error: {str(e)[:200]}"
    ctx["t8_reason"] = t8

    detail, sql_detail, src_detail = run_detail(client, filters); ctx["detail"] = detail
    print(f"    Detail: {len(detail):,} rows")

    sql_map = {}
    for status, sql, src in funnel_sqls:
        sql_map[f"Q-{status}"] = (sql, src)
    sql_map["Q-L5-Kho (registry §Báo cáo tổng hợp theo kho hệ thống)"] = (sql_kho, src_kho)
    sql_map["Q-L5-Khu vực (registry §Báo cáo tổng hợp theo khu vực)"] = (sql_kv, src_kv)
    sql_map["Q-L5-Customer (registry §Báo cáo tổng hợp theo NPP)"] = (sql_cust, src_cust)
    sql_map["Q-L5-Kênh (registry §Báo cáo tổng hợp theo kênh bán hàng)"] = (sql_kenh, src_kenh)
    sql_map["Q-L2-Kho (registry §L2 Điểm nóng — Kho)"] = (sql_l2k, src_l2k)
    sql_map["Q-L2-Drop (registry §L2 Điểm nóng — Drop + Lý do)"] = (sql_l2d, src_l2d)
    sql_map["Q-L2-Region (registry §L2 Điểm nóng — Khu vực)"] = (sql_l2r, src_l2r)
    sql_map["Q-L4-Drop Trend 14d (registry §L4 Trend tỷ lệ rớt 14 ngày)"] = (sql_l4, src_l4)
    sql_map["Q-T7-Drop bucket (registry §Bổ sung report hàng rớt)"] = (sql_t7, src_t7)
    sql_map["Q-T8-Drop reason (registry §Bổ sung report lý do rớt đơn)"] = (sql_t8, src_t8)
    sql_map["Q-T9-Detail orders (ad-hoc — widget detailTable config)"] = (sql_detail, src_detail)

    print(f"[*] Build workbook…")
    wb = Workbook()
    build_readme(wb.active, ctx)
    build_l1_l3_sheet(wb.create_sheet(), ctx)
    build_l5_sheet(wb.create_sheet(), ctx)
    build_l2_sheet(wb.create_sheet(), ctx)
    build_l4_sheet(wb.create_sheet(), ctx)
    build_detail_sheet(wb.create_sheet(), ctx)
    build_formula_guide_sheet(wb.create_sheet(), ctx)
    build_ux_checklist_sheet(wb.create_sheet(), ctx)
    build_sql_sheet(wb.create_sheet(), sql_map)
    build_t7_t8_sheet(wb.create_sheet(), ctx)

    wb.save(OUT_FILE)
    print(f"[OK] Wrote {OUT_FILE}")


if __name__ == "__main__":
    main()
