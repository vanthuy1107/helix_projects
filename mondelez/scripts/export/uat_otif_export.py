"""
UAT OTIF export — sinh file Excel test case với số SQL thật cho khách hàng confirm.

Window: 2026-05-18 → 2026-05-22 (5 ngày).
Filter mặc định: Loại ngày = "ETA gửi thầu (đơn)" → cột eta_giao_hang_cho_npp.
Source: analytics_workspace.mv_otif (Mondelez tenant CH).

SQL provenance:
  • KPI hero, 4 dim (kho/khu vực/kênh/NVC), Trend, Fail Ontime classifier — load CANONICAL từ
    `projects/mondelez/02-data/data-sources/sql-registry.md` (OTIF section). KHÔNG tự sinh.
    Khi registry update, script tự pick up không cần sửa code.
  • Loại hàng dim (chartByCategory), Detail Orders, Fail Infull classifier — ad-hoc (registry
    không có hoặc đây là FE-side classifier). Comment rõ ở từng function.

Output: projects/mondelez/01-sections/otif/uat/otif-uat-numbers-2026-05-18_to_2026-05-22.xlsx
"""
from __future__ import annotations

import os
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_DIR = next(p for p in Path(__file__).resolve().parents
                    if (p / "da.toml").exists())  # tenant root = nơi có da.toml (relocation-proof)
_OUTPUT_DIR = _PROJECT_DIR / "01-sections" / "otif" / "uat"

try:
    from dotenv import load_dotenv
    _env = _PROJECT_DIR / ".env"
    if _env.exists():
        load_dotenv(_env)
    else:
        load_dotenv()
except ImportError:
    pass

import clickhouse_connect
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _pct2(v):
    """Normalize % values → Decimal('XX.XX') với scale=2 fixed (trailing zero preserved).
    Excel cell raw value = đúng 2 decimal places, consistent giữa các metric."""
    if v is None:
        return None
    return Decimal(str(round(float(v), 2))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _num2(v):
    """Normalize số có .XX (CSE) → Decimal scale=2."""
    if v is None:
        return None
    return Decimal(str(round(float(v), 2))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# ── Config ───────────────────────────────────────────────
WINDOW_START = "2026-05-18 00:00:00"
WINDOW_END = "2026-05-22 23:59:59"
WINDOW_START_DATE = "2026-05-18"
WINDOW_END_DATE = "2026-05-22"
DATE_TYPE_LABEL = "ETA gửi thầu (đơn)"
DATE_COL = "eta_giao_hang_cho_npp"
SOURCE = "analytics_workspace.mv_otif"
EXCLUDE_CLAUSE = "otif_status != 'Không có dữ liệu STM'"
TENANT = "Mondelez"

OUT_FILE = _OUTPUT_DIR / "otif-uat-numbers-2026-05-18_to_2026-05-22.xlsx"

# RAG bands per PRD §13.2
RAG = {
    "otif":   {"target": 90.0, "green": 90.0, "yellow_lo": 85.0},
    "ontime": {"target": 95.0, "green": 95.0, "yellow_lo": 90.0},
    "infull": {"target": 97.0, "green": 97.0, "yellow_lo": 92.0},
}


def get_client():
    return clickhouse_connect.get_client(
        host=os.getenv("CLICKHOUSE_HOST", ""),
        port=int(os.getenv("CLICKHOUSE_PORT", "8443")),
        username=os.getenv("CLICKHOUSE_USER", ""),
        password=os.getenv("CLICKHOUSE_PASSWORD", ""),
        secure=os.getenv("CLICKHOUSE_SECURE", "true").lower() in ("1", "true", "yes"),
        connect_timeout=15,
        send_receive_timeout=60,
    )


# ── Registry SQL loader + filter substitution ────────────
REGISTRY_PATH = _PROJECT_DIR / "02-data" / "data-sources" / "sql-registry.md"
_REGISTRY_TEXT_CACHE = None


def _registry_text():
    global _REGISTRY_TEXT_CACHE
    if _REGISTRY_TEXT_CACHE is None:
        _REGISTRY_TEXT_CACHE = REGISTRY_PATH.read_text(encoding="utf-8")
    return _REGISTRY_TEXT_CACHE


def load_registry_sql(section_title):
    """Đọc sql-registry.md, tìm `### <section_title>` trong scope `## OTIF`, lấy block ClickHouse SQL.
    Trả về (sql, line_ref) — sql là raw string với placeholder, line_ref dùng cho appendix.
    """
    import re
    text = _registry_text()
    # Locate ## OTIF section start
    otif_match = re.search(r"^## OTIF\b", text, re.MULTILINE)
    if not otif_match:
        raise RuntimeError("sql-registry.md không có `## OTIF` section")
    otif_start = otif_match.start()
    # Pattern: ### <title> ... **ClickHouse SQL:** ```sql ... ```
    # Bám sát format của registry: backtick `New` optional
    pat = rf"^###\s+{re.escape(section_title)}\b.*?\*\*ClickHouse SQL:\*\*\s*```sql\s*\n(.*?)\n```"
    m = re.search(pat, text[otif_start:], re.MULTILINE | re.DOTALL)
    if not m:
        raise RuntimeError(f"Không tìm được section '{section_title}' trong OTIF (ClickHouse block)")
    sql = m.group(1).strip()
    # Compute absolute line number of ```sql start (offset-by-1)
    abs_offset = otif_start + m.start(1)
    line_no = text[:abs_offset].count("\n") + 1
    return sql, line_no


def query_filter_values(client):
    """Query mv_filter_* để lấy full list giá trị (dùng để substitute placeholder khi ALL filter,
    match đúng bypass-detect pattern của registry SQL).
    """
    out = {}
    src = [
        ("whseid", "whseid", "mv_filter_warehouse"),
        ("group_of_cargo", "group_of_cargo_code", "mv_filter_cargo_brand"),
        ("transporter", "vendor_code", "mv_filter_vendor"),
        ("area", "group_area_code", "mv_filter_region"),
    ]
    for key, col, tbl in src:
        rows = client.query(
            f"SELECT DISTINCT {col} FROM analytics_workspace.{tbl} WHERE {col} IS NOT NULL ORDER BY 1"
        ).result_rows
        # Escape single quotes, format as quoted CSV list
        vals = ",".join("'" + str(r[0]).replace("'", "''") + "'" for r in rows)
        out[key] = vals
    return out


def substitute_placeholders(sql, filters, date_type, from_date, to_date):
    """Substitute {{...}} placeholders + strip [[ ]] optional brackets, giữ semantics 100%."""
    sql = sql.replace("{{whseid}}", filters["whseid"])
    sql = sql.replace("{{group_of_cargo}}", filters["group_of_cargo"])
    sql = sql.replace("{{transporter}}", filters["transporter"])
    sql = sql.replace("{{area}}", filters["area"])
    sql = sql.replace("{{date_type}}", f"'{date_type}'")
    sql = sql.replace("{{from_date}}", f"'{from_date}'")
    sql = sql.replace("{{to_date}}", f"'{to_date}'")
    sql = sql.replace("[[", "").replace("]]", "")
    return sql


# ── Excel formula helpers ────────────────────────────────
DETAIL_SHEET = "'05 — Detail Orders'"
# Detail Orders column letters (fixed by build_detail_sheet order):
# A=SO B=Kho code C=Kho group D=Loại hàng E=Kênh F=NVC G=Khu vực
# H=Customer code I=Customer name J=ETA(UTC+7) K=ATA(UTC+7)
# L=Planned CSE M=Shipped CSE N=Delivered CSE
# O=Ontime status P=Infull status Q=OTIF status
DETAIL_DIM_COL = {
    "NVC": "F",
    "Kho": "C",
    "Loại hàng": "D",
    "Kênh": "E",
    "Khu vực": "G",
}


def f_total_orders():
    """Tổng đơn = OTIF + Failed OTIF (đã loại STM ở Detail nguồn)."""
    return f'=COUNTIF({DETAIL_SHEET}!Q:Q,"OTIF")+COUNTIF({DETAIL_SHEET}!Q:Q,"Failed OTIF")'


def f_count_status(col, status):
    return f'=COUNTIF({DETAIL_SHEET}!{col}:{col},"{status}")'


def f_pct_status(col, status):
    """% = count(status) / total × 100, round 2."""
    return (f'=ROUND(COUNTIF({DETAIL_SHEET}!{col}:{col},"{status}")'
            f'/(COUNTIF({DETAIL_SHEET}!Q:Q,"OTIF")+COUNTIF({DETAIL_SHEET}!Q:Q,"Failed OTIF"))*100,2)')


def f_dim_total(dim_col, value_cell):
    return f'=COUNTIF({DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell})'


def f_dim_status(dim_col, value_cell, status_col, status):
    return (f'=COUNTIFS({DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell},'
            f'{DETAIL_SHEET}!{status_col}:{status_col},"{status}")')


def f_dim_pct(dim_col, value_cell, status_col, status):
    return (f'=IFERROR(ROUND(COUNTIFS({DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell},'
            f'{DETAIL_SHEET}!{status_col}:{status_col},"{status}")'
            f'/COUNTIF({DETAIL_SHEET}!{dim_col}:{dim_col},{value_cell})*100,2),0)')


def _date_range_args(date_str):
    """Excel COUNTIFS args cho ETA datetime trên 1 ngày — return tuple bound expressions.
    date_str format: 'YYYY-MM-DD'. Bounds: >= DATE(...), < DATE(next_day)."""
    y, m, d = date_str.split("-")
    next_d_excel = f'DATE({y},{m},{d})+1'
    return f'">="&DATE({y},{m},{d}),{DETAIL_SHEET}!J:J,"<"&{next_d_excel}'


def f_trend_count(date_str, status_col, status):
    rng = _date_range_args(date_str)
    return (f'=COUNTIFS({DETAIL_SHEET}!J:J,{rng},'
            f'{DETAIL_SHEET}!{status_col}:{status_col},"{status}")')


def f_trend_total(date_str):
    rng = _date_range_args(date_str)
    return f'=COUNTIFS({DETAIL_SHEET}!J:J,{rng})'


def f_trend_pct(date_str, status_col, status):
    rng = _date_range_args(date_str)
    return (f'=IFERROR(ROUND(COUNTIFS({DETAIL_SHEET}!J:J,{rng},'
            f'{DETAIL_SHEET}!{status_col}:{status_col},"{status}")'
            f'/COUNTIFS({DETAIL_SHEET}!J:J,{rng})*100,2),0)')


# ── Styles ───────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="606060")
BODY_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="BFC9D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
GREEN = PatternFill("solid", fgColor="DCFCE7")
YELLOW = PatternFill("solid", fgColor="FEF3C7")
RED = PatternFill("solid", fgColor="FEE2E2")
GREY = PatternFill("solid", fgColor="F1F5F9")
EMPTY_CUSTOMER = PatternFill("solid", fgColor="FFFBEB")


def rag_fill(metric: str, value):
    if value is None:
        return GREY
    cfg = RAG[metric]
    if value >= cfg["green"]:
        return GREEN
    if value >= cfg["yellow_lo"]:
        return YELLOW
    return RED


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_row(ws, text, span):
    ws.append([text] + [""] * (span - 1))
    last = ws.max_row
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=span)
    c = ws.cell(row=last, column=1)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[last].height = 24


def header_row(ws, headers):
    ws.append(headers)
    last = ws.max_row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=last, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[last].height = 28


def body_cell(c, value=None, fill=None, fmt=None, align=None):
    c.font = BODY_FONT
    c.border = BORDER
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if value is not None:
        c.value = value


# ── Queries — load canonical SQL từ sql-registry.md ──────
PCT_FIELDS = ("pct_otif", "pct_ontime", "pct_infull")


def _normalize_row(d):
    """Apply _pct2 cho mọi field % trong dict — đảm bảo consistent 2-decimal storage."""
    for k in PCT_FIELDS:
        if k in d:
            d[k] = _pct2(d[k])
    return d


def run_kpi(client, filters):
    """KPI hero — load registry "Tổng đơn" (CH). Source: sql-registry.md §OTIF."""
    tpl, line_ref = load_registry_sql("Tổng đơn")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    row = dict(zip(r.column_names, r.result_rows[0]))
    return _normalize_row(row), sql, f"sql-registry.md:{line_ref}"


def run_fail_ontime(client, filters):
    """Fail Ontime — registry "Phân rã nguyên nhân fail ontime" (CH). Bucket theo not_ontime_reason raw column."""
    tpl, line_ref = load_registry_sql("Phân rã nguyên nhân fail ontime")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [dict(zip(r.column_names, row)) for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_fail_infull(client, filters):
    """Fail Infull — registry "Phân rã nguyên nhân fail infull" (CH)."""
    tpl, line_ref = load_registry_sql("Phân rã nguyên nhân fail infull")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [dict(zip(r.column_names, row)) for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


REGISTRY_DIM_SECTION = {
    "kho":     "OTIF/ Ontime/ Infull theo kho",
    "khu_vuc": "OTIF/ Ontime/ Infull theo khu vực",
    "kenh":    "OTIF/ Ontime/ Infull theo kênh bán hàng",
    "nvc":     "OTIF/ Ontime/ Infull theo nhà vận tải",
}


def run_by_registry_dim(client, dim_key, filters):
    """Load registry SQL cho 1 trong 4 dim chart. Trả về (rows, sql, source_ref)."""
    section = REGISTRY_DIM_SECTION[dim_key]
    tpl, line_ref = load_registry_sql(section)
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    rows = [_normalize_row(dict(zip(r.column_names, row))) for row in r.result_rows]
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_by_loai_hang(client, filters):
    """Loại hàng (chartByCategory) — KHÔNG có trong registry.
    Ad-hoc theo cùng pattern dim khác để consistency. Note rõ trong SQL appendix.
    """
    sql = f"""
WITH filtered_data AS (
    SELECT t.*
    FROM analytics_workspace.mv_otif AS t
    WHERE 1 = 1
    AND if(arraySort([{filters['whseid']}]) = (SELECT arraySort(groupArray(DISTINCT whseid)) FROM analytics_workspace.mv_filter_warehouse), 1=1, t.whseid IN ({filters['whseid']}))
    AND if(arraySort([{filters['group_of_cargo']}]) = (SELECT arraySort(groupArray(DISTINCT group_of_cargo_code)) FROM analytics_workspace.mv_filter_cargo_brand), 1=1, coalesce(t.group_of_cago, 'Unclassified') IN ({filters['group_of_cargo']}))
    AND if(arraySort([{filters['transporter']}]) = (SELECT arraySort(groupArray(DISTINCT vendor_code)) FROM analytics_workspace.mv_filter_vendor), 1=1, t.ten_ngan_nha_van_tai IN ({filters['transporter']}))
    AND if(arraySort([{filters['area']}]) = (SELECT arraySort(groupArray(DISTINCT group_area_code)) FROM analytics_workspace.mv_filter_region), 1=1, t.khu_vuc_doi_xe IN ({filters['area']}))
    AND ( toDate(
      CASE
        WHEN '{DATE_TYPE_LABEL}' = 'ETA gửi thầu (đơn)' THEN t.eta_giao_hang_cho_npp
        WHEN '{DATE_TYPE_LABEL}' = 'ATA chi tiết chuyến' THEN t.ata_den
      END)
      BETWEEN toDate('{WINDOW_START_DATE}') AND toDate('{WINDOW_END_DATE}')
    )
)
SELECT
    coalesce(group_of_cago, 'Unclassified')                            AS loai_hang,
    COUNT(so)                                                          AS total_so,
    countIf(ontime_status = 'Ontime')                                  AS ontime_so,
    round(100.0 * countIf(ontime_status = 'Ontime') / nullIf(COUNT(so), 0), 2) AS pct_ontime,
    countIf(infull_status = 'Infull')                                  AS infull_so,
    round(100.0 * countIf(infull_status = 'Infull') / nullIf(COUNT(so), 0), 2) AS pct_infull,
    countIf(otif_status = 'OTIF')                                      AS otif_so,
    round(100.0 * countIf(otif_status = 'OTIF')    / nullIf(COUNT(so), 0), 2) AS pct_otif
FROM filtered_data
WHERE ontime_status != 'Không có dữ liệu STM'
GROUP BY loai_hang
ORDER BY loai_hang
    """
    r = client.query(sql)
    rows = [_normalize_row(dict(zip(r.column_names, row))) for row in r.result_rows]
    return rows, sql.strip(), "ad-hoc (registry không có chartByCategory; pattern bám registry kho/khu vực)"


def run_detail(client, filters):
    """Detail DO-level — KHÔNG có trong registry (widget detailTable config nằm trong widget.config DB).
    Ad-hoc nhưng bám pattern filter của registry để consistent. Note rõ trong SQL appendix.
    """
    q = f"""
    SELECT
      so                                                       AS so,
      whseid                                                   AS kho_code,
      multiIf(
        whseid IN ('BKD1','BKD2','BKD3'), 'BKD',
        whseid = 'NKD',                    'NKD',
        whseid = 'VN821',                  'Kho ngoài BKD',
        whseid = 'VN831',                  'Kho ngoài NKD',
        concat('(khác) ', whseid)
      )                                                        AS kho_group,
      if(group_of_cago = '' OR group_of_cago IS NULL, '(rỗng)', group_of_cago) AS loai_hang,
      if(group_name = '', '(rỗng)', group_name)                AS kenh,
      if(ten_ngan_nha_van_tai = '', '(rỗng)', ten_ngan_nha_van_tai) AS nvc,
      if(khu_vuc_doi_xe = '', '(rỗng)', khu_vuc_doi_xe)        AS khu_vuc,
      coalesce(customer_code, '')                              AS customer_code,
      coalesce(customer_name, '')                              AS customer_name,
      toDateTime(eta_giao_hang_cho_npp, 'Asia/Ho_Chi_Minh')    AS eta_utc7,
      toDateTime(ata_den, 'Asia/Ho_Chi_Minh')                  AS ata_utc7,
      toFloat64(coalesce(sum_original_cse, 0))                 AS planned_cse,
      toFloat64(coalesce(sum_shipped_cse, 0))                  AS shipped_cse,
      toFloat64(coalesce(sum_san_luong_giao_cse, 0))           AS delivered_cse,
      coalesce(ontime_status, '')                              AS ontime_status,
      coalesce(infull_status, '')                              AS infull_status,
      otif_status                                              AS otif_status,
      coalesce(not_ontime_reason, '')                          AS not_ontime_reason,
      coalesce(not_infull_reason, '')                          AS not_infull_reason
    FROM analytics_workspace.mv_otif
    WHERE 1 = 1
    AND if(arraySort([{filters['whseid']}]) = (SELECT arraySort(groupArray(DISTINCT whseid)) FROM analytics_workspace.mv_filter_warehouse), 1=1, whseid IN ({filters['whseid']}))
    AND if(arraySort([{filters['group_of_cargo']}]) = (SELECT arraySort(groupArray(DISTINCT group_of_cargo_code)) FROM analytics_workspace.mv_filter_cargo_brand), 1=1, coalesce(group_of_cago, 'Unclassified') IN ({filters['group_of_cargo']}))
    AND if(arraySort([{filters['transporter']}]) = (SELECT arraySort(groupArray(DISTINCT vendor_code)) FROM analytics_workspace.mv_filter_vendor), 1=1, ten_ngan_nha_van_tai IN ({filters['transporter']}))
    AND if(arraySort([{filters['area']}]) = (SELECT arraySort(groupArray(DISTINCT group_area_code)) FROM analytics_workspace.mv_filter_region), 1=1, khu_vuc_doi_xe IN ({filters['area']}))
    AND ( toDate(
      CASE
        WHEN '{DATE_TYPE_LABEL}' = 'ETA gửi thầu (đơn)' THEN eta_giao_hang_cho_npp
        WHEN '{DATE_TYPE_LABEL}' = 'ATA chi tiết chuyến' THEN ata_den
      END)
      BETWEEN toDate('{WINDOW_START_DATE}') AND toDate('{WINDOW_END_DATE}')
    )
    AND ontime_status != 'Không có dữ liệu STM'
    ORDER BY eta_giao_hang_cho_npp ASC, so ASC
    """
    r = client.query(q)
    rows = []
    for row in r.result_rows:
        d = dict(zip(r.column_names, row))
        # Normalize CSE → 2 decimals (consistent display + reliable L=M comparison cho SUMPRODUCT)
        for k in ("planned_cse", "shipped_cse", "delivered_cse"):
            d[k] = _num2(d.get(k))
        # Strip tzinfo từ datetime (CH trả tz-aware, Excel không support)
        for k in ("eta_utc7", "ata_utc7"):
            v = d.get(k)
            if v is not None and hasattr(v, "tzinfo") and v.tzinfo is not None:
                d[k] = v.replace(tzinfo=None)
        rows.append(d)
    return rows, q.strip(), "ad-hoc (detailTable widget config — đồng pattern filter với registry)"


def run_trend(client, filters):
    """Trend daily — load registry "%OTIF và khối lượng đơn theo thời gian" (CH).
    Lưu ý: registry SQL chỉ trả % OTIF (1 metric); để có % Ontime/Infull cho RAG color → enrich thêm.
    """
    tpl, line_ref = load_registry_sql("%OTIF và khối lượng đơn theo thời gian")
    sql = substitute_placeholders(tpl, filters, DATE_TYPE_LABEL, WINDOW_START_DATE, WINDOW_END_DATE)
    r = client.query(sql)
    cols = list(r.column_names)
    rows = []
    for row in r.result_rows:
        d = dict(zip(cols, row))
        rows.append({
            "d": d.get("day"),
            "total_so": d.get("total_so"),
            "otif_so": d.get("otif_so"),
            "pct_otif": _pct2(d.get("pct_otif")),
            # Registry trend chỉ có pct_otif — Ontime/Infull tính bổ sung qua formula trong Excel
            "ontime_so": None, "infull_so": None,
            "pct_ontime": None, "pct_infull": None,
        })
    return rows, sql, f"sql-registry.md:{line_ref}"


def run_mv_refresh(client):
    """Lần MV refresh gần nhất (xấp xỉ qua max timestamp INSERT — fallback)."""
    try:
        q = "SELECT now('Asia/Ho_Chi_Minh') AS server_now"
        r = client.query(q)
        return r.result_rows[0][0]
    except Exception:
        return None


# ── Excel build ──────────────────────────────────────────
def build_readme(ws, ctx):
    ws.title = "00 — README"
    set_widths(ws, [22, 80])
    title_row(ws, "UAT OTIF — Số SQL thật cho khách Mondelez confirm", 2)

    meta = [
        ("Section", "OTIF — Smartlog Control Tower"),
        ("Tenant", TENANT),
        ("Window test", f"{WINDOW_START_DATE} → {WINDOW_END_DATE} (5 ngày)"),
        ("Loại ngày", f'{DATE_TYPE_LABEL}  (cột MV: {DATE_COL})'),
        ("Filter còn lại", "ALL (Kho/Khu vực/Loại hàng/NVC/Kênh không chọn)"),
        ("Nguồn dữ liệu", SOURCE),
        ("Loại trừ canonical", "ontime_status != 'Không có dữ liệu STM' (per registry §OTIF)"),
        ("SQL provenance", "KPI hero + 4 dim chart (Kho/Khu vực/Kênh/NVC) + Trend + Fail Ontime/Infull breakdown — load CANONICAL từ sql-registry.md (xem sheet 07). Loại hàng dim + Detail Orders — ad-hoc (registry không có hoặc widget-specific). Khi registry update, re-run script để pick up."),
        ("KPI vs Detail row count", "KPI Hero dùng uniqExact(SO) — đếm sales order unique (= 2984). Detail Orders sheet có 2986 row vì 1 SO có thể split sang nhiều whseid (mv_otif key = SO+whseid). Diff ~0.07% chấp nhận được, không phải bug."),
        ("Format chuẩn (consistent)", "% metric: số decimal places fixed = 2, format '0.00\"%\"' (display 93.30%, 82.94%, ...). Count: format '#,##0' (display 2,986). CSE / số lượng: format '#,##0.00' (display 110.00, 1,326.50, ...). Datetime ETA/ATA: format 'yyyy-mm-dd hh:mm:ss' (datetime cell, sort + filter range OK)."),
        ("Lưu ý raw value vs display", "Excel store float — nếu giá trị có trailing zero (vd 93.30 → raw 93.3), formula bar hiện 93.3 nhưng DISPLAY trong cell vẫn 93.30% nhờ number_format. Logic equivalent: 93.3 == 93.30. Đây là Excel quirk, KHÔNG phải bug."),
        ("Run timestamp", ctx["run_ts"]),
        ("CH server now (UTC+7)", str(ctx["server_now"]) if ctx["server_now"] else "n/a"),
        ("Số đơn (sau loại trừ)", f"{ctx['kpi']['total_so']:,} đơn"),
        ("", ""),
        ("Mục đích file", "PM/BA mang số SQL thật xuống cho khách Mondelez confirm 3 việc: (1) Định nghĩa Ontime/Infull/OTIF có khớp công thức MDLZ chuẩn không; (2) Top NVC/Kho/Loại hàng/Kênh/Khu vực kéo OTIF xuống có đúng kỳ vọng vận hành không; (3) Trend 5 ngày có phù hợp narrative không."),
        ("", ""),
        ("Hướng dẫn cho khách", "Khách chỉ cần nhìn cột 'SQL value', so với số nội bộ MDLZ (Excel/pipeline báo cáo). Nếu khớp → tick Status = OK. Nếu lệch ngoài tolerance → ghi rõ vào cột 'Khách ghi chú' để Smartlog điều tra."),
        ("", ""),
        ("Tolerance default", "Số đếm tuyệt đối ≤ 1% · % metric ≤ 0.5pp · Top N ≥ 4/5 tên match"),
        ("RAG bands per metric", "OTIF Green≥90 / Yellow 85–<90 / Red <85 · Ontime Green≥95 / Yellow 90–<95 / Red <90 · Infull Green≥97 / Yellow 92–<97 / Red <92"),
        ("", ""),
        ("Lưu ý window", "User yêu cầu ban đầu là 18-22/06/2026 — nhưng ETA data hôm nay (2026-05-26) chưa có đơn tương lai. Dùng proxy 18-22/05/2026 (cùng pattern thứ ngày, 1 tháng trước). Khi UAT chính thức gần ngày 06-03, sẽ re-run script với window thực."),
        ("Timezone artifact", "Dashboard filter so sánh dạng UTC nhưng customer think UTC+7. Hệ quả: Trend sheet 04 có thể hiện thị ngày 23/05 (leak ~200 đơn). Xem sheet 06 Formula Guide chi tiết — cần xác nhận với khách dashboard date cutoff."),
        ("", ""),
        ("Sheet trong file", "01 — KPI Hero (4 KPI card, có cột công thức live) · 02 — Health Matrix 5 dim (cell tính live từ sheet 05) · 03 — Fail Reason breakdown (SUMPRODUCT live) · 04 — Trend daily (COUNTIFS theo ngày) · 05 — Detail Orders (raw — 1 row = 1 SO, autofilter, là nguồn cho mọi công thức) · 06 — Formula Guide (mapping cột + pattern công thức copy paste) · 07 — UX & Filter checklist (35 mục cho khách verify visual/filter/interaction/storytelling/master data — KHÔNG compute từ data) · 08 — SQL Appendix (provenance: registry vs ad-hoc)"),
        ("Insight file riêng", "Operational insight (master data scorecard, 6 red flags, 9 KPI extension, 7 open questions) ở file md riêng: otif-uat-ops-insight-2026-05-18_to_2026-05-22.md (same folder) — KHÔNG nhồi vào Excel để giữ UAT pack tập trung vào reconciliation."),
        ("Cách verify", "Số trong sheet 01-04 đều có 2 dạng — SQL value (frozen, do Python ghi) và Excel formula (live, tính lại từ sheet 05). Hai phải khớp. Nếu lệch → có gap filter ở 1 trong 2. Xem sheet 06 Formula Guide để copy công thức custom."),
    ]
    for k, v in meta:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP
    for r_ in range(2, ws.max_row + 1):
        ws.row_dimensions[r_].height = max(18, min(80, 18 + len(str(ws.cell(row=r_, column=2).value or "")) // 60 * 16))


def _add_kpi_row(ws, label, sql_val, formula, metric_key=None, fmt=None, tolerance=""):
    ws.append([label, sql_val, formula, "", "", "", tolerance, ""])
    r_ = ws.max_row
    for col in range(1, 9):
        body_cell(ws.cell(row=r_, column=col))
    sv = ws.cell(row=r_, column=2)
    fv = ws.cell(row=r_, column=3)
    if fmt:
        sv.number_format = fmt
        fv.number_format = fmt
    if metric_key and isinstance(sql_val, (int, float)):
        sv.fill = rag_fill(metric_key, float(sql_val))
    # Dashboard / Khách columns — light yellow
    for col in (4, 5):
        ws.cell(row=r_, column=col).fill = EMPTY_CUSTOMER
    # Diff formula (Dashboard − SQL)
    df = ws.cell(row=r_, column=6)
    df.value = f"=IFERROR(D{r_}-B{r_}, \"\")"
    df.number_format = fmt or "0.00"


def build_kpi_sheet(ws, ctx):
    ws.title = "01 — KPI Hero"
    set_widths(ws, [40, 14, 60, 16, 16, 16, 18, 24])
    title_row(ws, f"L1 Hero KPI — {WINDOW_START_DATE} → {WINDOW_END_DATE} ({DATE_TYPE_LABEL})", 8)
    header_row(ws, [
        "Metric (KPI card)", "SQL value\n(frozen)", "Công thức Excel\n(live từ sheet 05)",
        "Dashboard\n(khách verify)", "Số MDLZ\n(golden file)",
        "Diff\n(Dashboard − SQL)", "Tolerance", "Khách ghi chú / Status",
    ])

    k = ctx["kpi"]
    _add_kpi_row(ws, "Tổng đơn (count DISTINCT SO)", k["total_so"],
                 f_total_orders(), fmt="#,##0", tolerance="≤ 1%")
    _add_kpi_row(ws, "Số đơn Ontime", k["ontime_so"],
                 f_count_status("O", "Ontime"), fmt="#,##0", tolerance="≤ 1%")
    _add_kpi_row(ws, "Số đơn Infull", k["infull_so"],
                 f_count_status("P", "Infull"), fmt="#,##0", tolerance="≤ 1%")
    _add_kpi_row(ws, "Số đơn OTIF", k["otif_so"],
                 f_count_status("Q", "OTIF"), fmt="#,##0", tolerance="≤ 1%")
    _add_kpi_row(ws, "% Ontime", k["pct_ontime"],
                 f_pct_status("O", "Ontime"), metric_key="ontime", fmt="0.00\"%\"", tolerance="≤ 0.5 pp")
    _add_kpi_row(ws, "% Infull", k["pct_infull"],
                 f_pct_status("P", "Infull"), metric_key="infull", fmt="0.00\"%\"", tolerance="≤ 0.5 pp")
    _add_kpi_row(ws, "% OTIF", k["pct_otif"],
                 f_pct_status("Q", "OTIF"), metric_key="otif", fmt="0.00\"%\"", tolerance="≤ 0.5 pp")

    ws.append([])
    ws.append(["Note RAG (PRD §13.2)", "OTIF target 90% · Ontime 95% (PM tentative) · Infull 97% (PM tentative)"])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)
    ws.append([])
    ws.append(["Cách verify", "Cột 'SQL value' (cố định khi chạy Python) phải = cột 'Công thức Excel' (live tính từ sheet 05 Detail Orders). Nếu lệch → có bug filter ở 1 trong 2. Xem sheet 07 Formula Guide."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).font = NOTE_FONT
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)


def build_dim_sheet(ws, ctx):
    ws.title = "02 — Health Matrix"
    set_widths(ws, [28, 30, 14, 14, 14, 14, 14, 14, 14])
    title_row(ws, "L2 Health Matrix — 5 dimensions (worst-first theo %OTIF) — số tính live từ sheet 05", 9)
    header_row(ws, ["Dimension", "Giá trị", "Tổng đơn", "Ontime", "Infull", "OTIF", "% Ontime", "% Infull", "% OTIF"])

    sections = [
        ("NVC (Nhà vận tải)", ctx["dim_nvc"], "NVC"),
        ("Kho (đã group)", ctx["dim_kho"], "Kho"),
        ("Loại hàng", ctx["dim_loai_hang"], "Loại hàng"),
        ("Kênh bán hàng", ctx["dim_kenh"], "Kênh"),
        ("Khu vực", ctx["dim_khu_vuc"], "Khu vực"),
    ]
    for label, rows, dim_key in sections:
        ws.append([label, "", "", "", "", "", "", "", ""])
        last = ws.max_row
        ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=9)
        c = ws.cell(row=last, column=1)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="14283F")
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[last].height = 22

        if not rows:
            ws.append(["", "(không có data)", "", "", "", "", "", "", ""])
            continue

        dim_col = DETAIL_DIM_COL[dim_key]
        first_val_key = list(rows[0].keys())[0]
        for row in rows:
            ws.append(["", row[first_val_key], "", "", "", "", "", "", ""])
            r_ = ws.max_row
            v_ref = f"$B{r_}"
            ws.cell(row=r_, column=3).value = f_dim_total(dim_col, v_ref)
            ws.cell(row=r_, column=4).value = f_dim_status(dim_col, v_ref, "O", "Ontime")
            ws.cell(row=r_, column=5).value = f_dim_status(dim_col, v_ref, "P", "Infull")
            ws.cell(row=r_, column=6).value = f_dim_status(dim_col, v_ref, "Q", "OTIF")
            ws.cell(row=r_, column=7).value = f_dim_pct(dim_col, v_ref, "O", "Ontime")
            ws.cell(row=r_, column=8).value = f_dim_pct(dim_col, v_ref, "P", "Infull")
            ws.cell(row=r_, column=9).value = f_dim_pct(dim_col, v_ref, "Q", "OTIF")
            for col in range(1, 10):
                body_cell(ws.cell(row=r_, column=col))
            for col in (3, 4, 5, 6):
                ws.cell(row=r_, column=col).number_format = "#,##0"
            # Apply RAG fill based on SQL-frozen pct value (since formula not evaluated until Excel opens)
            for col, mkey, src_key in ((7, "ontime", "pct_ontime"), (8, "infull", "pct_infull"), (9, "otif", "pct_otif")):
                c2 = ws.cell(row=r_, column=col)
                c2.number_format = "0.00\"%\""
                if row.get(src_key) is not None:
                    c2.fill = rag_fill(mkey, float(row[src_key]))
        ws.append([])


def build_fail_sheet(ws, ctx):
    """Fail Reason — số SQL từ registry "Phân rã nguyên nhân fail ontime/infull" (bucket theo
    not_ontime_reason / not_infull_reason raw column từ mv_otif). Excel formula live count theo
    cùng raw column từ sheet 05 (col R + S)."""
    ws.title = "03 — Fail Reason"
    set_widths(ws, [50, 14, 60, 16, 16, 16, 22, 24])
    title_row(ws, "L3 Fail Reason — bucket theo not_ontime_reason / not_infull_reason (registry §Phân rã)", 8)

    def _add_block(title, rows_data, reason_col_letter):
        ws.append([title])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=8)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="14283F")
        header_row(ws, ["Lý do (raw column từ mv_otif)", "Fail count\n(SQL frozen)",
                        "Công thức Excel\n(live)", "Dashboard\n(khách)", "Số MDLZ\n(golden)",
                        "% share", "Tolerance", "Khách ghi chú"])
        total = sum((r.get("fail_so") or 0) for r in rows_data) or 1
        for row in rows_data:
            reason = row.get("reason") or "(rỗng)"
            n_val = row.get("fail_so") or 0
            share = _pct2(100.0 * n_val / total)
            formula = (f'=COUNTIF({DETAIL_SHEET}!{reason_col_letter}:{reason_col_letter},"{reason}")'
                       if reason != "(rỗng)" else
                       f'=COUNTIF({DETAIL_SHEET}!{reason_col_letter}:{reason_col_letter},"")')
            ws.append([reason, n_val, formula, "", "", share, "≤ 1%", ""])
            r_ = ws.max_row
            for col in range(1, 9):
                body_cell(ws.cell(row=r_, column=col))
            ws.cell(row=r_, column=2).number_format = "#,##0"
            ws.cell(row=r_, column=3).number_format = "#,##0"
            ws.cell(row=r_, column=6).number_format = "0.00\"%\""
            for col in (4, 5):
                ws.cell(row=r_, column=col).fill = EMPTY_CUSTOMER
        # Total row
        ws.append([f"Tổng {title}", total, f'=SUM(B{ws.max_row - len(rows_data) + 1}:B{ws.max_row})',
                   "", "", _pct2(100), "", ""])
        r_ = ws.max_row
        for col in range(1, 9):
            body_cell(ws.cell(row=r_, column=col))
        ws.cell(row=r_, column=1).font = Font(bold=True)
        ws.cell(row=r_, column=2).number_format = "#,##0"
        ws.cell(row=r_, column=2).font = Font(bold=True)
        ws.cell(row=r_, column=6).number_format = "0.00\"%\""

    _add_block("Fail Ontime breakdown (Tier 2)", ctx["fail_ontime"], "R")
    ws.append([])
    _add_block("Fail Infull breakdown (Tier 2)", ctx["fail_infull"], "S")


def build_trend_sheet(ws, ctx):
    ws.title = "04 — Trend daily"
    set_widths(ws, [14, 14, 14, 14, 14, 14, 14, 14, 24])
    title_row(ws, "L4 Trend daily 5 ngày — số live tính từ sheet 05 (match cột ETA prefix yyyy-mm-dd)", 9)
    header_row(ws, ["Ngày", "Tổng đơn", "Ontime", "Infull", "OTIF", "% Ontime", "% Infull", "% OTIF", "Khách ghi chú"])
    for row in ctx["trend"]:
        date_str = row["d"].strftime("%Y-%m-%d")
        ws.append([row["d"], "", "", "", "", "", "", "", ""])
        r_ = ws.max_row
        ws.cell(row=r_, column=2).value = f_trend_total(date_str)
        ws.cell(row=r_, column=3).value = f_trend_count(date_str, "O", "Ontime")
        ws.cell(row=r_, column=4).value = f_trend_count(date_str, "P", "Infull")
        ws.cell(row=r_, column=5).value = f_trend_count(date_str, "Q", "OTIF")
        ws.cell(row=r_, column=6).value = f_trend_pct(date_str, "O", "Ontime")
        ws.cell(row=r_, column=7).value = f_trend_pct(date_str, "P", "Infull")
        ws.cell(row=r_, column=8).value = f_trend_pct(date_str, "Q", "OTIF")
        for col in range(1, 10):
            body_cell(ws.cell(row=r_, column=col))
        ws.cell(row=r_, column=1).number_format = "yyyy-mm-dd"
        for col in (2, 3, 4, 5):
            ws.cell(row=r_, column=col).number_format = "#,##0"
        # RAG fill from SQL-frozen pct
        for col, mkey, src_key in ((6, "ontime", "pct_ontime"), (7, "infull", "pct_infull"), (8, "otif", "pct_otif")):
            c2 = ws.cell(row=r_, column=col)
            c2.number_format = "0.00\"%\""
            if row.get(src_key) is not None:
                c2.fill = rag_fill(mkey, float(row[src_key]))


def build_detail_sheet(ws, ctx):
    ws.title = "05 — Detail Orders"
    headers = [
        "SO", "Kho code", "Kho group",
        "Loại hàng", "Kênh", "NVC", "Khu vực",
        "Customer code", "Customer name",
        "ETA (UTC+7)", "ATA (UTC+7)",
        "Planned CSE", "Shipped CSE", "Delivered CSE",
        "Ontime status", "Infull status", "OTIF status",
        "Lý do not Ontime", "Lý do not Infull",
        "Khớp golden? (Y/N)", "Khách ghi chú",
    ]
    widths = [
        18, 14, 18,
        16, 16, 18, 18,
        14, 28,
        20, 20,
        14, 14, 14,
        18, 18, 16,
        28, 28,
        18, 30,
    ]
    set_widths(ws, widths)
    title_row(ws, f"L6 Detail Orders — {len(ctx['detail']):,} đơn (1 row = 1 SO) cho khách đối chiếu line-by-line", len(headers))
    header_row(ws, headers)

    status_fill = {
        "Ontime": GREEN, "Infull": GREEN, "OTIF": GREEN,
        "Failed Ontime": RED, "Failed Infull": RED, "Failed OTIF": RED,
    }

    for row in ctx["detail"]:
        ws.append([
            row["so"], row["kho_code"], row["kho_group"],
            row["loai_hang"], row["kenh"], row["nvc"], row["khu_vuc"],
            row["customer_code"], row["customer_name"],
            row["eta_utc7"], row["ata_utc7"],
            row["planned_cse"], row["shipped_cse"], row["delivered_cse"],
            row["ontime_status"], row["infull_status"], row["otif_status"],
            row["not_ontime_reason"], row["not_infull_reason"],
            "", "",
        ])
        r_ = ws.max_row
        for col in range(1, len(headers) + 1):
            body_cell(ws.cell(row=r_, column=col))
        # ETA + ATA — proper datetime format (datetime serial number, not text)
        for col in (10, 11):
            ws.cell(row=r_, column=col).number_format = "yyyy-mm-dd hh:mm:ss"
        # CSE cols — number with 2 decimals
        for col in (12, 13, 14):
            ws.cell(row=r_, column=col).number_format = "#,##0.00"
        # Color status cols
        for col in (15, 16, 17):
            c2 = ws.cell(row=r_, column=col)
            fill = status_fill.get(c2.value)
            if fill:
                c2.fill = fill
        # Empty customer-fill cols (Khớp + Ghi chú)
        for col in (20, 21):
            ws.cell(row=r_, column=col).fill = EMPTY_CUSTOMER

    # Freeze pane on header + autofilter
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"


def build_formula_guide_sheet(ws, ctx):
    ws.title = "06 — Formula Guide"
    set_widths(ws, [28, 90])
    title_row(ws, "Excel Formula Guide — copy paste khi PM/Khách hàng cần verify thủ công", 2)

    intro = [
        ("Mục đích", "Tất cả số trong sheet 01-04 đều có công thức live tính lại từ sheet 05 (Detail Orders). PM/khách có thể copy công thức từ đây sang Excel nội bộ hoặc dùng để audit số dashboard."),
        ("Lưu ý timezone (QUAN TRỌNG)", "ETA + ATA ở sheet 05 đã convert UTC+7. WHERE filter của SQL/dashboard so sánh dạng UTC (vì cột stored DateTime64 UTC). → Window 18-22/05 UTC trên dashboard ≠ window 18-22/05 UTC+7 mà khách kỳ vọng. Hệ quả: sheet 04 Trend daily có thể xuất hiện **ngày 23/05** (leak) — đó là đơn có ETA UTC <= 22/05 23:59 UTC nhưng UTC+7 thực tế = sáng 23/05. Đây là **finding cần xác nhận với khách**: dashboard date filter cutoff theo UTC hay UTC+7? Phân biệt rõ trước khi accept reconciliation."),
        ("ETA/ATA dtype = datetime (KHÔNG phải text)", "Cột J + K ở sheet 05 là **datetime cell** (serial number, format yyyy-mm-dd hh:mm:ss). Có thể sort theo thời gian, filter theo range, dùng INT()/DATE()/HOUR() trong công thức. Trend sheet 04 dùng COUNTIFS với DATE() bound thay vì prefix-text match — robust với datetime."),
        ("Sheet name reference", "Tên sheet detail trong công thức: '05 — Detail Orders' (có space + em-dash, BẮT BUỘC giữ nguyên hoặc đổi đồng bộ tất cả công thức)."),
    ]
    for k, v in intro:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP

    ws.append([])
    ws.append(["Bảng 1 — Mapping cột Detail Orders", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="14283F")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    col_map = [
        ("A", "SO (mã đơn)"),
        ("B", "Kho code (whseid raw)"),
        ("C", "Kho group (đã group BKD/NKD/...)"),
        ("D", "Loại hàng (group_of_cago)"),
        ("E", "Kênh bán hàng (group_name)"),
        ("F", "NVC (ten_ngan_nha_van_tai)"),
        ("G", "Khu vực (khu_vuc_doi_xe)"),
        ("H", "Customer code"),
        ("I", "Customer name"),
        ("J", "ETA UTC+7 (text: yyyy-mm-dd hh:mm:ss)"),
        ("K", "ATA UTC+7 (text)"),
        ("L", "Planned CSE"),
        ("M", "Shipped CSE"),
        ("N", "Delivered CSE"),
        ("O", "Ontime status (Ontime / Failed Ontime)"),
        ("P", "Infull status (Infull / Failed Infull)"),
        ("Q", "OTIF status (OTIF / Failed OTIF) — đã loại STM"),
        ("R", "Lý do not Ontime"),
        ("S", "Lý do not Infull"),
        ("T", "Khớp golden? (Y/N) — khách điền"),
        ("U", "Khách ghi chú — khách điền"),
    ]
    for col, desc in col_map:
        ws.append([col, desc])
        for c in range(1, 3):
            body_cell(ws.cell(row=ws.max_row, column=c))

    ws.append([])
    ws.append(["Bảng 2 — Pattern công thức chuẩn", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="14283F")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)

    n_rows = len(ctx["detail"])
    last = n_rows + 2

    patterns = [
        ("Tổng đơn (đã loại STM)",
         f'=COUNTIF({DETAIL_SHEET}!Q:Q,"OTIF")+COUNTIF({DETAIL_SHEET}!Q:Q,"Failed OTIF")'),
        ("Số đơn Ontime", f'=COUNTIF({DETAIL_SHEET}!O:O,"Ontime")'),
        ("Số đơn Infull", f'=COUNTIF({DETAIL_SHEET}!P:P,"Infull")'),
        ("Số đơn OTIF", f'=COUNTIF({DETAIL_SHEET}!Q:Q,"OTIF")'),
        ("% Ontime",
         f'=ROUND(COUNTIF({DETAIL_SHEET}!O:O,"Ontime")/(COUNTIF({DETAIL_SHEET}!Q:Q,"OTIF")+COUNTIF({DETAIL_SHEET}!Q:Q,"Failed OTIF"))*100,2)'),
        ("% OTIF theo 1 NVC cụ thể (vd 'HDA' ở $B5)",
         f'=IFERROR(ROUND(COUNTIFS({DETAIL_SHEET}!F:F,$B5,{DETAIL_SHEET}!Q:Q,"OTIF")/COUNTIF({DETAIL_SHEET}!F:F,$B5)*100,2),0)'),
        ("% OTIF theo Kho group ($B5)",
         f'=IFERROR(ROUND(COUNTIFS({DETAIL_SHEET}!C:C,$B5,{DETAIL_SHEET}!Q:Q,"OTIF")/COUNTIF({DETAIL_SHEET}!C:C,$B5)*100,2),0)'),
        ("% OTIF theo Loại hàng ($B5)",
         f'=IFERROR(ROUND(COUNTIFS({DETAIL_SHEET}!D:D,$B5,{DETAIL_SHEET}!Q:Q,"OTIF")/COUNTIF({DETAIL_SHEET}!D:D,$B5)*100,2),0)'),
        ("Số đơn theo ngày ETA (datetime range, không phải text prefix)",
         f'=COUNTIFS({DETAIL_SHEET}!J:J,">="&DATE(2026,5,18),{DETAIL_SHEET}!J:J,"<"&DATE(2026,5,19))'),
        ("% OTIF theo ngày ETA (datetime range)",
         f'=IFERROR(ROUND(COUNTIFS({DETAIL_SHEET}!J:J,">="&DATE(2026,5,18),{DETAIL_SHEET}!J:J,"<"&DATE(2026,5,19),{DETAIL_SHEET}!Q:Q,"OTIF")/COUNTIFS({DETAIL_SHEET}!J:J,">="&DATE(2026,5,18),{DETAIL_SHEET}!J:J,"<"&DATE(2026,5,19))*100,2),0)'),
        ("Fail Ontime — Lỗi transport giao trễ (planned = shipped)",
         f'=SUMPRODUCT(({DETAIL_SHEET}!O3:O{last}="Failed Ontime")*({DETAIL_SHEET}!L3:L{last}={DETAIL_SHEET}!M3:M{last}))'),
        ("Fail Ontime — Lỗi WH gọi xe trễ (shipped < planned)",
         f'=SUMPRODUCT(({DETAIL_SHEET}!O3:O{last}="Failed Ontime")*({DETAIL_SHEET}!M3:M{last}<{DETAIL_SHEET}!L3:L{last}))'),
        ("Fail Infull — Warehouse failure (shipped < planned, delivered ≥ shipped)",
         f'=SUMPRODUCT(({DETAIL_SHEET}!P3:P{last}="Failed Infull")*({DETAIL_SHEET}!M3:M{last}<{DETAIL_SHEET}!L3:L{last})*({DETAIL_SHEET}!N3:N{last}>={DETAIL_SHEET}!M3:M{last}))'),
        ("Fail Infull — Transport failure (shipped = planned, delivered < shipped)",
         f'=SUMPRODUCT(({DETAIL_SHEET}!P3:P{last}="Failed Infull")*({DETAIL_SHEET}!M3:M{last}={DETAIL_SHEET}!L3:L{last})*({DETAIL_SHEET}!N3:N{last}<{DETAIL_SHEET}!M3:M{last}))'),
        ("Lateness median (phút) — đơn Failed Ontime",
         "Excel KHÔNG có hàm dateDiff sẵn cho text format. Tính tay từ sheet 05: thêm cột phụ V=K-J×24×60 (nếu K/J là datetime serial). Hoặc dùng SQL Q-INS-7 trong file ops insight."),
        ("Lateness >12h count",
         "Tương tự — cần datetime serial. Workaround: chạy lại script Python uat_otif_export.py để cập nhật + reference cột K/J convert datetime."),
    ]
    for label, formula in patterns:
        ws.append([label, formula])
        r_ = ws.max_row
        body_cell(ws.cell(row=r_, column=1))
        c = ws.cell(row=r_, column=2)
        c.font = Font(name="Consolas", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        ws.row_dimensions[r_].height = max(20, min(80, 16 + len(formula) // 80 * 14))

    ws.append([])
    ws.append(["Cảnh báo", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor="DC2626")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    warnings = [
        ("Range fragile", f"SUMPRODUCT dùng range L3:L{last} cố định theo số dòng hiện tại (n={n_rows}). Nếu sheet 05 bị xoá/thêm dòng → công thức Fail Reason có thể lệch. KPI Hero + Health Matrix + Trend dùng full-column COUNTIF — an toàn hơn."),
        ("STM exclusion", f"Sheet 05 chỉ chứa đơn đã loại 'Không có dữ liệu STM'. Nếu cần test bao gồm STM → re-run script Python với EXCLUDE_CLAUSE rỗng."),
        ("Sheet rename", f"Đổi tên sheet 05 → công thức ở 01-04 sẽ broken. Nếu phải rename, dùng Find&Replace trên formula bar: '05 — Detail Orders' → tên mới."),
        ("Recalculation", f"Mở file lần đầu Excel hỏi 'Update external links' = KHÔNG (vì không có external link). Nếu công thức không refresh → F9 hoặc Ctrl+Alt+F9 (full recalc)."),
    ]
    for k, v in warnings:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = WRAP
        for c in range(1, 3):
            body_cell(ws.cell(row=ws.max_row, column=c))


UX_CHECKLIST_ITEMS = [
    # (category, what_to_check, expected, severity_if_fail, ref)
    # === UX Visual ===
    ("UX Visual", "KPI cards (Tổng đơn / % Ontime / % Infull / % OTIF) — value đủ to readable trên viewport 1366×768",
     "Headline ≥ display weight (lớn hơn body-large). Label dim ≥ 12px. Không dùng 10px subtitle.", "Minor", "PRD §13.6"),
    ("UX Visual", "RAG color border/background của 3 KPI có target (% Ontime/Infull/OTIF) visible trên HiDPI / retina",
     "RAG color rõ ràng trên mọi resolution, không dùng border 0.5px (vô hình trên retina).", "Minor", "PRD §13.6"),
    ("UX Visual", "KPI cards hiển thị 4 phần tử per AC-10: giá trị + gap-to-target + RAG indicator + 2 delta vs tuần trước / tháng trước",
     "Đủ 4 phần tử. KPI 'Tổng đơn' không có target — chỉ value + 2 delta + subtitle context.", "Major", "PRD AC-10"),
    ("UX Visual", "Health Matrix cell coloring theo per-metric band §13.2 (OTIF: G≥90/Y85–<90/R<85; Ontime: G≥95/Y90–<95/R<90; Infull: G≥97/Y92–<97/R<92)",
     "3 metric với 3 band khác nhau, không dùng cùng 1 threshold cho cả 3.", "Major", "PRD §13.2 + §13.5"),
    ("UX Visual", "Trend chart có vùng nền xanh nhạt (target band) từ target % OTIF (90%) tới 100% + reference dashed line tại 90",
     "User spot được 'đạt/dưới target' không cần đọc số.", "Major", "PRD AC-11 + §13.7"),
    ("UX Visual", "Action title của mỗi chart nói insight (vd 'NVC TLL kéo OTIF chung xuống 4pt') thay vì chỉ tên dim ('OTIF theo NVC')",
     "Title chứa: giá trị xấu nhất + magnitude + gap to target khi áp dụng.", "Major", "PRD AC-12"),
    ("UX Visual", "Tier 1 (4 KPI + Health Matrix + Mini sparkline) fit 1 fold trên viewport 1366×768 — user trả lời Q1+Q2 không scroll",
     "Đủ 3 component Tier 1 visible trên 1 fold đầu, không bị clip.", "Major", "PRD AC-14"),
    ("UX Visual", "Tier 3 (5 dim drill-down NVC/Kho/Loại hàng/Kênh/Khu vực) **expanded mặc định** (post-v1.2.6 reversal)",
     "5 chart Tier 3 render sẵn khi load, header có toggle để collapse manual. KHÔNG cần click để expand.", "Major", "PRD AC-15"),
    ("UX Visual", "KPI bar sticky top khi scroll xuống Tier 2/3 (sau filter bar)",
     "KPI context giữ visible suốt tab Chart, user luôn thấy '% chính' khi đang nhìn detail.", "Minor", "PRD §13.6"),
    # === Filter behavior ===
    ("Filter", "Filter Kho dropdown đủ 4 group + ALL (BKD / NKD / Kho ngoài BKD / Kho ngoài NKD)",
     "4 group hiển thị đúng tên, ánh xạ whseid đúng PRD §4 (BKD = BKD1+BKD2+BKD3 etc).", "Critical", "PRD §4 + AC-09"),
    ("Filter", "Filter NVC dropdown hiển thị tên ngắn VN đúng (HDA, ANH SON, HOA PHAT, TLL, NJV-Nhất Tín, ...) — không phải code raw",
     "Tên VN có dấu (Nhất Tín, không phải 'Nhat Tin'). Loại trừ '(rỗng)' nếu master data có gap.", "Minor", "PRD §4"),
    ("Filter", "Filter Loại hàng dropdown đúng 7 option theo OTIF_CATEGORY_ORDER (FRESH/DRY/MOONCAKE/POSM-OFFBOM/PM/EQUIPMENT/TEST) + ALL",
     "Đủ 7 type. Order trong dropdown theo priority hoặc alpha, consistent với chart by category.", "Minor", "PRD §5.2"),
    ("Filter", "Filter Loại ngày — chỉ 2 option (ETA gửi thầu / ATA chi tiết chuyến). Mặc định = ETA gửi thầu",
     "2 option duy nhất; OQ-05 SQL 7-branch là tech-debt không expose user.", "Minor", "PRD §3"),
    ("Filter", "Filter Khoảng ngày — chọn > 2 năm → toast lỗi + KHÔNG gọi API",
     "isDateStringRangeOver2Years() trigger, MAX_DATE_RANGE_MS check. Filter trước vẫn được giữ, không submit request.", "Major", "PRD AC-02"),
    ("Filter", "Filter mặc định khi mở dashboard lần đầu — Khoảng ngày = hôm nay; còn lại = ALL",
     "Default state chuẩn không phải state cũ từ session khác.", "Minor", "PRD §4"),
    ("Filter", "Filter persist sau F5 refresh (localStorage `dashboard-widget-filter:{dashboardId}:{widgetId}`)",
     "Quay lại trang giữ filter state. Reset chỉ khi user click Reset.", "Minor", "PRD §12"),
    ("Filter", "Multi-select 2+ value cho 1 dim (vd NVC=ANH SON + HOA PHAT) → query AND/IN logic đúng",
     "Số dashboard = SQL `whseid IN ('A','B')` không phải `=`. Verify với Health Matrix sum.", "Major", "AC-09"),
    ("Filter", "Filter combo 5-dim (NPP + Kho + Khu vực + Loại hàng + NVC) response time < 3s",
     "Trace Network tab; nếu > 3s flag perf defect.", "Minor", "Plan §8"),
    ("Filter", "Filter Reset / Clear all button — restore default state đầy đủ",
     "Click Reset → filter về default 100%, không sót dim nào.", "Minor", "UX standard"),
    # === Interaction ===
    ("Interaction", "Click row Health Matrix → smooth scroll xuống Tier 3 chart tương ứng + ensure auto-expand (nếu đang collapsed)",
     "Click NVC=TLL → scroll smooth tới chart 'By Transporter' + chart đã open. KHÔNG cross-filter các chart khác.", "Major", "PRD AC-15 + §13.5"),
    ("Interaction", "Click 1 DO row trong Detail Table → mở Order Monitor với context DO + load < 2s",
     "URL/state có DO code; load wall-clock < 2s.", "Minor", "AC-16 (proposed)"),
    ("Interaction", "Nút Export trên mỗi chart — chọn PNG hoặc CSV; file download có đúng tên + data",
     "Export PNG: file size > 0, mở được. Export CSV: header + rows match chart data.", "Minor", "PRD AC-08"),
    ("Interaction", "Trend chart toggle Time Bucket Day / Week / Month — đổi grouping ngay, KHÔNG gọi lại API",
     "Toggle chuyển local data already-fetched; không có request Network mới khi click.", "Minor", "PRD AC-04"),
    ("Interaction", "Hover trên bar chart → tooltip hiển thị % + count + dim value rõ ràng",
     "Tooltip không bị clipped ngoài viewport; có cả 2 chỉ số % và count tuyệt đối.", "Cosmetic", "UX standard"),
    ("Interaction", "Operation Summary tab — chọn / bỏ chọn 4 dim (NVC/Kênh/Loại hàng/Khu vực) → bảng pivot recompute đúng",
     "Bảng pivot group lại theo dim đã chọn, % OTIF aggregate lại đúng tổng cấp trên.", "Major", "PRD AC-05"),
    ("Interaction", "Detail Table — sort theo bất kỳ column nào (ETA, ATA, NVC, Status, ...)",
     "Click header → asc; click lần 2 → desc; sort indicator visible.", "Minor", "UX standard"),
    # === Storytelling (mental model) ===
    ("Storytelling", "Q1 — 'OTIF tổng có đỏ không?' trả lời được ≤ 2 giây nhìn KPI Hero",
     "% OTIF + RAG color visible ngay, không cần scroll/decode.", "Major", "PRD §13.9"),
    ("Storytelling", "Q2 — 'Vấn đề ở chiều nào?' trả lời được ≤ 10 giây từ Health Matrix",
     "Customer chỉ đúng row dim worst-performing.", "Major", "PRD §13.9"),
    ("Storytelling", "Q3 — 'Lý do fail là gì?' trả lời được ≤ 20 giây từ Fail Reason chart",
     "Customer phân biệt được transport/warehouse root cause.", "Minor", "PRD §13.9"),
    ("Storytelling", "Q4 — 'Xu hướng có xấu đi không?' trả lời được ≤ 30 giây từ Trend chart",
     "Customer thấy được pattern lên/xuống trong window.", "Minor", "PRD §13.9"),
    ("Storytelling", "Q5 — 'Cần action gì? Ai chịu trách nhiệm?' — từ Tier 3 + Exception Spotlight",
     "Customer match được vendor/kho/khu vực cần intervene.", "Minor", "PRD §13.9"),
    ("Storytelling", "Exception Spotlight liệt kê top 3-5 items off-target xuyên 5 dim, sort theo magnitude (gap × volume)",
     "Mỗi item: tên dim + value + % OTIF + gap signed + drill-down link.", "Major", "PRD AC-13 + §13.8"),
    # === Master Data display ===
    ("Master Data", "Tên NVC trong filter dropdown + chart label — đầy đủ ngữ nghĩa, không phải mã raw",
     "Hiển thị 'Nhất Tín' không phải 'NJV1'; 'Hỏa Phát' tiếng Việt có dấu nếu master data chứa dấu.", "Minor", "Master data hygiene"),
    ("Master Data", "Customer name hiển thị trong Detail Table + Exception Spotlight — không truncate, không thiếu space",
     "Verify mẫu 5 customer: 'CN BINH DUONG - CONG TY CP DV THUONG MAI...' đầy đủ space (gap 'DV'/'THUONG'), không bị 'DVTHUONG'.", "Minor", "Master data — Q-INS-9"),
    ("Master Data", "Đơn có NVC empty/null trong DB hiển thị thế nào trong dashboard",
     "Hiển thị '(rỗng)' hoặc 'Chưa phân loại' với mã rõ; KHÔNG để empty string làm trắng row.", "Minor", "Q-INS-4"),
    ("Master Data", "Number format — số đếm có dấu phẩy ngăn ngàn (vd 2,986 không phải 2986)",
     "Convention VN/EN — chốt với khách. Hiện UI dùng 2,986 (en-US).", "Cosmetic", "i18n"),
    ("Master Data", "% format — '82.94%' với 2 chữ số thập phân, không phải '82.9' hoặc '82.94000%'",
     "Round 2 decimal places, có ký hiệu %.", "Cosmetic", "i18n"),
]


def build_ux_checklist_sheet(ws, ctx):
    """UX/Filter/Interaction/Storytelling/Master Data checklist — customer fill trong session."""
    ws.title = "07 — UX & Filter checklist"
    set_widths(ws, [6, 20, 48, 48, 14, 12, 16, 30])
    title_row(ws, "UX / Filter / Interaction / Storytelling / Master Data checklist — KHÔNG compute từ data, customer fill trong session UAT", 8)
    header_row(ws, [
        "#", "Category",
        "Mục cần kiểm tra (Quan sát thế nào?)",
        "Expected behavior",
        "Customer\nverdict",
        "Severity\nnếu Fail",
        "Reference\n(PRD/AC/source)",
        "Customer ghi chú",
    ])

    category_fill = {
        "UX Visual":      PatternFill("solid", fgColor="DBEAFE"),  # light blue
        "Filter":         PatternFill("solid", fgColor="FEF3C7"),  # light yellow
        "Interaction":    PatternFill("solid", fgColor="E0E7FF"),  # light indigo
        "Storytelling":   PatternFill("solid", fgColor="FCE7F3"),  # light pink
        "Master Data":    PatternFill("solid", fgColor="DCFCE7"),  # light green
    }
    severity_fill = {"Critical": RED, "Major": YELLOW, "Minor": GREEN, "Cosmetic": GREY}

    for i, (cat, check, expected, severity, ref) in enumerate(UX_CHECKLIST_ITEMS, 1):
        ws.append([i, cat, check, expected, "", severity, ref, ""])
        r_ = ws.max_row
        for col in range(1, 9):
            body_cell(ws.cell(row=r_, column=col))
            ws.cell(row=r_, column=col).alignment = WRAP
        # Category fill
        ws.cell(row=r_, column=2).fill = category_fill.get(cat, GREY)
        ws.cell(row=r_, column=2).font = Font(bold=True)
        # Verdict cell — empty yellow for customer to fill
        ws.cell(row=r_, column=5).fill = EMPTY_CUSTOMER
        ws.cell(row=r_, column=5).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Severity color
        ws.cell(row=r_, column=6).fill = severity_fill.get(severity, GREY)
        ws.cell(row=r_, column=6).alignment = Alignment(horizontal="center", vertical="center")
        # Notes cell yellow
        ws.cell(row=r_, column=8).fill = EMPTY_CUSTOMER
        # Row height
        max_text = max(len(check), len(expected))
        ws.row_dimensions[r_].height = max(36, min(96, 20 + max_text // 50 * 16))

    # Freeze pane + summary footer
    ws.freeze_panes = "A3"

    ws.append([])
    ws.append(["Hướng dẫn customer", "Mỗi row = 1 mục cần verify trên dashboard. Customer điền cột 'Verdict' (Pass / Fail / N/A / cần làm rõ) + 'Customer ghi chú' nếu có. Severity cho biết mức độ nếu Fail. Reference trỏ về acceptance criteria gốc trong PRD."])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)
    ws.append(["Tham chiếu skill", "Skill /da-uat đã chốt 4 lớp test (A. Data Reconciliation / B. Business Logic / C. UX & Storytelling / D. Performance). Sheet này tập trung lớp C + D + filter behavior (cross-cutting), bổ sung cho sheet 01-05 chỉ kiểm Data."])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)


def build_sql_sheet(ws, sql_map):
    ws.title = "08 — SQL Appendix"
    set_widths(ws, [40, 32, 110])
    title_row(ws, "SQL queries — provenance rõ ràng (canonical từ sql-registry.md OR ad-hoc)", 3)
    header_row(ws, ["Query ID — mục đích", "Source / provenance", "SQL (đã substitute placeholder)"])
    for qid, (q, src) in sql_map.items():
        ws.append([qid, src, q])
        r_ = ws.max_row
        ws.cell(row=r_, column=1).font = Font(bold=True)
        ws.cell(row=r_, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        src_cell = ws.cell(row=r_, column=2)
        src_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if src.startswith("sql-registry.md"):
            src_cell.fill = GREEN
            src_cell.font = Font(bold=True)
        else:
            src_cell.fill = YELLOW
        for col in (1, 2, 3):
            ws.cell(row=r_, column=col).border = BORDER
        sql_cell = ws.cell(row=r_, column=3)
        sql_cell.font = Font(name="Consolas", size=9)
        sql_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_].height = max(40, min(420, 18 + q.count("\n") * 13))


# ── Main ─────────────────────────────────────────────────
def main():
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"[*] Connect CH…")
    client = get_client()
    print(f"[*] Run queries window {WINDOW_START_DATE}→{WINDOW_END_DATE}…")

    ctx = {
        "run_ts": datetime.now(timezone(timedelta(hours=7))).strftime("%Y-%m-%d %H:%M UTC+7"),
        "server_now": run_mv_refresh(client),
    }

    print(f"[*] Query mv_filter_* for ALL substitution…")
    filters = query_filter_values(client)
    ctx["filters"] = filters

    kpi, sql_kpi, src_kpi = run_kpi(client, filters); ctx["kpi"] = kpi; print(f"    KPI: total={kpi['total_so']:,} pct_otif={kpi['pct_otif']}")
    fo, sql_fo, src_fo = run_fail_ontime(client, filters); ctx["fail_ontime"] = fo
    fi, sql_fi, src_fi = run_fail_infull(client, filters); ctx["fail_infull"] = fi
    nvc, sql_nvc, src_nvc = run_by_registry_dim(client, "nvc", filters); ctx["dim_nvc"] = nvc
    kho, sql_kho, src_kho = run_by_registry_dim(client, "kho", filters); ctx["dim_kho"] = kho
    lh, sql_lh, src_lh = run_by_loai_hang(client, filters); ctx["dim_loai_hang"] = lh
    kenh, sql_kenh, src_kenh = run_by_registry_dim(client, "kenh", filters); ctx["dim_kenh"] = kenh
    kv, sql_kv, src_kv = run_by_registry_dim(client, "khu_vuc", filters); ctx["dim_khu_vuc"] = kv
    trend, sql_trend, src_trend = run_trend(client, filters); ctx["trend"] = trend
    detail, sql_detail, src_detail = run_detail(client, filters); ctx["detail"] = detail; print(f"    Detail: {len(detail):,} rows")

    sql_map = {
        "Q-01 — KPI hero (registry §Tổng đơn)": (sql_kpi, src_kpi),
        "Q-02 — Fail Ontime (registry §Phân rã nguyên nhân fail ontime)": (sql_fo, src_fo),
        "Q-03 — Fail Infull (registry §Phân rã nguyên nhân fail infull)": (sql_fi, src_fi),
        "Q-04 — Theo NVC (registry §theo nhà vận tải)": (sql_nvc, src_nvc),
        "Q-05 — Theo Kho (registry §theo kho)": (sql_kho, src_kho),
        "Q-06 — Theo Loại hàng (ad-hoc)": (sql_lh, src_lh),
        "Q-07 — Theo Kênh (registry §theo kênh bán hàng)": (sql_kenh, src_kenh),
        "Q-08 — Theo Khu vực (registry §theo khu vực)": (sql_kv, src_kv),
        "Q-09 — Trend daily (registry §%OTIF và khối lượng đơn theo thời gian)": (sql_trend, src_trend),
        "Q-10 — Detail orders (ad-hoc)": (sql_detail, src_detail),
    }

    print(f"[*] Build workbook…")
    wb = Workbook()
    build_readme(wb.active, ctx)
    build_kpi_sheet(wb.create_sheet(), ctx)
    build_dim_sheet(wb.create_sheet(), ctx)
    build_fail_sheet(wb.create_sheet(), ctx)
    build_trend_sheet(wb.create_sheet(), ctx)
    build_detail_sheet(wb.create_sheet(), ctx)
    build_formula_guide_sheet(wb.create_sheet(), ctx)
    build_ux_checklist_sheet(wb.create_sheet(), ctx)
    build_sql_sheet(wb.create_sheet(), sql_map)

    wb.save(OUT_FILE)
    print(f"[OK] Wrote {OUT_FILE}")


if __name__ == "__main__":
    main()
