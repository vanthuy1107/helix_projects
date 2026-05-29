"""UAT Quick Check — Mondelez

Sinh 1 file Excel ĐƠN GIẢN cho khách đối chiếu nhanh giữa dashboard Control Tower
và số liệu khách đang theo dõi (file Excel khách tự làm / hệ thống cũ).

Cấu trúc:
  - README: hướng dẫn + time window + danh sách sheet
  - Tổng hợp: 1 row/section, KPI chính + tổng đơn
  - 15 sheet/section: 4 nhóm metric (Tổng đơn / Số lượng / Trạng thái / KPI)
    - Cột: # | Nhóm | Chỉ tiêu | Dashboard | WMS/TMS | Diff | Status | Ghi chú
    - Số liệu để TRỐNG — PM/khách điền Dashboard + số gốc từ WMS (kho) / TMS (vận tải)

Khác với uat_otif_export.py / uat_flash_daily_export.py / uat_vfr_export.py:
  - KHÔNG query DB, KHÔNG load SQL registry
  - KHÔNG có formula live — chỉ cell trống
  - Mục tiêu: đối chiếu nhanh để xác nhận số khớp/lệch, KHÔNG audit sâu

Re-run khi cần đổi time window — chỉnh WINDOW_FROM/WINDOW_TO rồi chạy lại.
"""
from __future__ import annotations
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

WINDOW_FROM = date(2026, 5, 22)
WINDOW_TO = date(2026, 5, 28)
TENANT = "Mondelez"
OUT_DIR = Path(__file__).resolve().parent.parent / "uat"
OUT_FILE = OUT_DIR / f"mondelez-uat-quick-check-{WINDOW_FROM:%Y-%m-%d}_to_{WINDOW_TO:%Y-%m-%d}.xlsx"

NAVY = "1E3A5F"
DARK = "14283F"
ACCENT = "2563EB"
PALE = "EFF4FB"
WHITE = "FFFFFF"
YELLOW = "FFF4CC"
GROUP_TONG_DON = "DBEAFE"
GROUP_SO_LUONG = "DCFCE7"
GROUP_TRANG_THAI = "FEF3C7"
GROUP_KPI = "FCE7F3"

GROUP_COLOR = {
    "Tổng đơn": GROUP_TONG_DON,
    "Số lượng": GROUP_SO_LUONG,
    "Trạng thái": GROUP_TRANG_THAI,
    "KPI": GROUP_KPI,
}

SOURCE_SYSTEM = {
    "OTIF": "TMS (vận tải)",
    "Flash Daily": "TMS (vận tải)",
    "VFR": "TMS (vận tải)",
    "Shipping Progress": "TMS (vận tải)",
    "Factory Inbound": "WMS (kho)",
    "Loose Picking": "WMS (kho)",
    "Stock Type": "WMS (kho)",
    "Transfer": "WMS (kho)",
    "Txn Move": "WMS (kho)",
    "WH Utilization": "WMS (kho)",
}

THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def header_font(color: str = WHITE, bold: bool = True, size: int = 11) -> Font:
    return Font(name="Calibri", color=color, bold=bold, size=size)


def body_font(bold: bool = False) -> Font:
    return Font(name="Calibri", color=DARK, bold=bold, size=11)


def fill(hex_: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_)


# Schema mỗi section: (sheet_name, view_title, mental_question, rows)
# rows = list of (group, metric_name, unit)
SECTIONS: list[tuple[str, str, str, list[tuple[str, str, str]]]] = [
    (
        "OTIF",
        "Section OTIF — On-Time In-Full",
        "Đơn hàng giao đúng hạn & đủ số lượng đạt bao nhiêu % so với target 90%?",
        [
            ("Tổng đơn", "Tổng đơn trong kỳ", "đơn"),
            ("Tổng đơn", "Tổng đơn có dữ liệu STM (đủ điều kiện tính OTIF)", "đơn"),
            ("Số lượng", "Tổng CSE giao (Case-Equivalent)", "CSE"),
            ("Trạng thái", "Đơn đạt OTIF", "đơn"),
            ("Trạng thái", "Đơn Fail Ontime", "đơn"),
            ("Trạng thái", "Đơn Fail Infull", "đơn"),
            ("KPI", "% OTIF (target 90%)", "%"),
            ("KPI", "% Ontime (target 95%)", "%"),
            ("KPI", "% Infull (target 97%)", "%"),
        ],
    ),
    (
        "Flash Daily",
        "Section Flash Daily — Báo cáo giao hàng theo ngày",
        "Hôm nay khách đặt bao nhiêu, đã giao xong bao nhiêu, còn lại bao nhiêu chưa giao?",
        [
            ("Tổng đơn", "Tổng đơn flash trong kỳ", "đơn"),
            ("Tổng đơn", "Tổng customer order", "customer"),
            ("Số lượng", "Tổng CSE đặt", "CSE"),
            ("Số lượng", "Tổng CSE đã giao", "CSE"),
            ("Trạng thái", "Đơn đã giao hoàn tất", "đơn"),
            ("Trạng thái", "Đơn đang giao / pending", "đơn"),
            ("Trạng thái", "Đơn huỷ / không giao được", "đơn"),
            ("KPI", "% hoàn thành theo đơn", "%"),
            ("KPI", "% hoàn thành theo CSE", "%"),
        ],
    ),
    (
        "VFR",
        "Section VFR — Vehicle Fill Rate",
        "Xe có chở đủ tải không? Bao nhiêu chuyến đạt VFR ≥ target?",
        [
            ("Tổng đơn", "Tổng số chuyến trong kỳ", "chuyến"),
            ("Tổng đơn", "Số chuyến có dữ liệu thực tế", "chuyến"),
            ("Số lượng", "Tổng tải trọng kế hoạch", "kg"),
            ("Số lượng", "Tổng tải trọng thực tế", "kg"),
            ("Trạng thái", "Chuyến đạt VFR", "chuyến"),
            ("Trạng thái", "Chuyến không đạt VFR", "chuyến"),
            ("KPI", "% VFR trung bình", "%"),
            ("KPI", "% chuyến đạt VFR", "%"),
        ],
    ),
    (
        "Factory Inbound",
        "Section Factory Inbound — Nhập kho từ nhà máy",
        "Hàng từ nhà máy về kho — đã nhận bao nhiêu, còn pending bao nhiêu?",
        [
            ("Tổng đơn", "Tổng PO inbound trong kỳ", "PO"),
            ("Tổng đơn", "Tổng ASN", "ASN"),
            ("Số lượng", "Tổng pallet kế hoạch nhận", "pallet"),
            ("Số lượng", "Tổng pallet thực tế nhận", "pallet"),
            ("Trạng thái", "PO đã nhận đủ", "PO"),
            ("Trạng thái", "PO đang nhận một phần", "PO"),
            ("Trạng thái", "PO pending / chưa nhận", "PO"),
            ("KPI", "% PO hoàn tất", "%"),
            ("KPI", "% pallet đạt kế hoạch", "%"),
        ],
    ),
    (
        "Loose Picking",
        "Section Loose Picking — Soạn hàng lẻ",
        "Tổng pick lẻ trong kỳ, độ chính xác bao nhiêu %, có pick error không?",
        [
            ("Tổng đơn", "Tổng pick task trong kỳ", "task"),
            ("Số lượng", "Tổng pick lines", "line"),
            ("Số lượng", "Tổng CSE đã pick", "CSE"),
            ("Trạng thái", "Pick task hoàn tất", "task"),
            ("Trạng thái", "Pick task có error / re-pick", "task"),
            ("KPI", "% pick accuracy", "%"),
            ("KPI", "% pick hoàn thành đúng giờ", "%"),
        ],
    ),
    (
        "Shipping Progress",
        "Section Shipping Progress — Tiến độ vận chuyển",
        "Đơn đang ở giai đoạn nào của hành trình giao hàng?",
        [
            ("Tổng đơn", "Tổng đơn ship trong kỳ", "đơn"),
            ("Số lượng", "Tổng CSE đang vận chuyển", "CSE"),
            ("Trạng thái", "Đơn đã giao (Delivered)", "đơn"),
            ("Trạng thái", "Đơn đang trên đường (In-transit)", "đơn"),
            ("Trạng thái", "Đơn chờ xuất kho (Pending dispatch)", "đơn"),
            ("Trạng thái", "Đơn failed / trả về", "đơn"),
            ("KPI", "% on-time delivery", "%"),
            ("KPI", "% đơn còn lại trong kỳ", "%"),
        ],
    ),
    (
        "Stock Type",
        "Section Stock Type — Phân loại tồn kho",
        "Cơ cấu tồn kho theo loại (available / allocated / damaged) ra sao?",
        [
            ("Tổng đơn", "Tổng SKU đang theo dõi", "SKU"),
            ("Số lượng", "Tổng CSE tồn", "CSE"),
            ("Số lượng", "Tổng pallet tồn", "pallet"),
            ("Trạng thái", "Stock available (sẵn bán)", "CSE"),
            ("Trạng thái", "Stock allocated (đã chốt cho đơn)", "CSE"),
            ("Trạng thái", "Stock damaged / hold", "CSE"),
            ("KPI", "% stock available trên tổng", "%"),
            ("KPI", "% stock damaged trên tổng", "%"),
        ],
    ),
    (
        "Transfer",
        "Section Transfer — Điều chuyển nội bộ",
        "Có bao nhiêu lệnh điều chuyển giữa kho, hoàn thành bao nhiêu?",
        [
            ("Tổng đơn", "Tổng transfer order trong kỳ", "TO"),
            ("Số lượng", "Tổng CSE kế hoạch điều chuyển", "CSE"),
            ("Số lượng", "Tổng CSE đã điều chuyển", "CSE"),
            ("Trạng thái", "TO hoàn tất", "TO"),
            ("Trạng thái", "TO đang điều chuyển", "TO"),
            ("Trạng thái", "TO pending / chưa bắt đầu", "TO"),
            ("KPI", "% TO hoàn thành", "%"),
            ("KPI", "% CSE đạt kế hoạch", "%"),
        ],
    ),
    (
        "Txn Move",
        "Section Txn Move — Giao dịch dịch chuyển trong kho",
        "Tổng giao dịch dịch chuyển nội kho (putaway/replenish/relocation), thành công bao nhiêu?",
        [
            ("Tổng đơn", "Tổng giao dịch move trong kỳ", "txn"),
            ("Số lượng", "Tổng CSE liên quan", "CSE"),
            ("Trạng thái", "Txn thành công", "txn"),
            ("Trạng thái", "Txn pending", "txn"),
            ("Trạng thái", "Txn error / rollback", "txn"),
            ("KPI", "% txn thành công", "%"),
            ("KPI", "% txn error", "%"),
        ],
    ),
    (
        "WH Utilization",
        "Section WH Utilization — Độ sử dụng kho",
        "Các kho đang dùng bao nhiêu % công suất? Kho nào full, kho nào trống?",
        [
            ("Tổng đơn", "Tổng số kho theo dõi", "kho"),
            ("Số lượng", "Tổng pallet location khả dụng", "location"),
            ("Số lượng", "Tổng pallet location đang dùng", "location"),
            ("Trạng thái", "Kho utilization > 90%", "kho"),
            ("Trạng thái", "Kho utilization 50–90%", "kho"),
            ("Trạng thái", "Kho utilization < 50%", "kho"),
            ("KPI", "% utilization trung bình toàn mạng", "%"),
            ("KPI", "% location available", "%"),
        ],
    ),
]


def _set_col_widths(ws: Worksheet, widths: list[int]) -> None:
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _title_block(ws: Worksheet, title: str, subtitle: str | None, top_row: int = 1) -> int:
    ws.cell(row=top_row, column=1, value=title).font = header_font(WHITE, True, 14)
    ws.cell(row=top_row, column=1).fill = fill(NAVY)
    ws.cell(row=top_row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[top_row].height = 26
    ws.merge_cells(start_row=top_row, start_column=1, end_row=top_row, end_column=8)
    if subtitle:
        ws.cell(row=top_row + 1, column=1, value=subtitle).font = body_font(False)
        ws.cell(row=top_row + 1, column=1).fill = fill(PALE)
        ws.cell(row=top_row + 1, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[top_row + 1].height = 22
        ws.merge_cells(start_row=top_row + 1, start_column=1, end_row=top_row + 1, end_column=8)
        return top_row + 2
    return top_row + 1


def _info_row(ws: Worksheet, row: int, label: str, value: str) -> None:
    ws.cell(row=row, column=1, value=label).font = body_font(True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=row, column=2, value=value).font = body_font(False)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("00 — README")
    _set_col_widths(ws, [22, 18, 18, 14, 14, 14, 16, 24])

    row = _title_block(
        ws,
        f"UAT Quick Check — {TENANT}",
        f"Đối chiếu nhanh giữa Dashboard Control Tower và số liệu khách đang theo dõi",
    )
    _info_row(ws, row, "Thời gian (window):", f"{WINDOW_FROM:%Y-%m-%d}  →  {WINDOW_TO:%Y-%m-%d}  (UTC)")
    row += 1
    _info_row(ws, row, "Tenant:", TENANT)
    row += 1
    _info_row(ws, row, "Số view (sheet):", str(len(SECTIONS)))
    row += 1
    _info_row(ws, row, "Người tạo:", "Smartlog PM/BA")
    row += 2

    ws.cell(row=row, column=1, value="Hướng dẫn sử dụng").font = header_font(DARK, True, 12)
    row += 1
    instructions = [
        "1. Mở mỗi sheet view, điền số vào cột [Dashboard] và [WMS/TMS].",
        "2. Cột [Dashboard] = số đọc trên màn hình Control Tower (sau khi áp đúng filter + time window).",
        "3. Cột [WMS/TMS] = số từ HỆ THỐNG GỐC — TMS (vận tải) cho OTIF / Flash Daily / VFR / Shipping Progress; WMS (kho) cho Factory Inbound / Loose Picking / Stock Type / Transfer / Txn Move / WH Utilization.",
        "4. Cột [Diff] = Dashboard − WMS/TMS. PM/khách tự tính bằng tay (file để trống công thức).",
        "5. Cột [Status]: ghi 'OK' nếu khớp / trong dung sai; 'Lệch' nếu vượt dung sai; 'Chưa rõ' nếu cần kiểm tra thêm.",
        "6. Cột [Ghi chú]: nếu Lệch — ghi giả thuyết nguyên nhân (vd 'timezone', 'filter khác', 'cutoff giờ', 'WMS/TMS chưa đồng bộ').",
        "7. Sau khi điền xong tất cả sheet, mở sheet 'Tổng hợp' để xem bức tranh chung.",
    ]
    for line in instructions:
        ws.cell(row=row, column=1, value=line).font = body_font(False)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Quy ước dung sai (tolerance) đề xuất").font = header_font(DARK, True, 12)
    row += 1
    tols = [
        ("Số đếm tuyệt đối (đơn, CSE, pallet, ...)", "≤ 1%"),
        ("% metric (OTIF%, Ontime%, Infull%, ...)", "≤ 0.5 điểm %"),
        ("Tổng amount/tải trọng/CSE", "≤ 0.5%"),
        ("Ranking Top N (top kho late, top customer)", "≥ 4/5 tên trùng"),
    ]
    for label, val in tols:
        ws.cell(row=row, column=1, value=label).font = body_font(False)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=row, column=1).fill = fill(PALE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=5, value=val).font = body_font(True)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5).fill = fill(PALE)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Danh sách sheet").font = header_font(DARK, True, 12)
    row += 1
    headers = ["#", "Sheet", "View / Section", "Câu hỏi nghiệp vụ chính"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font(WHITE, True, 11)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    row += 1
    sheet_index = [("Tổng hợp", "Tất cả view", "Bức tranh tổng hợp — OK / Lệch / Chưa rõ cho từng view")]
    for sn, vt, mq, _ in SECTIONS:
        sheet_index.append((sn, vt, mq))
    for i, (sn, vt, mq) in enumerate(sheet_index, start=1):
        ws.cell(row=row, column=1, value=i).font = body_font(False)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=sn).font = body_font(True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=3, value=vt).font = body_font(False)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=row, column=3).border = BORDER
        ws.cell(row=row, column=4, value=mq).font = body_font(False)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="left", indent=1, wrap_text=True)
        ws.cell(row=row, column=4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 22
        row += 1


def build_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("01 — Tổng hợp")
    _set_col_widths(ws, [4, 22, 38, 14, 14, 14, 12, 28])

    row = _title_block(
        ws,
        f"Tổng hợp UAT Quick Check — {WINDOW_FROM:%Y-%m-%d} → {WINDOW_TO:%Y-%m-%d}",
        "1 dòng / view — chỉ tiêu KPI chính + Tổng đơn. Số liệu copy từ sheet con sau khi điền xong.",
    )

    headers = ["#", "View", "KPI chính (đại diện)", "Dashboard", "WMS/TMS", "Diff", "Status", "Ghi chú"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font(WHITE, True, 11)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 32
    row += 1

    primary_kpi = {
        "OTIF": "% OTIF (target 90%)",
        "Flash Daily": "% hoàn thành theo CSE",
        "VFR": "% chuyến đạt VFR",
        "Factory Inbound": "% PO hoàn tất",
        "Loose Picking": "% pick accuracy",
        "Shipping Progress": "% on-time delivery",
        "Stock Type": "% stock available trên tổng",
        "Transfer": "% TO hoàn thành",
        "Txn Move": "% txn thành công",
        "WH Utilization": "% utilization trung bình toàn mạng",
    }

    section_block_start = row
    for i, (sn, _vt, _mq, _rows) in enumerate(SECTIONS, start=1):
        kpi = primary_kpi.get(sn, "—")
        values = [i, sn, kpi, None, None, None, None, None]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER
            c.font = body_font(False)
            if col == 1:
                c.alignment = Alignment(horizontal="center")
            elif col in (4, 5, 6):
                c.alignment = Alignment(horizontal="right")
                c.fill = fill(YELLOW)
            elif col == 7:
                c.alignment = Alignment(horizontal="center")
                c.fill = fill(YELLOW)
            elif col == 8:
                c.alignment = Alignment(horizontal="left", indent=1, wrap_text=True)
                c.fill = fill(YELLOW)
            else:
                c.alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Tóm tắt:").font = header_font(DARK, True, 12)
    row += 1
    summary_lines = [
        ("Tổng số view check:", str(len(SECTIONS))),
        ("Số view OK (khớp / trong dung sai):", ""),
        ("Số view Lệch:", ""),
        ("Số view Chưa rõ / cần check lại:", ""),
        ("Kết luận tổng:", ""),
    ]
    for label, val in summary_lines:
        ws.cell(row=row, column=2, value=label).font = body_font(True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", indent=1)
        ws.cell(row=row, column=3, value=val).font = body_font(False)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=row, column=3).fill = fill(YELLOW)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        row += 1

    ws.freeze_panes = ws.cell(row=section_block_start, column=1)


def build_section_sheet(wb: Workbook, idx: int, sn: str, vt: str, mq: str, rows: list[tuple[str, str, str]]) -> None:
    sheet_name = f"{idx + 1:02d} — {sn}"[:31]
    ws = wb.create_sheet(sheet_name)
    _set_col_widths(ws, [4, 14, 42, 14, 14, 14, 12, 28])

    row = _title_block(ws, f"{vt}", f"Câu hỏi nghiệp vụ: {mq}")
    _info_row(ws, row, "Nguồn gốc đối chiếu:", SOURCE_SYSTEM.get(sn, "WMS/TMS"))
    row += 1
    _info_row(ws, row, "Thời gian (window):", f"{WINDOW_FROM:%Y-%m-%d}  →  {WINDOW_TO:%Y-%m-%d}  (UTC)")
    row += 1
    _info_row(ws, row, "Filter mặc định:", "ALL (NVC, kho, customer, cargo). PM/khách ghi rõ filter nếu áp khác.")
    row += 2

    headers = ["#", "Nhóm", "Chỉ tiêu", "Dashboard", "WMS/TMS", "Diff (D−W)", "Status", "Ghi chú nếu lệch"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font(WHITE, True, 11)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 36
    header_row = row
    row += 1

    for i, (group, metric, unit) in enumerate(rows, start=1):
        values = [i, group, f"{metric}  ({unit})", None, None, None, None, None]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER
            c.font = body_font(False)
            if col == 1:
                c.alignment = Alignment(horizontal="center")
            elif col == 2:
                c.alignment = Alignment(horizontal="left", indent=1)
                c.fill = fill(GROUP_COLOR.get(group, PALE))
                c.font = body_font(True)
            elif col == 3:
                c.alignment = Alignment(horizontal="left", indent=1, wrap_text=True)
            elif col in (4, 5, 6):
                c.alignment = Alignment(horizontal="right")
                c.fill = fill(YELLOW)
            elif col == 7:
                c.alignment = Alignment(horizontal="center")
                c.fill = fill(YELLOW)
            elif col == 8:
                c.alignment = Alignment(horizontal="left", indent=1, wrap_text=True)
                c.fill = fill(YELLOW)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="Verdict view này:").font = body_font(True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value="").fill = fill(YELLOW)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=row, column=3).font = body_font(True)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    row += 1
    ws.cell(row=row, column=2, value="Ghi chú PM/BA:").font = body_font(True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value="").fill = fill(YELLOW)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", indent=1, wrap_text=True)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 44

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:H{header_row + len(rows)}"


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb)
    build_summary(wb)
    for idx, (sn, vt, mq, rows) in enumerate(SECTIONS, start=1):
        build_section_sheet(wb, idx, sn, vt, mq, rows)
    wb.save(OUT_FILE)
    print(f"WROTE  {OUT_FILE}")
    print(f"SHEETS {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"  - {name.encode('ascii', 'replace').decode('ascii')}")


if __name__ == "__main__":
    main()
